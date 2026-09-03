#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AnyLine —— 在线事务管理网站 (Flask + SQLite)"""
import json
import os
import re
import sqlite3
import base64
import binascii
from zipfile import BadZipFile
from io import BytesIO
from datetime import date, datetime, timezone

from flask import (
    Flask, g, jsonify, request, send_file, send_from_directory, session,
)
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps, UnidentifiedImageError
from werkzeug.exceptions import HTTPException
from werkzeug.security import check_password_hash, generate_password_hash

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "anyline.db")
DEFAULT_STATUS_ENUM = ["未启动", "进行中", "有风险", "等待中", "已暂停", "已闭环", "已取消"]
DEFAULT_STATUS_COLORS = {
    "未启动": "#8c959f",
    "进行中": "#0969da",
    "有风险": "#d4a72c",
    "等待中": "#0e7490",
    "已暂停": "#8250df",
    "已闭环": "#1a7f37",
    "已取消": "#57606a",
}
FALLBACK_STATUS_COLOR = "#6e7781"
PRIORITY_ENUM = ["低", "中", "高", "紧急"]
TASK_IMAGE_TYPES = {"image/png", "image/jpeg", "image/gif", "image/webp"}
MAX_TASK_IMAGES = 8
MAX_TASK_IMAGE_BYTES = 5 * 1024 * 1024
MAX_TASK_IMAGES_BYTES = 20 * 1024 * 1024
AVATAR_IMAGE_TYPES = {"image/png", "image/jpeg", "image/webp"}
MAX_AVATAR_SOURCE_BYTES = 5 * 1024 * 1024
MAX_AVATAR_SOURCE_PIXELS = 25_000_000
AVATAR_SIZE = 256
MAX_AVATAR_BYTES = 1024 * 1024
MAX_TASK_ATTACHMENTS = 8
MAX_TASK_ATTACHMENT_BYTES = 5 * 1024 * 1024
MAX_TASK_ATTACHMENTS_BYTES = 20 * 1024 * 1024
MAX_ATTACHMENT_FILENAME_LENGTH = 180
MAX_TASK_IMPORT_BYTES = 5 * 1024 * 1024
MAX_TASK_IMPORT_ROWS = 2000
MAX_LINE_IMPORT_ROWS = 2000
LINE_EXPORT_COLUMNS = (
    ("线ID", "id", 12),
    ("线类型", "type", 12),
    ("线路径", "path", 36),
    ("父线ID", "parent_id", 12),
    ("父线路径", "parent_path", 36),
    ("线名", "name", 24),
    ("描述", "description", 36),
    ("颜色", "color", 14),
    ("起始日期", "fork_date", 14),
    ("反合日期", "merge_date", 14),
    ("更新日期", "updated_at", 14),
)
LINE_IMPORT_COLUMNS = (
    ("线标识", "import_key", 16, True),
    ("父线标识", "parent_key", 16, False),
    ("线名", "name", 24, True),
    ("描述", "description", 36, False),
    ("颜色", "color", 14, False),
    ("起始日期", "fork_date", 14, True),
    ("反合日期", "merge_date", 14, False),
)
DATA_TASK_EXPORT_COLUMNS = (
    ("事务ID", "id", 12),
    ("所属线ID", "line_id", 14),
    ("所属线路径", "line_path", 36),
    ("事务名", "name", 24),
    ("事务内容", "content", 36),
    ("闭环目标", "goal", 28),
    ("下一步动作", "next_action", 28),
    ("风险原因", "risk_reason", 28),
    ("优先级", "priority", 12),
    ("责任人", "owner", 16),
    ("进展状态", "status", 14),
    ("起始日期", "start_date", 14),
    ("结束日期", "end_date", 14),
    ("状态起始日期", "status_since", 16),
    ("更新日期", "updated_at", 14),
)
TASK_EXPORT_COLUMNS = (
    ("事务ID", "id", 12),
    ("线名", "line_name", 20),
    ("线类型", "line_type", 12),
    ("父线", "parent_name", 20),
    ("事务名", "name", 24),
    ("事务内容", "content", 36),
    ("闭环目标", "goal", 28),
    ("下一步动作", "next_action", 28),
    ("风险原因", "risk_reason", 28),
    ("优先级", "priority", 12),
    ("责任人", "owner", 16),
    ("进展状态", "status", 14),
    ("起始日期", "start_date", 14),
    ("结束日期", "end_date", 14),
    ("状态起始日期", "status_since", 16),
    ("更新日期", "updated_at", 14),
)
TASK_IMPORT_COLUMNS = (
    ("所属线ID", "line_id", 14, False),
    ("所属线路径", "line_path", 32, False),
    ("事务名", "name", 24, True),
    ("事务内容", "content", 36, True),
    ("闭环目标", "goal", 28, False),
    ("下一步动作", "next_action", 28, False),
    ("风险原因", "risk_reason", 28, False),
    ("优先级", "priority", 12, False),
    ("责任人", "owner", 16, True),
    ("进展状态", "status", 14, True),
    ("起始日期", "start_date", 14, True),
    ("结束日期", "end_date", 14, True),
)

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.config.update(
    DATABASE=os.environ.get("ANYLINE_DB_PATH", DB_PATH),
    MAX_CONTENT_LENGTH=64 * 1024 * 1024,
    SECRET_KEY=os.environ.get(
        "ANYLINE_SECRET_KEY", "anyline-local-secret-change-in-production"
    ),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)


# ---------------------------------------------------------------- db helpers
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(app.config["DATABASE"])
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db(db_path=None):
    db = sqlite3.connect(db_path or app.config["DATABASE"])
    db.row_factory = sqlite3.Row
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS lines (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL,
            description TEXT DEFAULT '',
            color      TEXT,
            parent_id  INTEGER,                 -- NULL = 主线
            fork_date  TEXT NOT NULL,           -- 线的起点(支线=分叉日)
            merge_date TEXT,                    -- 反合回父线的日期, NULL=未反合
            deleted    INTEGER NOT NULL DEFAULT 0,
            del_batch  INTEGER,
            deleted_at TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS tasks (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            line_id      INTEGER NOT NULL,
            name         TEXT NOT NULL,
            content      TEXT DEFAULT '',
            goal         TEXT DEFAULT '',
            owner        TEXT DEFAULT '',
            priority     TEXT NOT NULL DEFAULT '中',
            next_action  TEXT DEFAULT '',
            risk_reason  TEXT DEFAULT '',
            status       TEXT NOT NULL DEFAULT '未启动',
            start_date   TEXT NOT NULL,
            end_date     TEXT,
            status_since TEXT NOT NULL,         -- 当前进展状态的开始日, 用于计算停留时长
            deleted      INTEGER NOT NULL DEFAULT 0,
            del_batch    INTEGER,
            deleted_at   TEXT,
            updated_at   TEXT
        );
        CREATE TABLE IF NOT EXISTS task_dependencies (
            workspace_id        INTEGER NOT NULL,
            dependent_task_id   INTEGER NOT NULL,
            prerequisite_task_id INTEGER NOT NULL,
            PRIMARY KEY(workspace_id,dependent_task_id,prerequisite_task_id),
            CHECK(dependent_task_id <> prerequisite_task_id)
        );
        CREATE TABLE IF NOT EXISTS milestones (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id       INTEGER NOT NULL,
            line_id            INTEGER NOT NULL,
            name               TEXT NOT NULL,
            target_description TEXT NOT NULL DEFAULT '',
            milestone_date     TEXT NOT NULL,
            deleted            INTEGER NOT NULL DEFAULT 0,
            del_batch          INTEGER,
            deleted_at         TEXT,
            created_at         TEXT NOT NULL,
            updated_at         TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS milestone_tasks (
            workspace_id INTEGER NOT NULL,
            milestone_id INTEGER NOT NULL,
            task_id      INTEGER NOT NULL,
            PRIMARY KEY(workspace_id,milestone_id,task_id)
        );
        CREATE TABLE IF NOT EXISTS task_images (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id INTEGER NOT NULL,
            task_id      INTEGER NOT NULL,
            mime_type    TEXT NOT NULL,
            data         BLOB NOT NULL,
            created_at   TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS task_attachments (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id INTEGER NOT NULL,
            task_id      INTEGER NOT NULL,
            filename     TEXT NOT NULL,
            mime_type    TEXT NOT NULL,
            data         BLOB NOT NULL,
            created_at   TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS meta (
            key   TEXT PRIMARY KEY,
            value TEXT
        );
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT NOT NULL COLLATE NOCASE UNIQUE,
            display_name  TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            managed_by    INTEGER,
            active        INTEGER NOT NULL DEFAULT 1,
            created_at    TEXT NOT NULL,
            updated_at    TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS workspaces (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            created_by  INTEGER NOT NULL,
            created_at  TEXT NOT NULL,
            updated_at  TEXT NOT NULL,
            archived_at TEXT
        );
        CREATE TABLE IF NOT EXISTS workspace_members (
            workspace_id INTEGER NOT NULL,
            user_id      INTEGER NOT NULL,
            role         TEXT NOT NULL CHECK(role IN ('admin','member')),
            joined_at    TEXT NOT NULL,
            PRIMARY KEY(workspace_id,user_id)
        );
        CREATE TABLE IF NOT EXISTS workspace_meta (
            workspace_id INTEGER NOT NULL,
            key          TEXT NOT NULL,
            value        TEXT,
            PRIMARY KEY(workspace_id,key)
        );
        CREATE TABLE IF NOT EXISTS undo_snapshots (
            workspace_id INTEGER PRIMARY KEY,
            snapshot     TEXT NOT NULL,
            created_at   TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS redo_snapshots (
            workspace_id INTEGER PRIMARY KEY,
            snapshot     TEXT NOT NULL,
            created_at   TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS dashboard_snapshots (
            workspace_id  INTEGER NOT NULL,
            snapshot_date TEXT NOT NULL,
            total         INTEGER NOT NULL,
            done          INTEGER NOT NULL,
            overdue       INTEGER NOT NULL,
            risk          INTEGER NOT NULL,
            blocked       INTEGER NOT NULL,
            status_counts TEXT NOT NULL,
            PRIMARY KEY(workspace_id,snapshot_date)
        );
        CREATE TABLE IF NOT EXISTS task_followers (
            workspace_id INTEGER NOT NULL,
            task_id      INTEGER NOT NULL,
            user_id      INTEGER NOT NULL,
            created_at   TEXT NOT NULL,
            PRIMARY KEY(workspace_id,task_id,user_id)
        );
        CREATE TABLE IF NOT EXISTS task_comments (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id INTEGER NOT NULL,
            task_id      INTEGER NOT NULL,
            author_id    INTEGER NOT NULL,
            content      TEXT NOT NULL,
            created_at   TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS task_activities (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id INTEGER NOT NULL,
            task_id      INTEGER NOT NULL,
            actor_id     INTEGER,
            event_type   TEXT NOT NULL,
            summary      TEXT NOT NULL,
            metadata     TEXT NOT NULL DEFAULT '{}',
            created_at   TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS notifications (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id INTEGER NOT NULL,
            user_id      INTEGER NOT NULL,
            task_id      INTEGER,
            actor_id     INTEGER,
            kind         TEXT NOT NULL,
            message      TEXT NOT NULL,
            dedupe_key   TEXT,
            read_at      TEXT,
            created_at   TEXT NOT NULL
        );
        """
    )
    ensure_column(db, "lines", "description", "TEXT DEFAULT ''")
    ensure_column(db, "lines", "color", "TEXT")
    ensure_column(db, "lines", "deleted_at", "TEXT")
    ensure_column(db, "lines", "updated_at", "TEXT")
    ensure_column(db, "lines", "workspace_id", "INTEGER")
    ensure_column(db, "tasks", "priority", "TEXT NOT NULL DEFAULT '中'")
    ensure_column(db, "tasks", "next_action", "TEXT DEFAULT ''")
    ensure_column(db, "tasks", "risk_reason", "TEXT DEFAULT ''")
    ensure_column(db, "tasks", "deleted_at", "TEXT")
    ensure_column(db, "tasks", "updated_at", "TEXT")
    ensure_column(db, "tasks", "workspace_id", "INTEGER")
    ensure_column(db, "users", "managed_by", "INTEGER")
    ensure_column(db, "users", "avatar_mime", "TEXT")
    ensure_column(db, "users", "avatar_data", "BLOB")
    ensure_column(db, "users", "avatar_updated_at", "TEXT")
    ensure_column(db, "workspaces", "archived_at", "TEXT")
    today = date.today().isoformat()

    admin_username = os.environ.get("ANYLINE_ADMIN_USERNAME", "admin").strip() or "admin"
    admin_password = os.environ.get("ANYLINE_ADMIN_PASSWORD", "admin123")
    admin = db.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()
    if not admin:
        cur = db.execute(
            "INSERT INTO users(username,display_name,password_hash,created_at,updated_at) "
            "VALUES(?,?,?,?,?)",
            (
                admin_username, "系统管理员", generate_password_hash(admin_password),
                today, today,
            ),
        )
        admin_id = cur.lastrowid
    else:
        admin_id = admin["id"]
    db.execute(
        "UPDATE users SET managed_by=id WHERE managed_by IS NULL AND id=?", (admin_id,)
    )

    workspace = db.execute("SELECT id FROM workspaces ORDER BY id LIMIT 1").fetchone()
    if not workspace:
        cur = db.execute(
            "INSERT INTO workspaces(name,description,created_by,created_at,updated_at) "
            "VALUES(?,?,?,?,?)",
            ("默认项目", "由原有 AnyLine 数据自动迁移", admin_id, today, today),
        )
        workspace_id = cur.lastrowid
    else:
        workspace_id = workspace["id"]
    db.execute(
        "INSERT OR IGNORE INTO workspace_members(workspace_id,user_id,role,joined_at) "
        "VALUES(?,?,?,?)", (workspace_id, admin_id, "admin", today),
    )
    db.execute(
        "UPDATE lines SET workspace_id=? WHERE workspace_id IS NULL", (workspace_id,)
    )
    db.execute(
        "UPDATE tasks SET workspace_id=(SELECT workspace_id FROM lines "
        "WHERE lines.id=tasks.line_id) WHERE workspace_id IS NULL"
    )
    db.execute(
        "UPDATE tasks SET workspace_id=? WHERE workspace_id IS NULL", (workspace_id,)
    )
    db.execute(
        "INSERT OR IGNORE INTO workspace_meta(workspace_id,key,value) "
        "SELECT ?,key,value FROM meta", (workspace_id,)
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_lines_workspace_deleted "
        "ON lines(workspace_id,deleted)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_tasks_workspace_deleted "
        "ON tasks(workspace_id,deleted)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_task_dependencies_prerequisite "
        "ON task_dependencies(workspace_id,prerequisite_task_id)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_milestones_workspace_line "
        "ON milestones(workspace_id,line_id,deleted,milestone_date)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_milestone_tasks_task "
        "ON milestone_tasks(workspace_id,task_id,milestone_id)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_task_images_task "
        "ON task_images(workspace_id,task_id)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_task_attachments_task "
        "ON task_attachments(workspace_id,task_id)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_members_user ON workspace_members(user_id)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_dashboard_snapshots_date "
        "ON dashboard_snapshots(workspace_id,snapshot_date)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_task_followers_user "
        "ON task_followers(workspace_id,user_id,task_id)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_task_comments_task "
        "ON task_comments(workspace_id,task_id,created_at,id)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_task_activities_task "
        "ON task_activities(workspace_id,task_id,created_at,id)"
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_notifications_user "
        "ON notifications(workspace_id,user_id,read_at,created_at,id)"
    )
    db.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_notifications_dedupe "
        "ON notifications(workspace_id,user_id,dedupe_key) "
        "WHERE dedupe_key IS NOT NULL"
    )
    db.execute("UPDATE lines SET updated_at=? WHERE updated_at IS NULL", (today,))
    db.execute("UPDATE tasks SET updated_at=? WHERE updated_at IS NULL", (today,))
    db.commit()
    db.close()


def ensure_column(db, table, column, ddl):
    cols = {r["name"] for r in db.execute(f"PRAGMA table_info({table})")}
    if column not in cols:
        db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {ddl}")


class ApiError(ValueError):
    def __init__(self, message, status=400):
        super().__init__(message)
        self.status = status


def json_object():
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        raise ApiError("请求体必须是 JSON 对象")
    return data


def required_id(value, field):
    if not isinstance(value, int) or isinstance(value, bool):
        raise ApiError(f"{field} 必须是整数")
    return value


def text_field(data, key, label, default="", nullable=False):
    value = data.get(key, default)
    if nullable and value is None:
        return None
    if not isinstance(value, str):
        raise ApiError(f"{label} 必须是字符串")
    return value


def line_color(value):
    if value is None or value == "":
        return None
    if not isinstance(value, str) or len(value) != 7 or value[0] != "#" or \
            any(char not in "0123456789abcdefABCDEF" for char in value[1:]):
        raise ApiError("颜色必须是 #RRGGBB 格式")
    return value.lower()


def append_excel_value(cell, value, is_date=False):
    if is_date and value:
        try:
            cell.value = date.fromisoformat(value)
            cell.number_format = "yyyy-mm-dd"
            return
        except ValueError:
            pass
    cell.value = "" if value is None else value
    if isinstance(cell.value, str):
        # Prevent user-entered values from being evaluated as Excel formulas.
        cell.data_type = "s"


def line_export_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "线导入"
    sheet.append([column[0] for column in LINE_EXPORT_COLUMNS])
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append([None] * len(LINE_EXPORT_COLUMNS))
        for column_index, (_label, key, _width) in enumerate(LINE_EXPORT_COLUMNS, 1):
            append_excel_value(
                sheet.cell(sheet.max_row, column_index), row.get(key),
                key in {"fork_date", "merge_date", "updated_at"},
            )
    last_column = get_column_letter(len(LINE_EXPORT_COLUMNS))
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 22
    for index, (_label, _key, width) in enumerate(LINE_EXPORT_COLUMNS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    instructions = workbook.create_sheet("填报说明")
    for row in (
        ("项目", "说明"),
        ("导入方式", "该导出文件可直接作为线导入文件使用，也可按需修改后导入。"),
        ("线ID", "导入时作为文件内唯一的线标识；可使用数字或文本，导入后会生成新的系统 ID。"),
        ("父线ID", "主线留空；支线填写同一工作表内父线的线ID。行顺序不影响导入。"),
        ("只读列", "线类型、线路径、父线路径和更新日期仅供查看，导入时会自动忽略。"),
        ("导入规则", "任意一行校验失败时整批不导入；一次成功导入可使用 Ctrl+Z 整批撤销。"),
    ):
        instructions.append(row)
    instructions["A1"].font = instructions["B1"].font = Font(bold=True)
    instructions["A1"].fill = instructions["B1"].fill = header_fill
    instructions.column_dimensions["A"].width = 16
    instructions.column_dimensions["B"].width = 88
    for row in instructions.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def line_import_template_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "线导入"
    sheet.append([column[0] for column in LINE_IMPORT_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(LINE_IMPORT_COLUMNS))}1"
    sheet.row_dimensions[1].height = 24
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    required_fill = PatternFill("solid", fgColor="FCE8E6")
    for index, (label, _key, width, required) in enumerate(LINE_IMPORT_COLUMNS, 1):
        cell = sheet.cell(1, index)
        cell.font = Font(bold=True, color="9C0006" if required else "1F2328")
        cell.fill = required_fill if required else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment("必填字段", "AnyLine") if required else None
        sheet.column_dimensions[get_column_letter(index)].width = width

    instructions = workbook.create_sheet("填报说明")
    for row in (
        ("项目", "说明"),
        ("导入规则", "从第 2 行开始填写；空白行会自动忽略。任意一行校验失败时整批不导入。"),
        ("线标识", "文件内必须唯一，可使用便于识别的数字或文本，仅用于关联父线。"),
        ("主线", "父线标识留空。"),
        ("支线", "父线标识填写同一工作表中父线的线标识；父线可以写在支线之后。"),
        ("日期格式", "使用 YYYY-MM-DD；支线不能早于父线，反合日期不能早于支线起始日期。"),
        ("颜色", "可留空，或填写 #RRGGBB 格式的颜色。"),
        ("导入范围", f"单次最多 {MAX_LINE_IMPORT_ROWS} 条主线与支线，仅创建新线。"),
        ("撤销", "一次导入作为一次编辑，可在画布视图使用 Ctrl+Z 整批撤销。"),
    ):
        instructions.append(row)
    instructions["A1"].font = instructions["B1"].font = Font(bold=True)
    instructions["A1"].fill = instructions["B1"].fill = header_fill
    instructions.column_dimensions["A"].width = 16
    instructions.column_dimensions["B"].width = 88
    for row in instructions.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def task_export_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事务"
    headers = [column[0] for column in TASK_EXPORT_COLUMNS]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    date_keys = {"start_date", "end_date", "status_since", "updated_at"}
    for row in rows:
        sheet.append([None] * len(TASK_EXPORT_COLUMNS))
        row_index = sheet.max_row
        for column_index, (_, key, _) in enumerate(TASK_EXPORT_COLUMNS, 1):
            append_excel_value(
                sheet.cell(row_index, column_index), row[key], key in date_keys
            )

    last_column = get_column_letter(len(TASK_EXPORT_COLUMNS))
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 22
    for index, (_, _, width) in enumerate(TASK_EXPORT_COLUMNS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def workspace_line_records(db, workspace_id):
    rows = [dict(row) for row in db.execute(
        "SELECT id,name,description,color,parent_id,fork_date,merge_date,updated_at FROM lines "
        "WHERE workspace_id=? AND deleted=0 ORDER BY id", (workspace_id,)
    )]
    by_id = {row["id"]: row for row in rows}
    paths = {}

    def line_path(line_id, visiting=None):
        if line_id in paths:
            return paths[line_id]
        row = by_id[line_id]
        visiting = set() if visiting is None else visiting
        if line_id in visiting or row["parent_id"] not in by_id:
            result = row["name"]
        elif row["parent_id"] is None:
            result = row["name"]
        else:
            result = f"{line_path(row['parent_id'], visiting | {line_id})} / {row['name']}"
        paths[line_id] = result
        return result

    for row in rows:
        row["path"] = line_path(row["id"])
        row["type"] = "主线" if row["parent_id"] is None else "支线"
        row["parent_path"] = paths.get(row["parent_id"], "")
    return rows


def data_export_workbook(line_rows, task_rows):
    line_output = line_export_workbook(line_rows)
    workbook = load_workbook(line_output)
    sheet = workbook.create_sheet("事务导入", 1)
    sheet.append([column[0] for column in DATA_TASK_EXPORT_COLUMNS])
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    date_keys = {"start_date", "end_date", "status_since", "updated_at"}
    for row in task_rows:
        sheet.append([None] * len(DATA_TASK_EXPORT_COLUMNS))
        for index, (_label, key, _width) in enumerate(DATA_TASK_EXPORT_COLUMNS, 1):
            append_excel_value(
                sheet.cell(sheet.max_row, index), row.get(key), key in date_keys
            )
    last_column = get_column_letter(len(DATA_TASK_EXPORT_COLUMNS))
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 22
    for index, (_label, _key, width) in enumerate(DATA_TASK_EXPORT_COLUMNS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    instructions = workbook["填报说明"]
    instructions.append((
        "事务导入",
        "事务通过所属线ID或所属线路径关联“线导入”工作表中的线；事务ID、状态起始日期和更新日期仅供查看。",
    ))
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


def task_import_template_workbook(db, workspace_id):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事务导入"
    headers = [column[0] for column in TASK_IMPORT_COLUMNS]
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    sheet.row_dimensions[1].height = 24

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    required_fill = PatternFill("solid", fgColor="FCE8E6")
    for index, (label, _key, width, required) in enumerate(TASK_IMPORT_COLUMNS, 1):
        cell = sheet.cell(1, index)
        cell.font = Font(bold=True, color="9C0006" if required else "1F2328")
        cell.fill = required_fill if required else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width
        if label in {"所属线ID", "所属线路径"}:
            cell.comment = Comment("两列至少填写一项；同时填写时必须指向同一条线。", "AnyLine")
        elif required:
            cell.comment = Comment("必填字段", "AnyLine")

    instructions = workbook.create_sheet("填报说明")
    instruction_rows = [
        ("项目", "说明"),
        ("导入规则", "从第 2 行开始填写；空白行会自动忽略。任意一行校验失败时整批不导入。"),
        ("所属线", "所属线ID与所属线路径至少填写一项，建议从“项目数据”工作表复制。"),
        ("必填字段", "事务名、事务内容、责任人、进展状态、起始日期、结束日期。"),
        ("责任人", "必须从当前项目空间成员中选择；成员变更后请重新下载模板。"),
        ("日期格式", "使用 YYYY-MM-DD；结束日期不能早于起始日期。"),
        ("优先级", "留空时默认为“中”。"),
        ("导入范围", f"单次最多 {MAX_TASK_IMPORT_ROWS} 条事务，仅创建新事务，不导入图片和依赖关系。"),
        ("撤销", "一次导入作为一次编辑，可在画布视图使用 Ctrl+Z 整批撤销。"),
    ]
    for row in instruction_rows:
        instructions.append(row)
    instructions["A1"].font = instructions["B1"].font = Font(bold=True)
    instructions["A1"].fill = instructions["B1"].fill = header_fill
    instructions.column_dimensions["A"].width = 16
    instructions.column_dimensions["B"].width = 88
    for row in instructions.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    lines = workspace_line_records(db, workspace_id)
    project_data = workbook.create_sheet("项目数据")
    project_data.append(["线ID", "所属线路径", "线类型", "线起始日期"])
    for line in lines:
        project_data.append([line["id"], line["path"], line["type"], date.fromisoformat(line["fork_date"])])
        project_data.cell(project_data.max_row, 4).number_format = "yyyy-mm-dd"
    for cell in project_data[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    project_data.freeze_panes = "A2"
    project_data.auto_filter.ref = f"A1:D{max(1, project_data.max_row)}"
    for column, width in zip("ABCD", (12, 40, 12, 16)):
        project_data.column_dimensions[column].width = width

    options = workbook.create_sheet("选项")
    statuses = get_statuses(db)
    owners = get_workspace_member_names(db, workspace_id)
    options.append(["优先级", "进展状态", "责任人"])
    for index in range(max(len(PRIORITY_ENUM), len(statuses), len(owners))):
        options.append([
            PRIORITY_ENUM[index] if index < len(PRIORITY_ENUM) else None,
            statuses[index] if index < len(statuses) else None,
            owners[index] if index < len(owners) else None,
        ])
    key_columns = {key: get_column_letter(index) for index, (_, key, _, _) in enumerate(TASK_IMPORT_COLUMNS, 1)}
    validations = [
        ("priority", 1, len(PRIORITY_ENUM), "ImportPriorities"),
        ("status", 2, len(statuses), "ImportStatuses"),
        ("owner", 3, len(owners), "ImportOwners"),
    ]
    for key, option_column, count, range_name in validations:
        if not count:
            continue
        option_letter = get_column_letter(option_column)
        workbook.defined_names.add(DefinedName(
            range_name,
            attr_text=f"'选项'!${option_letter}$2:${option_letter}${count + 1}",
        ))
        validation = DataValidation(
            type="list",
            formula1=range_name,
            allow_blank=(key == "priority"),
        )
        validation.error = "请选择模板提供的有效选项"
        validation.errorTitle = "无效选项"
        validation.showErrorMessage = True
        sheet.add_data_validation(validation)
        column = key_columns[key]
        validation.add(f"{column}2:{column}{MAX_TASK_IMPORT_ROWS + 1}")
    options.sheet_state = "hidden"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def data_import_template_workbook(db, workspace_id):
    task_output = task_import_template_workbook(db, workspace_id)
    workbook = load_workbook(task_output)
    sheet = workbook.create_sheet("线导入", 0)
    sheet.append([column[0] for column in LINE_IMPORT_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(LINE_IMPORT_COLUMNS))}1"
    sheet.row_dimensions[1].height = 24
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    required_fill = PatternFill("solid", fgColor="FCE8E6")
    for index, (_label, _key, width, required) in enumerate(LINE_IMPORT_COLUMNS, 1):
        cell = sheet.cell(1, index)
        cell.font = Font(bold=True, color="9C0006" if required else "1F2328")
        cell.fill = required_fill if required else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment("必填字段", "AnyLine") if required else None
        sheet.column_dimensions[get_column_letter(index)].width = width
    instructions = workbook["填报说明"]
    line_instructions = (
        ("线导入", "“线导入”与“事务导入”可只填写一张，也可同时填写；任意数据校验失败时整批不导入。"),
        ("线标识", "文件内必须唯一；主线的父线标识留空，支线填写同一工作表内父线的线标识。"),
        ("事务关联新线", "事务的所属线ID可填写“线导入”工作表中的线标识，也可填写当前项目已有线ID或路径。"),
    )
    for row in line_instructions:
        instructions.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


def import_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return str(value).strip()


def import_cell_date(value, label):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = import_cell_text(value)
    if not text:
        raise ValueError(f"{label}不能为空")
    parse_iso_date(text, label)
    return text


def parse_line_import_sheet(sheet):
    if sheet.max_row > MAX_LINE_IMPORT_ROWS + 1:
        return [], [{
            "row": MAX_LINE_IMPORT_ROWS + 2,
            "message": f"单次最多导入 {MAX_LINE_IMPORT_ROWS} 条线，请删除多余行",
        }], MAX_LINE_IMPORT_ROWS + 1
    if sheet.max_column > 100:
        return [], [{"row": 1, "message": "导入工作表列数异常，请使用下载的导入模板"}], 1
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not header_cells:
        return [], [{"row": 1, "message": "工作表缺少表头"}], 1
    headers = [import_cell_text(cell.value) for cell in header_cells]
    nonempty_headers = [header for header in headers if header]
    duplicates = sorted({header for header in nonempty_headers if nonempty_headers.count(header) > 1})
    if duplicates:
        return [], [{"row": 1, "message": f"表头重复：{'、'.join(duplicates)}"}], 1
    header_indexes = {header: index for index, header in enumerate(headers) if header}
    aliases = {
        "线标识": "import_key", "线ID": "import_key",
        "父线标识": "parent_key", "父线ID": "parent_key",
        "线名": "name", "描述": "description", "颜色": "color",
        "起始日期": "fork_date", "反合日期": "merge_date",
    }
    missing = []
    for label, alternatives in (
        ("线标识", {"线标识", "线ID"}),
        ("线名", {"线名"}),
        ("起始日期", {"起始日期"}),
    ):
        if not alternatives & set(header_indexes):
            missing.append(label)
    if missing:
        return [], [{"row": 1, "message": f"缺少必需表头：{'、'.join(missing)}"}], 1

    parsed_rows = []
    errors = []
    seen_keys = {}
    data_count = 0

    def add_error(row_number, message):
        item = {"row": row_number, "message": message}
        if item not in errors:
            errors.append(item)

    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), 2):
        values = [cell.value for cell in cells]
        if not any(import_cell_text(value) for value in values):
            continue
        data_count += 1
        raw = {}
        for label, index in header_indexes.items():
            key = aliases.get(label)
            if key:
                raw[key] = values[index] if index < len(values) else None
        try:
            import_key = import_cell_text(raw.get("import_key"))
            parent_key = import_cell_text(raw.get("parent_key"))
            name = import_cell_text(raw.get("name"))
            if not import_key:
                raise ValueError("线标识不能为空")
            if import_key in seen_keys:
                raise ValueError(f"线标识与第 {seen_keys[import_key]} 行重复")
            seen_keys[import_key] = row_number
            if not name:
                raise ValueError("线名不能为空")
            color_text = import_cell_text(raw.get("color"))
            fork_date = import_cell_date(raw.get("fork_date"), "起始日期")
            merge_text = import_cell_text(raw.get("merge_date"))
            merge_date = import_cell_date(raw.get("merge_date"), "反合日期") if merge_text else None
            parsed_rows.append({
                "_row": row_number,
                "_key": import_key,
                "_parent_key": parent_key,
                "name": name,
                "description": import_cell_text(raw.get("description")),
                "color": line_color(color_text),
                "fork_date": fork_date,
                "merge_date": merge_date,
            })
        except ValueError as exc:
            add_error(row_number, str(exc))

    by_key = {row["_key"]: row for row in parsed_rows}
    for row in parsed_rows:
        parent_key = row["_parent_key"]
        if not parent_key:
            if row["merge_date"]:
                add_error(row["_row"], "主线不能填写反合日期")
            continue
        if parent_key == row["_key"]:
            add_error(row["_row"], "父线标识不能与线标识相同")
            continue
        parent = by_key.get(parent_key)
        if not parent:
            add_error(row["_row"], "父线标识不存在于当前导入文件")
            continue
        if row["fork_date"] < parent["fork_date"]:
            add_error(row["_row"], "支线起始日期不能早于父线起始日期")
        if row["merge_date"] and row["merge_date"] < row["fork_date"]:
            add_error(row["_row"], "反合日期不能早于支线起始日期")

    ordered = []
    visit_state = {}

    def visit(row, trail):
        state = visit_state.get(row["_key"], 0)
        if state == 2:
            return
        if state == 1:
            for key in trail[trail.index(row["_key"]):]:
                add_error(by_key[key]["_row"], "父线关系不能形成循环")
            return
        visit_state[row["_key"]] = 1
        parent = by_key.get(row["_parent_key"])
        if parent:
            visit(parent, trail + [row["_key"]])
        visit_state[row["_key"]] = 2
        if row not in ordered:
            ordered.append(row)

    for row in parsed_rows:
        visit(row, [])
    errors.sort(key=lambda item: item["row"])
    return ordered, errors, data_count


def imported_line_records(rows):
    by_key = {row["_key"]: row for row in rows}
    paths = {}

    def line_path(key):
        if key in paths:
            return paths[key]
        row = by_key[key]
        parent = by_key.get(row["_parent_key"])
        paths[key] = f"{line_path(parent['_key'])} / {row['name']}" if parent else row["name"]
        return paths[key]

    return [{
        "id": None,
        "import_key": row["_key"],
        "name": row["name"],
        "path": line_path(row["_key"]),
        "fork_date": row["fork_date"],
    } for row in rows]


def parse_task_import_sheet(sheet, db, workspace_id, imported_lines=None):
    if sheet.max_row > MAX_TASK_IMPORT_ROWS + 1:
        return [], [{
            "row": MAX_TASK_IMPORT_ROWS + 2,
            "message": f"单次最多导入 {MAX_TASK_IMPORT_ROWS} 条事务，请删除多余行",
        }], MAX_TASK_IMPORT_ROWS + 1
    if sheet.max_column > 100:
        return [], [{"row": 1, "message": "导入工作表列数异常，请使用下载的导入模板"}], 1
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not header_cells:
        return [], [{"row": 1, "message": "工作表缺少表头"}], 1
    headers = [import_cell_text(cell.value) for cell in header_cells]
    nonempty_headers = [header for header in headers if header]
    duplicates = sorted({header for header in nonempty_headers if nonempty_headers.count(header) > 1})
    if duplicates:
        return [], [{"row": 1, "message": f"表头重复：{'、'.join(duplicates)}"}], 1
    header_indexes = {header: index for index, header in enumerate(headers) if header}
    required = {label for label, _key, _width, is_required in TASK_IMPORT_COLUMNS if is_required}
    missing = sorted(required - set(header_indexes))
    if not ({"所属线ID", "所属线路径", "线名"} & set(header_indexes)):
        missing.append("所属线ID或所属线路径")
    if missing:
        return [], [{"row": 1, "message": f"缺少必需表头：{'、'.join(missing)}"}], 1

    lines = workspace_line_records(db, workspace_id)
    imported_lines = imported_lines or []
    all_lines = lines + imported_lines
    line_by_id = {line["id"]: line for line in lines}
    line_by_import_key = {line["import_key"]: line for line in imported_lines}
    line_by_path = {}
    line_by_name = {}
    for line in all_lines:
        line_by_path.setdefault(line["path"], []).append(line)
        line_by_name.setdefault(line["name"], []).append(line)
    statuses = set(get_statuses(db))
    owners = set(get_workspace_member_names(db, workspace_id))
    column_keys = {label: key for label, key, _width, _required in TASK_IMPORT_COLUMNS}
    column_keys["线名"] = "line_path"
    parsed_rows = []
    errors = []
    data_count = 0

    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), 2):
        values = [cell.value for cell in cells]
        if not any(import_cell_text(value) for value in values):
            continue
        data_count += 1
        if data_count > MAX_TASK_IMPORT_ROWS:
            errors.append({"row": row_number, "message": f"单次最多导入 {MAX_TASK_IMPORT_ROWS} 条事务"})
            break
        raw = {}
        for label, index in header_indexes.items():
            key = column_keys.get(label)
            if key:
                raw[key] = values[index] if index < len(values) else None
        try:
            line_id_text = import_cell_text(raw.get("line_id"))
            line_path_text = import_cell_text(raw.get("line_path"))
            line = None
            if line_id_text:
                line = line_by_import_key.get(line_id_text)
                if line is None:
                    try:
                        numeric_id = float(line_id_text)
                        if not numeric_id.is_integer():
                            raise ValueError
                        line_id = int(numeric_id)
                    except ValueError:
                        raise ValueError("所属线ID必须是整数或本文件中的线标识")
                    line = line_by_id.get(line_id)
                    if not line:
                        raise ValueError("所属线ID不存在于当前项目空间或本次线导入数据")
                if line_path_text and line_path_text not in {line["path"], line["name"]}:
                    raise ValueError("所属线ID与所属线路径不一致")
            elif line_path_text:
                candidates = line_by_path.get(line_path_text) or line_by_name.get(line_path_text) or []
                if not candidates:
                    raise ValueError("所属线路径不存在于当前项目空间")
                if len(candidates) > 1:
                    raise ValueError("所属线名称不唯一，请填写所属线ID或完整路径")
                line = candidates[0]
            else:
                raise ValueError("所属线ID与所属线路径至少填写一项")

            task = {
                "line_id": line["id"],
                "_line_import_key": line.get("import_key"),
                "name": import_cell_text(raw.get("name")),
                "content": import_cell_text(raw.get("content")),
                "goal": import_cell_text(raw.get("goal")),
                "next_action": import_cell_text(raw.get("next_action")),
                "risk_reason": import_cell_text(raw.get("risk_reason")),
                "priority": import_cell_text(raw.get("priority")) or "中",
                "owner": import_cell_text(raw.get("owner")),
                "status": import_cell_text(raw.get("status")),
                "start_date": import_cell_date(raw.get("start_date"), "起始日期"),
                "end_date": import_cell_date(raw.get("end_date"), "结束日期"),
            }
            for key, label in (("name", "事务名"), ("content", "事务内容"),
                               ("owner", "责任人"), ("status", "进展状态")):
                if not task[key]:
                    raise ValueError(f"{label}不能为空")
            if task["priority"] not in PRIORITY_ENUM:
                raise ValueError("非法的优先级")
            if task["status"] not in statuses:
                raise ValueError("非法的进展状态")
            if task["owner"] not in owners:
                raise ValueError("责任人不是当前项目空间成员")
            validate_date_range(task["start_date"], task["end_date"])
            if task["start_date"] < line["fork_date"]:
                raise ValueError("事务起始日期不能早于所属线起始日期")
            parsed_rows.append(task)
        except ValueError as exc:
            errors.append({"row": row_number, "message": str(exc)})
    return parsed_rows, errors, data_count


@app.errorhandler(ApiError)
def handle_api_error(exc):
    return jsonify({"error": str(exc)}), exc.status


@app.errorhandler(HTTPException)
def handle_http_error(exc):
    if request.path.startswith("/api/"):
        return jsonify({"error": exc.description}), exc.code
    return exc


@app.errorhandler(Exception)
def handle_unexpected_error(exc):
    app.logger.exception("Unhandled application error", exc_info=exc)
    if request.path.startswith("/api/"):
        return jsonify({"error": "服务器处理请求失败"}), 500
    return "Internal Server Error", 500


# ------------------------------------------------------------- auth helpers
def user_workspaces(db, user_id):
    return [dict(row) for row in db.execute(
        "SELECT w.id,w.name,w.description,w.archived_at,m.role FROM workspaces w "
        "JOIN workspace_members m ON m.workspace_id=w.id "
        "WHERE m.user_id=? ORDER BY w.name,w.id", (user_id,)
    )]


def current_workspace_id():
    return g.workspace["id"]


def require_workspace_admin(workspace_id=None):
    workspace_id = workspace_id or current_workspace_id()
    row = get_db().execute(
        "SELECT role FROM workspace_members WHERE workspace_id=? AND user_id=?",
        (workspace_id, g.user["id"]),
    ).fetchone()
    if not row or row["role"] != "admin":
        raise ApiError("仅项目管理员可执行此操作", 403)
    return row


def require_workspace_writable(workspace_id=None):
    workspace_id = workspace_id or current_workspace_id()
    row = get_db().execute(
        "SELECT archived_at FROM workspaces WHERE id=?", (workspace_id,)
    ).fetchone()
    if not row:
        raise ApiError("项目空间不存在", 404)
    if row["archived_at"]:
        raise ApiError("项目空间已归档，仅可浏览，不能编辑", 409)
    return row


CURRENT_WORKSPACE_WRITE_ENDPOINTS = {
    "api_set_statuses",
    "create_line", "update_line", "delete_line", "import_lines", "import_data",
    "create_task", "update_task", "task_dependency", "delete_task",
    "import_tasks", "bulk_tasks", "undo", "redo", "restore_trash", "purge_trash",
    "add_task_comment", "follow_task", "unfollow_task",
    "create_milestone", "update_milestone", "delete_milestone",
}


def validate_username(value):
    if not isinstance(value, str):
        raise ApiError("账号必须是字符串")
    value = value.strip()
    if len(value) < 2 or len(value) > 40 or any(c.isspace() for c in value):
        raise ApiError("账号长度应为 2-40 个字符且不能包含空格")
    return value


def validate_password(value, required=True):
    if value in (None, "") and not required:
        return None
    if not isinstance(value, str) or len(value) < 6 or len(value) > 128:
        raise ApiError("密码长度应为 6-128 个字符")
    return value


@app.before_request
def load_authenticated_context():
    if not request.path.startswith("/api/"):
        return None
    if request.endpoint in {"auth_login", "auth_session"}:
        return None
    user_id = session.get("user_id")
    if not user_id:
        raise ApiError("请先登录", 401)
    db = get_db()
    user = db.execute(
        "SELECT id,username,display_name,active FROM users WHERE id=?", (user_id,)
    ).fetchone()
    if not user or not user["active"]:
        session.clear()
        raise ApiError("账号不可用，请重新登录", 401)
    g.user = user
    workspaces = user_workspaces(db, user["id"])
    if not workspaces:
        raise ApiError("账号尚未加入任何项目空间", 403)
    workspace_id = session.get("workspace_id")
    current = next((w for w in workspaces if w["id"] == workspace_id), None)
    if current is None:
        current = workspaces[0]
        session["workspace_id"] = current["id"]
    g.workspace = current
    if request.endpoint in CURRENT_WORKSPACE_WRITE_ENDPOINTS:
        require_workspace_writable()
    return None


# ---------------------------------------------------- collaboration helpers
def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="microseconds").replace(
        "+00:00", "Z"
    )


def active_task(db, workspace_id, task_id):
    task = db.execute(
        "SELECT t.* FROM tasks t JOIN lines l ON l.id=t.line_id "
        "WHERE t.id=? AND t.workspace_id=? AND l.workspace_id=? "
        "AND t.deleted=0 AND l.deleted=0",
        (task_id, workspace_id, workspace_id),
    ).fetchone()
    if not task:
        raise ApiError("事务不存在", 404)
    return task


def workspace_member_rows(db, workspace_id):
    return db.execute(
        "SELECT u.id,u.username,u.display_name FROM workspace_members m "
        "JOIN users u ON u.id=m.user_id "
        "WHERE m.workspace_id=? AND u.active=1 ORDER BY u.display_name,u.id",
        (workspace_id,),
    ).fetchall()


def owner_user_ids(db, workspace_id, owner):
    if not owner:
        return set()
    return {
        row["id"] for row in db.execute(
            "SELECT u.id FROM workspace_members m JOIN users u ON u.id=m.user_id "
            "WHERE m.workspace_id=? AND u.active=1 AND u.display_name=?",
            (workspace_id, owner),
        )
    }


def task_audience_user_ids(db, workspace_id, task_id, owner=None):
    followers = {
        row["user_id"] for row in db.execute(
            "SELECT user_id FROM task_followers WHERE workspace_id=? AND task_id=?",
            (workspace_id, task_id),
        )
    }
    if owner is None:
        row = db.execute(
            "SELECT owner FROM tasks WHERE id=? AND workspace_id=?",
            (task_id, workspace_id),
        ).fetchone()
        owner = row["owner"] if row else ""
    return followers | owner_user_ids(db, workspace_id, owner)


def add_notifications(db, workspace_id, user_ids, kind, message, task_id=None,
                      actor_id=None, dedupe_key=None):
    created_at = now_iso()
    for user_id in set(user_ids) - ({actor_id} if actor_id else set()):
        db.execute(
            "INSERT OR IGNORE INTO notifications("
            "workspace_id,user_id,task_id,actor_id,kind,message,dedupe_key,created_at"
            ") VALUES(?,?,?,?,?,?,?,?)",
            (
                workspace_id, user_id, task_id, actor_id, kind, message,
                dedupe_key, created_at,
            ),
        )


def add_task_activity(db, workspace_id, task_id, event_type, summary,
                      metadata=None, actor_id=None):
    db.execute(
        "INSERT INTO task_activities("
        "workspace_id,task_id,actor_id,event_type,summary,metadata,created_at"
        ") VALUES(?,?,?,?,?,?,?)",
        (
            workspace_id, task_id,
            g.user["id"] if actor_id is None else actor_id,
            event_type, summary,
            json.dumps(metadata or {}, ensure_ascii=False, separators=(",", ":")),
            now_iso(),
        ),
    )


def mentioned_user_ids(db, workspace_id, content):
    mentioned = set()
    for member in workspace_member_rows(db, workspace_id):
        tokens = {f"@{member['username']}", f"@{member['display_name']}"}
        if any(
            re.search(
                re.escape(token) + r"(?![A-Za-z0-9_\-\u4e00-\u9fff])",
                content,
            )
            for token in tokens if len(token) > 1
        ):
            mentioned.add(member["id"])
    return mentioned


def ensure_due_notifications(db, workspace_id, user_id):
    today = date.today()
    rows = db.execute(
        "SELECT DISTINCT t.id,t.name,t.end_date FROM tasks t "
        "JOIN lines l ON l.id=t.line_id "
        "LEFT JOIN users owner ON owner.display_name=t.owner AND owner.active=1 "
        "LEFT JOIN workspace_members owner_member ON owner_member.user_id=owner.id "
        "AND owner_member.workspace_id=t.workspace_id "
        "LEFT JOIN task_followers follower ON follower.workspace_id=t.workspace_id "
        "AND follower.task_id=t.id AND follower.user_id=? "
        "WHERE t.workspace_id=? AND l.workspace_id=? AND t.deleted=0 AND l.deleted=0 "
        "AND t.status NOT IN ('已闭环','已取消') AND t.end_date IS NOT NULL "
        "AND (owner_member.user_id=? OR follower.user_id=?)",
        (user_id, workspace_id, workspace_id, user_id, user_id),
    ).fetchall()
    for task in rows:
        try:
            days_left = (date.fromisoformat(task["end_date"]) - today).days
        except (TypeError, ValueError):
            continue
        if days_left > 7:
            continue
        kind = "overdue" if days_left < 0 else "due_soon"
        if days_left < 0:
            message = f"事务「{task['name']}」已超期 {-days_left} 天"
        elif days_left == 0:
            message = f"事务「{task['name']}」今天到期"
        else:
            message = f"事务「{task['name']}」将在 {days_left} 天后到期"
        add_notifications(
            db, workspace_id, {user_id}, kind, message, task["id"],
            dedupe_key=f"{kind}:{task['id']}:{task['end_date']}",
        )


def notify_dependents_unblocked(db, workspace_id, prerequisite_task_id,
                                prerequisite_name, actor_id):
    rows = db.execute(
        "SELECT dependent.id,dependent.name,dependent.owner "
        "FROM task_dependencies edge "
        "JOIN tasks dependent ON dependent.id=edge.dependent_task_id "
        "JOIN lines line ON line.id=dependent.line_id "
        "WHERE edge.workspace_id=? AND edge.prerequisite_task_id=? "
        "AND dependent.workspace_id=? AND dependent.deleted=0 AND line.deleted=0 "
        "AND dependent.status NOT IN ('已闭环','已取消') "
        "AND NOT EXISTS ("
        "SELECT 1 FROM task_dependencies remaining "
        "JOIN tasks prerequisite ON prerequisite.id=remaining.prerequisite_task_id "
        "WHERE remaining.workspace_id=edge.workspace_id "
        "AND remaining.dependent_task_id=dependent.id "
        "AND prerequisite.deleted=0 "
        "AND prerequisite.status NOT IN ('已闭环','已取消'))",
        (workspace_id, prerequisite_task_id, workspace_id),
    ).fetchall()
    for task in rows:
        add_notifications(
            db, workspace_id,
            task_audience_user_ids(db, workspace_id, task["id"], task["owner"]),
            "dependency_unblocked",
            f"前置事务「{prerequisite_name}」已完成，「{task['name']}」可以继续推进",
            task["id"], actor_id,
        )


def notify_task_if_unblocked(db, workspace_id, task_id, actor_id, message):
    task = db.execute(
        "SELECT id,name,owner,status FROM tasks WHERE id=? AND workspace_id=? "
        "AND deleted=0", (task_id, workspace_id),
    ).fetchone()
    if not task or task["status"] in {"已闭环", "已取消"}:
        return
    blocked = db.execute(
        "SELECT 1 FROM task_dependencies edge "
        "JOIN tasks prerequisite ON prerequisite.id=edge.prerequisite_task_id "
        "WHERE edge.workspace_id=? AND edge.dependent_task_id=? "
        "AND prerequisite.deleted=0 "
        "AND prerequisite.status NOT IN ('已闭环','已取消') LIMIT 1",
        (workspace_id, task_id),
    ).fetchone()
    if not blocked:
        add_notifications(
            db, workspace_id,
            task_audience_user_ids(db, workspace_id, task_id, task["owner"]),
            "dependency_unblocked", message, task_id, actor_id,
        )


# ------------------------------------------------------------- undo helpers
def get_meta(db, key):
    row = db.execute(
        "SELECT value FROM workspace_meta WHERE workspace_id=? AND key=?",
        (current_workspace_id(), key),
    ).fetchone()
    return row["value"] if row else None


def set_meta(db, key, value):
    if value is None:
        db.execute(
            "DELETE FROM workspace_meta WHERE workspace_id=? AND key=?",
            (current_workspace_id(), key),
        )
    else:
        db.execute(
            "INSERT INTO workspace_meta(workspace_id,key,value) VALUES(?,?,?) "
            "ON CONFLICT(workspace_id,key) DO UPDATE SET value=excluded.value",
            (current_workspace_id(), key, str(value)),
        )


def purge_deleted(db):
    """物理删除所有软删除的行, 并清空相关撤销状态。"""
    workspace_id = current_workspace_id()
    db.execute(
        "DELETE FROM milestone_tasks WHERE workspace_id=? AND ("
        "milestone_id IN (SELECT id FROM milestones WHERE workspace_id=? AND deleted=1) "
        "OR task_id IN (SELECT id FROM tasks WHERE workspace_id=? AND deleted=1))",
        (workspace_id, workspace_id, workspace_id),
    )
    db.execute(
        "DELETE FROM task_dependencies WHERE workspace_id=? AND ("
        "dependent_task_id IN (SELECT id FROM tasks WHERE workspace_id=? AND deleted=1) "
        "OR prerequisite_task_id IN (SELECT id FROM tasks WHERE workspace_id=? AND deleted=1))",
        (workspace_id, workspace_id, workspace_id),
    )
    db.execute(
        "DELETE FROM task_images WHERE workspace_id=? AND task_id IN "
        "(SELECT id FROM tasks WHERE workspace_id=? AND deleted=1)",
        (workspace_id, workspace_id),
    )
    db.execute(
        "DELETE FROM task_attachments WHERE workspace_id=? AND task_id IN "
        "(SELECT id FROM tasks WHERE workspace_id=? AND deleted=1)",
        (workspace_id, workspace_id),
    )
    for table in ("task_followers", "task_comments", "task_activities"):
        db.execute(
            f"DELETE FROM {table} WHERE workspace_id=? AND task_id IN "
            "(SELECT id FROM tasks WHERE workspace_id=? AND deleted=1)",
            (workspace_id, workspace_id),
        )
    db.execute(
        "DELETE FROM notifications WHERE workspace_id=? AND task_id IN "
        "(SELECT id FROM tasks WHERE workspace_id=? AND deleted=1)",
        (workspace_id, workspace_id),
    )
    db.execute(
        "DELETE FROM milestones WHERE deleted=1 AND workspace_id=?", (workspace_id,)
    )
    db.execute(
        "DELETE FROM tasks WHERE deleted=1 AND workspace_id=?", (workspace_id,)
    )
    db.execute(
        "DELETE FROM lines WHERE deleted=1 AND workspace_id=?", (workspace_id,)
    )
    db.execute("DELETE FROM undo_snapshots WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM redo_snapshots WHERE workspace_id=?", (workspace_id,))
    set_meta(db, "undo_batch", None)


def current_workspace_snapshot(db, workspace_id):
    return {
        "lines": [dict(row) for row in db.execute(
            "SELECT * FROM lines WHERE workspace_id=? ORDER BY id", (workspace_id,)
        )],
        "tasks": [dict(row) for row in db.execute(
            "SELECT * FROM tasks WHERE workspace_id=? ORDER BY id", (workspace_id,)
        )],
        "milestones": [dict(row) for row in db.execute(
            "SELECT * FROM milestones WHERE workspace_id=? ORDER BY id", (workspace_id,)
        )],
        "milestone_tasks": [dict(row) for row in db.execute(
            "SELECT * FROM milestone_tasks WHERE workspace_id=? "
            "ORDER BY milestone_id,task_id", (workspace_id,)
        )],
        "task_dependencies": [dict(row) for row in db.execute(
            "SELECT * FROM task_dependencies WHERE workspace_id=? "
            "ORDER BY dependent_task_id,prerequisite_task_id", (workspace_id,)
        )],
        "task_images": [
            {
                **{key: row[key] for key in row.keys() if key != "data"},
                "data": base64.b64encode(row["data"]).decode("ascii"),
            }
            for row in db.execute(
                "SELECT * FROM task_images WHERE workspace_id=? ORDER BY id",
                (workspace_id,),
            )
        ],
        "task_attachments": [
            {
                **{key: row[key] for key in row.keys() if key != "data"},
                "data": base64.b64encode(row["data"]).decode("ascii"),
            }
            for row in db.execute(
                "SELECT * FROM task_attachments WHERE workspace_id=? ORDER BY id",
                (workspace_id,),
            )
        ],
    }


def save_workspace_snapshot(db, table, workspace_id, snapshot):
    db.execute(
        f"INSERT INTO {table}(workspace_id,snapshot,created_at) VALUES(?,?,?) "
        "ON CONFLICT(workspace_id) DO UPDATE SET "
        "snapshot=excluded.snapshot,created_at=excluded.created_at",
        (
            workspace_id,
            json.dumps(snapshot, ensure_ascii=False, separators=(",", ":")),
            date.today().isoformat(),
        ),
    )


def on_edit(db):
    """保存当前空间的编辑前状态；新编辑会清除已有恢复点。"""
    workspace_id = current_workspace_id()
    save_workspace_snapshot(
        db, "undo_snapshots", workspace_id,
        current_workspace_snapshot(db, workspace_id),
    )
    db.execute("DELETE FROM redo_snapshots WHERE workspace_id=?", (workspace_id,))


def has_undo(db):
    workspace_id = current_workspace_id()
    snapshot = db.execute(
        "SELECT 1 FROM undo_snapshots WHERE workspace_id=?", (workspace_id,)
    ).fetchone()
    return snapshot is not None or get_meta(db, "undo_batch") is not None


def has_redo(db):
    return db.execute(
        "SELECT 1 FROM redo_snapshots WHERE workspace_id=?",
        (current_workspace_id(),),
    ).fetchone() is not None


def restore_snapshot(db, snapshot):
    workspace_id = current_workspace_id()
    db.execute("DELETE FROM task_dependencies WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM milestone_tasks WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM task_images WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM task_attachments WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM milestones WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM tasks WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM lines WHERE workspace_id=?", (workspace_id,))
    for table in (
        "lines", "tasks", "milestones", "task_images", "task_attachments",
        "task_dependencies", "milestone_tasks"
    ):
        rows = snapshot.get(table, [])
        if not isinstance(rows, list):
            raise ApiError("撤销数据已损坏", 500)
        valid_columns = {
            row["name"] for row in db.execute(f"PRAGMA table_info({table})")
        }
        for row in rows:
            if not isinstance(row, dict) or not set(row).issubset(valid_columns):
                raise ApiError("撤销数据已损坏", 500)
            row = dict(row)
            if table in {"task_images", "task_attachments"}:
                try:
                    row["data"] = base64.b64decode(row["data"], validate=True)
                except (KeyError, TypeError, ValueError, binascii.Error):
                    raise ApiError("撤销数据已损坏", 500)
            row["workspace_id"] = workspace_id
            columns = list(row)
            marks = ",".join("?" for _ in columns)
            db.execute(
                f"INSERT INTO {table}({','.join(columns)}) VALUES({marks})",
                [row[column] for column in columns],
            )


def new_batch(db):
    """新的删除操作: 返回新批次号, 回收站历史批次继续保留。"""
    batch = int(get_meta(db, "batch_seq") or 0) + 1
    set_meta(db, "batch_seq", batch)
    set_meta(db, "undo_batch", batch)
    return batch


def parse_iso_date(value, field):
    if value is None or value == "":
        return None
    if not isinstance(value, str):
        raise ValueError(f"{field} 必须是 YYYY-MM-DD 格式")
    try:
        parsed = date.fromisoformat(value)
    except ValueError:
        raise ValueError(f"{field} 必须是 YYYY-MM-DD 格式")
    if parsed.isoformat() != value:
        raise ValueError(f"{field} 必须是 YYYY-MM-DD 格式")
    return parsed


def validate_date_range(start, end):
    s = parse_iso_date(start, "起始日期")
    e = parse_iso_date(end, "结束日期")
    if s and e and e < s:
        raise ValueError("结束日期不能早于起始日期")


def validate_task_images(db, workspace_id, task_id, value):
    """校验事务图片描述，返回需保留的图片 ID 和待新增的二进制图片。"""
    if not isinstance(value, list):
        raise ApiError("事务图片必须是数组")
    if len(value) > MAX_TASK_IMAGES:
        raise ApiError(f"每个事务最多可添加 {MAX_TASK_IMAGES} 张图片")

    existing_ids = []
    new_images = []
    total_bytes = 0
    for item in value:
        if not isinstance(item, dict):
            raise ApiError("事务图片格式错误")
        if "id" in item:
            image_id = required_id(item["id"], "图片 id")
            if task_id is None:
                raise ApiError("新建事务不能引用已有图片")
            if image_id in existing_ids:
                raise ApiError("事务图片不能重复")
            row = db.execute(
                "SELECT length(data) AS size FROM task_images "
                "WHERE id=? AND task_id=? AND workspace_id=?",
                (image_id, task_id, workspace_id),
            ).fetchone()
            if not row:
                raise ApiError("事务图片不存在", 404)
            existing_ids.append(image_id)
            total_bytes += row["size"]
            continue

        data_url = item.get("data_url")
        if not isinstance(data_url, str) or "," not in data_url:
            raise ApiError("事务图片格式错误")
        header, encoded = data_url.split(",", 1)
        if not header.startswith("data:") or not header.endswith(";base64"):
            raise ApiError("事务图片格式错误")
        mime_type = header[5:-7].lower()
        if mime_type not in TASK_IMAGE_TYPES:
            raise ApiError("仅支持 PNG、JPEG、GIF 或 WebP 图片")
        try:
            image_data = base64.b64decode(encoded, validate=True)
        except (ValueError, binascii.Error):
            raise ApiError("事务图片数据无效")
        if not image_data:
            raise ApiError("事务图片不能为空")
        if len(image_data) > MAX_TASK_IMAGE_BYTES:
            raise ApiError("单张事务图片不能超过 5MB")
        signatures = {
            "image/png": image_data.startswith(b"\x89PNG\r\n\x1a\n"),
            "image/jpeg": image_data.startswith(b"\xff\xd8\xff"),
            "image/gif": image_data.startswith((b"GIF87a", b"GIF89a")),
            "image/webp": image_data.startswith(b"RIFF") and
                          image_data[8:12] == b"WEBP",
        }
        if not signatures[mime_type]:
            raise ApiError("事务图片内容与格式不符")
        total_bytes += len(image_data)
        new_images.append((mime_type, image_data))

    if total_bytes > MAX_TASK_IMAGES_BYTES:
        raise ApiError("单个事务的图片总大小不能超过 20MB")
    return existing_ids, new_images


def replace_task_images(db, workspace_id, task_id, image_changes):
    existing_ids, new_images = image_changes
    if existing_ids:
        marks = ",".join("?" for _ in existing_ids)
        db.execute(
            f"DELETE FROM task_images WHERE workspace_id=? AND task_id=? "
            f"AND id NOT IN ({marks})",
            [workspace_id, task_id] + existing_ids,
        )
    else:
        db.execute(
            "DELETE FROM task_images WHERE workspace_id=? AND task_id=?",
            (workspace_id, task_id),
        )
    if new_images:
        db.executemany(
            "INSERT INTO task_images(workspace_id,task_id,mime_type,data,created_at) "
            "VALUES(?,?,?,?,?)",
            [
                (workspace_id, task_id, mime_type, image_data,
                 date.today().isoformat())
                for mime_type, image_data in new_images
            ],
        )


def validate_task_attachments(db, workspace_id, task_id, value):
    """校验事务附件描述，返回需保留的附件 ID 和待新增附件。"""
    if not isinstance(value, list):
        raise ApiError("事务附件必须是数组")
    if len(value) > MAX_TASK_ATTACHMENTS:
        raise ApiError(f"每个事务最多可添加 {MAX_TASK_ATTACHMENTS} 个附件")

    existing_ids = []
    new_attachments = []
    total_bytes = 0
    for item in value:
        if not isinstance(item, dict):
            raise ApiError("事务附件格式错误")
        if "id" in item:
            attachment_id = required_id(item["id"], "附件 id")
            if task_id is None:
                raise ApiError("新建事务不能引用已有附件")
            if attachment_id in existing_ids:
                raise ApiError("事务附件不能重复")
            row = db.execute(
                "SELECT length(data) AS size FROM task_attachments "
                "WHERE id=? AND task_id=? AND workspace_id=?",
                (attachment_id, task_id, workspace_id),
            ).fetchone()
            if not row:
                raise ApiError("事务附件不存在", 404)
            existing_ids.append(attachment_id)
            total_bytes += row["size"]
            continue

        filename = item.get("name")
        if not isinstance(filename, str):
            raise ApiError("附件名称不能为空")
        filename = filename.replace("\\", "/").rsplit("/", 1)[-1].strip()
        if not filename or any(ord(char) < 32 for char in filename):
            raise ApiError("附件名称无效")
        if len(filename) > MAX_ATTACHMENT_FILENAME_LENGTH:
            raise ApiError(
                f"附件名称不能超过 {MAX_ATTACHMENT_FILENAME_LENGTH} 个字符"
            )

        data_url = item.get("data_url")
        if not isinstance(data_url, str) or "," not in data_url:
            raise ApiError("事务附件格式错误")
        header, encoded = data_url.split(",", 1)
        if not header.startswith("data:") or not header.endswith(";base64"):
            raise ApiError("事务附件格式错误")
        mime_type = header[5:-7].strip().lower() or "application/octet-stream"
        if len(mime_type) > 200 or "/" not in mime_type or any(
                char in mime_type for char in "\r\n"):
            raise ApiError("附件类型无效")
        try:
            attachment_data = base64.b64decode(encoded, validate=True)
        except (ValueError, binascii.Error):
            raise ApiError("事务附件数据无效")
        if not attachment_data:
            raise ApiError("事务附件不能为空")
        if len(attachment_data) > MAX_TASK_ATTACHMENT_BYTES:
            raise ApiError("单个事务附件不能超过 5MB")
        total_bytes += len(attachment_data)
        new_attachments.append((filename, mime_type, attachment_data))

    if total_bytes > MAX_TASK_ATTACHMENTS_BYTES:
        raise ApiError("单个事务的附件总大小不能超过 20MB")
    return existing_ids, new_attachments


def replace_task_attachments(db, workspace_id, task_id, attachment_changes):
    existing_ids, new_attachments = attachment_changes
    if existing_ids:
        marks = ",".join("?" for _ in existing_ids)
        db.execute(
            f"DELETE FROM task_attachments WHERE workspace_id=? AND task_id=? "
            f"AND id NOT IN ({marks})",
            [workspace_id, task_id] + existing_ids,
        )
    else:
        db.execute(
            "DELETE FROM task_attachments WHERE workspace_id=? AND task_id=?",
            (workspace_id, task_id),
        )
    if new_attachments:
        db.executemany(
            "INSERT INTO task_attachments("
            "workspace_id,task_id,filename,mime_type,data,created_at) "
            "VALUES(?,?,?,?,?,?)",
            [
                (
                    workspace_id, task_id, filename, mime_type,
                    attachment_data, date.today().isoformat(),
                )
                for filename, mime_type, attachment_data in new_attachments
            ],
        )


def dependency_id_list(value):
    if not isinstance(value, list) or not all(
            isinstance(item, int) and not isinstance(item, bool) for item in value):
        raise ApiError("依赖事务必须是整数数组")
    if len(value) != len(set(value)):
        raise ApiError("依赖事务不能包含重复项")
    return value


def current_dependency_ids(db, workspace_id, dependent_task_id):
    return [row["prerequisite_task_id"] for row in db.execute(
        "SELECT prerequisite_task_id FROM task_dependencies "
        "WHERE workspace_id=? AND dependent_task_id=? "
        "ORDER BY prerequisite_task_id",
        (workspace_id, dependent_task_id),
    )]


def validate_dependencies(db, workspace_id, dependent_task_id, prerequisite_ids):
    prerequisite_ids = dependency_id_list(prerequisite_ids)
    if dependent_task_id is not None and dependent_task_id in prerequisite_ids:
        raise ApiError("事务不能依赖自身")
    if prerequisite_ids:
        marks = ",".join("?" * len(prerequisite_ids))
        rows = db.execute(
            f"SELECT id FROM tasks WHERE workspace_id=? AND deleted=0 "
            f"AND id IN ({marks})",
            [workspace_id] + prerequisite_ids,
        ).fetchall()
        if len(rows) != len(prerequisite_ids):
            raise ApiError("部分依赖事务不存在或已删除", 404)

    if dependent_task_id is None:
        return prerequisite_ids

    graph = {}
    for row in db.execute(
        "SELECT d.dependent_task_id,d.prerequisite_task_id "
        "FROM task_dependencies d "
        "JOIN tasks dependent ON dependent.id=d.dependent_task_id "
        "JOIN tasks prerequisite ON prerequisite.id=d.prerequisite_task_id "
        "WHERE d.workspace_id=? AND dependent.workspace_id=? "
        "AND prerequisite.workspace_id=? AND dependent.deleted=0 "
        "AND prerequisite.deleted=0",
        (workspace_id, workspace_id, workspace_id),
    ):
        graph.setdefault(row["dependent_task_id"], set()).add(
            row["prerequisite_task_id"]
        )
    graph[dependent_task_id] = set(prerequisite_ids)

    visiting, visited = set(), set()

    def walk(task_id):
        if task_id in visiting:
            return True
        if task_id in visited:
            return False
        visiting.add(task_id)
        if any(walk(next_id) for next_id in graph.get(task_id, ())):
            return True
        visiting.remove(task_id)
        visited.add(task_id)
        return False

    if walk(dependent_task_id):
        raise ApiError("事务依赖不能形成循环")
    return prerequisite_ids


def ensure_dependencies_closed(db, workspace_id, prerequisite_ids):
    if not prerequisite_ids:
        return
    marks = ",".join("?" * len(prerequisite_ids))
    rows = db.execute(
        f"SELECT id,name FROM tasks WHERE workspace_id=? AND deleted=0 "
        f"AND status<>'已闭环' AND id IN ({marks}) ORDER BY id",
        [workspace_id] + prerequisite_ids,
    ).fetchall()
    if rows:
        names = "、".join(row["name"] for row in rows[:3])
        if len(rows) > 3:
            names += f"等 {len(rows)} 项"
        raise ApiError(f"被依赖事务尚未闭环：{names}", 409)


def replace_task_dependencies(db, workspace_id, dependent_task_id, prerequisite_ids):
    db.execute(
        "DELETE FROM task_dependencies WHERE workspace_id=? AND dependent_task_id=?",
        (workspace_id, dependent_task_id),
    )
    db.executemany(
        "INSERT INTO task_dependencies("
        "workspace_id,dependent_task_id,prerequisite_task_id) VALUES(?,?,?)",
        [
            (workspace_id, dependent_task_id, prerequisite_id)
            for prerequisite_id in prerequisite_ids
        ],
    )


def ensure_tasks_not_required(
        db, workspace_id, task_ids, excluded_milestone_ids=None):
    if not task_ids:
        return
    marks = ",".join("?" * len(task_ids))
    params = [workspace_id, workspace_id] + task_ids + task_ids
    row = db.execute(
        f"SELECT dependent.name AS dependent_name,prerequisite.name AS prerequisite_name "
        f"FROM task_dependencies d "
        f"JOIN tasks dependent ON dependent.id=d.dependent_task_id "
        f"JOIN tasks prerequisite ON prerequisite.id=d.prerequisite_task_id "
        f"WHERE d.workspace_id=? AND dependent.workspace_id=? "
        f"AND dependent.deleted=0 AND d.prerequisite_task_id IN ({marks}) "
        f"AND d.dependent_task_id NOT IN ({marks}) LIMIT 1",
        params,
    ).fetchone()
    if row:
        raise ApiError(
            f"事务“{row['prerequisite_name']}”仍被“{row['dependent_name']}”依赖，不能删除",
            409,
        )
    excluded_milestone_ids = list(excluded_milestone_ids or [])
    milestone_params = [workspace_id, workspace_id] + task_ids
    excluded_clause = ""
    if excluded_milestone_ids:
        excluded_marks = ",".join("?" * len(excluded_milestone_ids))
        excluded_clause = f" AND milestone.id NOT IN ({excluded_marks})"
        milestone_params.extend(excluded_milestone_ids)
    milestone = db.execute(
        f"SELECT milestone.name AS milestone_name,task.name AS task_name "
        f"FROM milestone_tasks relation "
        f"JOIN milestones milestone ON milestone.id=relation.milestone_id "
        f"JOIN tasks task ON task.id=relation.task_id "
        f"WHERE relation.workspace_id=? AND milestone.workspace_id=? "
        f"AND milestone.deleted=0 AND relation.task_id IN ({marks})"
        f"{excluded_clause} LIMIT 1",
        milestone_params,
    ).fetchone()
    if milestone:
        raise ApiError(
            f"事务“{milestone['task_name']}”仍是里程碑“{milestone['milestone_name']}”"
            "的验收条件，不能删除",
            409,
        )


def milestone_task_id_list(value):
    if not isinstance(value, list) or not all(
            isinstance(item, int) and not isinstance(item, bool) for item in value):
        raise ApiError("验收事务必须是整数数组")
    if len(value) != len(set(value)):
        raise ApiError("验收事务不能包含重复项")
    return value


def prepare_avatar_data(data_url):
    if not isinstance(data_url, str) or "," not in data_url:
        raise ApiError("头像数据格式不正确")
    header, encoded = data_url.split(",", 1)
    if not header.startswith("data:") or not header.endswith(";base64"):
        raise ApiError("头像数据格式不正确")
    declared_mime = header[5:-7].lower()
    if declared_mime not in AVATAR_IMAGE_TYPES:
        raise ApiError("头像仅支持 PNG、JPEG 或 WebP 格式")
    try:
        source_data = base64.b64decode(encoded, validate=True)
    except (binascii.Error, ValueError):
        raise ApiError("头像数据无法解析")
    if not source_data:
        raise ApiError("头像文件不能为空")
    if len(source_data) > MAX_AVATAR_SOURCE_BYTES:
        raise ApiError("头像源文件不能超过 5MB")

    try:
        with Image.open(BytesIO(source_data)) as opened:
            actual_mime = {
                "PNG": "image/png", "JPEG": "image/jpeg", "WEBP": "image/webp",
            }.get(opened.format)
            if actual_mime != declared_mime:
                raise ApiError("头像文件格式与内容不一致")
            width, height = opened.size
            if min(width, height) < 64:
                raise ApiError("头像宽度和高度均不能小于 64 像素")
            if width * height > MAX_AVATAR_SOURCE_PIXELS:
                raise ApiError("头像像素尺寸过大")
            opened.load()
            source = ImageOps.exif_transpose(opened)
            has_alpha = "A" in source.getbands() or (
                source.mode == "P" and "transparency" in source.info
            )
            source = source.convert("RGBA" if has_alpha else "RGB")
            avatar = ImageOps.fit(
                source, (AVATAR_SIZE, AVATAR_SIZE),
                method=Image.Resampling.LANCZOS, centering=(0.5, 0.5),
            )
            output = BytesIO()
            if has_alpha:
                avatar.save(output, format="PNG", optimize=True)
                mime_type = "image/png"
            else:
                avatar.save(output, format="JPEG", quality=88, optimize=True)
                mime_type = "image/jpeg"
    except ApiError:
        raise
    except (UnidentifiedImageError, OSError, ValueError, Image.DecompressionBombError):
        raise ApiError("头像不是有效的图片文件")

    avatar_data = output.getvalue()
    if len(avatar_data) > MAX_AVATAR_BYTES:
        raise ApiError("处理后的头像文件过大")
    return mime_type, avatar_data


def validate_milestone_tasks(db, workspace_id, task_ids):
    task_ids = milestone_task_id_list(task_ids)
    if task_ids:
        marks = ",".join("?" * len(task_ids))
        rows = db.execute(
            f"SELECT id FROM tasks WHERE workspace_id=? AND deleted=0 "
            f"AND id IN ({marks})",
            [workspace_id] + task_ids,
        ).fetchall()
        if len(rows) != len(task_ids):
            raise ApiError("部分验收事务不存在或已删除", 404)
    return task_ids


def replace_milestone_tasks(db, workspace_id, milestone_id, task_ids):
    db.execute(
        "DELETE FROM milestone_tasks WHERE workspace_id=? AND milestone_id=?",
        (workspace_id, milestone_id),
    )
    db.executemany(
        "INSERT INTO milestone_tasks(workspace_id,milestone_id,task_id) "
        "VALUES(?,?,?)",
        [(workspace_id, milestone_id, task_id) for task_id in task_ids],
    )


def get_statuses(db):
    raw = get_meta(db, "statuses")
    statuses = None
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list) and all(isinstance(s, str) for s in parsed):
                statuses = [s.strip() for s in parsed if s.strip()]
        except ValueError:
            statuses = None
    statuses = statuses or list(DEFAULT_STATUS_ENUM)
    # 保留历史数据中的状态, 避免配置变更后老事务无法编辑。
    rows = db.execute(
        "SELECT DISTINCT status FROM tasks WHERE workspace_id=? "
        "AND status IS NOT NULL ORDER BY status", (current_workspace_id(),)
    ).fetchall()
    for r in rows:
        if r["status"] and r["status"] not in statuses:
            statuses.append(r["status"])
    return statuses


def get_status_colors(db, statuses=None):
    statuses = statuses or get_statuses(db)
    stored = {}
    raw = get_meta(db, "status_colors")
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                for status, color in parsed.items():
                    try:
                        normalized = line_color(color)
                    except ApiError:
                        continue
                    if isinstance(status, str) and status and normalized:
                        stored[status] = normalized
        except ValueError:
            pass
    return {
        status: stored.get(
            status, DEFAULT_STATUS_COLORS.get(status, FALLBACK_STATUS_COLOR)
        )
        for status in statuses
    }


# ------------------------------------------------------------------- routes
@app.route("/")
def index():
    return send_from_directory(str(app.static_folder), "index.html")


def session_payload(db, user):
    workspaces = user_workspaces(db, user["id"])
    if not workspaces:
        session.clear()
        raise ApiError("账号尚未加入任何项目空间，请联系管理员", 403)
    workspace_id = session.get("workspace_id")
    current = next((w for w in workspaces if w["id"] == workspace_id), None)
    if current is None and workspaces:
        current = workspaces[0]
        session["workspace_id"] = current["id"]
    avatar = db.execute(
        "SELECT avatar_data IS NOT NULL AS has_avatar,avatar_updated_at "
        "FROM users WHERE id=?", (user["id"],),
    ).fetchone()
    avatar_url = None
    if avatar and avatar["has_avatar"]:
        avatar_url = f"/api/auth/avatar?v={avatar['avatar_updated_at'] or ''}"
    return {
        "authenticated": True,
        "user": {
            "id": user["id"], "username": user["username"],
            "display_name": user["display_name"], "avatar_url": avatar_url,
        },
        "workspaces": workspaces,
        "current_workspace": current,
    }


@app.route("/api/auth/session")
def auth_session():
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"authenticated": False})
    db = get_db()
    user = db.execute(
        "SELECT id,username,display_name,active FROM users WHERE id=?", (user_id,)
    ).fetchone()
    if not user or not user["active"]:
        session.clear()
        return jsonify({"authenticated": False})
    return jsonify(session_payload(db, user))


@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = json_object()
    username = validate_username(data.get("username"))
    password = data.get("password")
    if not isinstance(password, str):
        raise ApiError("密码必须是字符串")
    db = get_db()
    user = db.execute(
        "SELECT id,username,display_name,password_hash,active FROM users "
        "WHERE username=?", (username,)
    ).fetchone()
    if not user or not user["active"] or not check_password_hash(
            user["password_hash"], password):
        raise ApiError("账号或密码错误", 401)
    session.clear()
    session["user_id"] = user["id"]
    return jsonify(session_payload(db, user))


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/auth/password", methods=["PUT"])
def auth_change_password():
    data = json_object()
    current_password = data.get("current_password")
    new_password = validate_password(data.get("new_password"))
    db = get_db()
    user = db.execute(
        "SELECT password_hash FROM users WHERE id=?", (g.user["id"],)
    ).fetchone()
    if not isinstance(current_password, str) or not check_password_hash(
            user["password_hash"], current_password):
        raise ApiError("当前密码错误", 400)
    db.execute(
        "UPDATE users SET password_hash=?,updated_at=? WHERE id=?",
        (generate_password_hash(new_password), date.today().isoformat(), g.user["id"]),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/workspaces", methods=["POST"])
def create_workspace():
    data = json_object()
    name = text_field(data, "name", "项目空间名称").strip()
    description = text_field(data, "description", "项目空间描述").strip()
    if not name:
        raise ApiError("项目空间名称不能为空")
    db = get_db()
    can_create = db.execute(
        "SELECT 1 FROM workspace_members WHERE user_id=? AND role='admin' LIMIT 1",
        (g.user["id"],),
    ).fetchone()
    if not can_create:
        raise ApiError("仅管理员可创建项目空间", 403)
    today = date.today().isoformat()
    cur = db.execute(
        "INSERT INTO workspaces(name,description,created_by,created_at,updated_at) "
        "VALUES(?,?,?,?,?)", (name, description, g.user["id"], today, today),
    )
    workspace_id = cur.lastrowid
    db.execute(
        "INSERT INTO workspace_members(workspace_id,user_id,role,joined_at) "
        "VALUES(?,?,?,?)", (workspace_id, g.user["id"], "admin", today),
    )
    db.commit()
    session["workspace_id"] = workspace_id
    return jsonify({"id": workspace_id}), 201


@app.route("/api/workspaces/<int:workspace_id>", methods=["PATCH"])
def update_workspace(workspace_id):
    require_workspace_admin(workspace_id)
    require_workspace_writable(workspace_id)
    data = json_object()
    fields, values = [], []
    for key, label in (("name", "项目空间名称"), ("description", "项目空间描述")):
        if key in data:
            value = text_field(data, key, label).strip()
            if key == "name" and not value:
                raise ApiError("项目空间名称不能为空")
            fields.append(f"{key}=?")
            values.append(value)
    if fields:
        fields.append("updated_at=?")
        values.extend([date.today().isoformat(), workspace_id])
        get_db().execute(
            f"UPDATE workspaces SET {','.join(fields)} WHERE id=?", values
        )
        get_db().commit()
    return jsonify({"ok": True})


@app.route("/api/workspaces/<int:workspace_id>/archive", methods=["POST"])
def archive_workspace(workspace_id):
    require_workspace_admin(workspace_id)
    db = get_db()
    workspace = db.execute(
        "SELECT archived_at FROM workspaces WHERE id=?", (workspace_id,)
    ).fetchone()
    if not workspace:
        raise ApiError("项目空间不存在", 404)
    if workspace["archived_at"]:
        raise ApiError("项目空间已经归档", 409)
    archived_at = date.today().isoformat()
    db.execute(
        "UPDATE workspaces SET archived_at=?,updated_at=? WHERE id=?",
        (archived_at, archived_at, workspace_id),
    )
    db.commit()
    return jsonify({"ok": True, "archived_at": archived_at})


@app.route("/api/workspaces/<int:workspace_id>/restore", methods=["POST"])
def restore_workspace(workspace_id):
    require_workspace_admin(workspace_id)
    db = get_db()
    workspace = db.execute(
        "SELECT archived_at FROM workspaces WHERE id=?", (workspace_id,)
    ).fetchone()
    if not workspace:
        raise ApiError("项目空间不存在", 404)
    if not workspace["archived_at"]:
        raise ApiError("项目空间尚未归档", 409)
    db.execute(
        "UPDATE workspaces SET archived_at=NULL,updated_at=? WHERE id=?",
        (date.today().isoformat(), workspace_id),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/workspaces/<int:workspace_id>", methods=["DELETE"])
def delete_workspace(workspace_id):
    require_workspace_admin(workspace_id)
    db = get_db()
    workspace = db.execute(
        "SELECT name FROM workspaces WHERE id=?", (workspace_id,)
    ).fetchone()
    if not workspace:
        raise ApiError("项目空间不存在", 404)
    confirmation = text_field(json_object(), "confirmation", "确认项目名称")
    if confirmation != workspace["name"]:
        raise ApiError("输入的项目空间名称不匹配，无法删除")
    remaining = [
        workspace for workspace in user_workspaces(db, g.user["id"])
        if workspace["id"] != workspace_id
    ]
    if not remaining:
        raise ApiError("至少需要保留一个可访问的项目空间", 409)

    db.execute("DELETE FROM task_dependencies WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM milestone_tasks WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM task_images WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM task_attachments WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM task_followers WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM task_comments WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM task_activities WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM notifications WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM milestones WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM tasks WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM lines WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM workspace_meta WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM undo_snapshots WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM redo_snapshots WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM dashboard_snapshots WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM workspace_members WHERE workspace_id=?", (workspace_id,))
    db.execute("DELETE FROM workspaces WHERE id=?", (workspace_id,))
    db.commit()

    if session.get("workspace_id") == workspace_id:
        session["workspace_id"] = remaining[0]["id"]
    return jsonify({
        "ok": True,
        "current_workspace_id": session.get("workspace_id"),
    })


@app.route("/api/workspaces/<int:workspace_id>/select", methods=["POST"])
def select_workspace(workspace_id):
    row = get_db().execute(
        "SELECT 1 FROM workspace_members WHERE workspace_id=? AND user_id=?",
        (workspace_id, g.user["id"]),
    ).fetchone()
    if not row:
        raise ApiError("无权访问该项目空间", 403)
    session["workspace_id"] = workspace_id
    return jsonify({"ok": True})


@app.route("/api/workspaces/<int:workspace_id>/members", methods=["GET", "POST"])
def workspace_members(workspace_id):
    require_workspace_admin(workspace_id)
    db = get_db()
    if request.method == "GET":
        rows = [dict(row) for row in db.execute(
            "SELECT u.id,u.username,u.display_name,m.role,m.joined_at,"
            "CASE WHEN u.managed_by=? THEN 1 ELSE 0 END AS can_manage_account "
            "FROM workspace_members m JOIN users u ON u.id=m.user_id "
            "WHERE m.workspace_id=? ORDER BY m.role,u.display_name,u.id",
            (g.user["id"], workspace_id),
        )]
        return jsonify({"members": rows})

    require_workspace_writable(workspace_id)
    data = json_object()
    username = validate_username(data.get("username"))
    display_name = text_field(data, "display_name", "姓名", username).strip() or username
    role = data.get("role", "member")
    if role not in {"admin", "member"}:
        raise ApiError("角色必须是 admin 或 member")
    user = db.execute(
        "SELECT id FROM users WHERE username=?", (username,)
    ).fetchone()
    today = date.today().isoformat()
    if user:
        user_id = user["id"]
    else:
        password = validate_password(data.get("password"))
        cur = db.execute(
            "INSERT INTO users(username,display_name,password_hash,managed_by,created_at,updated_at) "
            "VALUES(?,?,?,?,?,?)",
            (
                username, display_name, generate_password_hash(password),
                g.user["id"], today, today,
            ),
        )
        user_id = cur.lastrowid
    try:
        db.execute(
            "INSERT INTO workspace_members(workspace_id,user_id,role,joined_at) "
            "VALUES(?,?,?,?)", (workspace_id, user_id, role, today),
        )
    except sqlite3.IntegrityError:
        raise ApiError("该账号已在项目空间中", 409)
    db.commit()
    return jsonify({"user_id": user_id}), 201


@app.route(
    "/api/workspaces/<int:workspace_id>/members/<int:user_id>",
    methods=["PATCH", "DELETE"],
)
def workspace_member(workspace_id, user_id):
    require_workspace_admin(workspace_id)
    require_workspace_writable(workspace_id)
    db = get_db()
    member = db.execute(
        "SELECT m.role,u.managed_by FROM workspace_members m "
        "JOIN users u ON u.id=m.user_id "
        "WHERE m.workspace_id=? AND m.user_id=?",
        (workspace_id, user_id),
    ).fetchone()
    if not member:
        raise ApiError("成员不存在", 404)
    if user_id == g.user["id"] and request.method == "DELETE":
        raise ApiError("不能将自己移出当前项目空间")

    def ensure_not_last_admin(new_role=None):
        if member["role"] != "admin" or new_role == "admin":
            return
        count = db.execute(
            "SELECT COUNT(*) AS count FROM workspace_members "
            "WHERE workspace_id=? AND role='admin'", (workspace_id,),
        ).fetchone()["count"]
        if count <= 1:
            raise ApiError("项目空间至少需要保留一名管理员")

    if request.method == "DELETE":
        ensure_not_last_admin()
        db.execute(
            "DELETE FROM task_followers WHERE workspace_id=? AND user_id=?",
            (workspace_id, user_id),
        )
        db.execute(
            "DELETE FROM notifications WHERE workspace_id=? AND user_id=?",
            (workspace_id, user_id),
        )
        db.execute(
            "DELETE FROM workspace_members WHERE workspace_id=? AND user_id=?",
            (workspace_id, user_id),
        )
        db.commit()
        return jsonify({"ok": True})

    data = json_object()
    if "role" in data:
        role = data["role"]
        if role not in {"admin", "member"}:
            raise ApiError("角色必须是 admin 或 member")
        ensure_not_last_admin(role)
        db.execute(
            "UPDATE workspace_members SET role=? WHERE workspace_id=? AND user_id=?",
            (role, workspace_id, user_id),
        )
    fields, values = [], []
    account_changes = "display_name" in data or data.get("password") not in (None, "")
    if account_changes and member["managed_by"] != g.user["id"]:
        raise ApiError("该账号由其他管理员维护，仅可调整其空间角色", 403)
    if "display_name" in data:
        display_name = text_field(data, "display_name", "姓名").strip()
        if not display_name:
            raise ApiError("姓名不能为空")
        fields.append("display_name=?")
        values.append(display_name)
    if data.get("password") not in (None, ""):
        fields.append("password_hash=?")
        values.append(generate_password_hash(validate_password(data["password"])))
    if fields:
        fields.append("updated_at=?")
        values.extend([date.today().isoformat(), user_id])
        db.execute(f"UPDATE users SET {','.join(fields)} WHERE id=?", values)
    db.commit()
    return jsonify({"ok": True})


def update_dashboard_snapshot(db, workspace_id, tasks, dependencies):
    today = date.today().isoformat()
    done_statuses = {"已闭环", "已取消"}
    task_by_id = {task["id"]: task for task in tasks}
    unfinished_ids = {
        task["id"] for task in tasks if task["status"] not in done_statuses
    }
    blocked_ids = {
        dependency["dependent_task_id"]
        for dependency in dependencies
        if dependency["dependent_task_id"] in unfinished_ids
        and dependency["prerequisite_task_id"] in unfinished_ids
        and dependency["prerequisite_task_id"] in task_by_id
    }
    status_counts = {}
    for task in tasks:
        status = task["status"] or "未设置"
        status_counts[status] = status_counts.get(status, 0) + 1

    metrics = {
        "total": len(tasks),
        "done": sum(task["status"] in done_statuses for task in tasks),
        "overdue": sum(
            task["id"] in unfinished_ids
            and bool(task["end_date"])
            and task["end_date"] < today
            for task in tasks
        ),
        "risk": sum(task["status"] == "有风险" for task in tasks),
        "blocked": len(blocked_ids),
    }
    db.execute(
        "INSERT INTO dashboard_snapshots("
        "workspace_id,snapshot_date,total,done,overdue,risk,blocked,status_counts"
        ") VALUES(?,?,?,?,?,?,?,?) "
        "ON CONFLICT(workspace_id,snapshot_date) DO UPDATE SET "
        "total=excluded.total,done=excluded.done,overdue=excluded.overdue,"
        "risk=excluded.risk,blocked=excluded.blocked,"
        "status_counts=excluded.status_counts",
        (
            workspace_id, today, metrics["total"], metrics["done"],
            metrics["overdue"], metrics["risk"], metrics["blocked"],
            json.dumps(status_counts, ensure_ascii=False, sort_keys=True),
        ),
    )
    db.commit()
    rows = db.execute(
        "SELECT snapshot_date,total,done,overdue,risk,blocked,status_counts "
        "FROM dashboard_snapshots WHERE workspace_id=? "
        "ORDER BY snapshot_date DESC LIMIT 90",
        (workspace_id,),
    ).fetchall()
    snapshots = []
    for row in reversed(rows):
        snapshot = dict(row)
        try:
            snapshot["status_counts"] = json.loads(snapshot["status_counts"])
        except (TypeError, json.JSONDecodeError):
            snapshot["status_counts"] = {}
        snapshots.append(snapshot)
    return snapshots


@app.route("/api/state")
def api_state():
    db = get_db()
    workspace_id = current_workspace_id()
    statuses = get_statuses(db)
    lines = [dict(r) for r in db.execute(
        "SELECT id,name,description,color,parent_id,fork_date,merge_date,updated_at "
        "FROM lines "
        "WHERE workspace_id=? AND deleted=0 ORDER BY id", (workspace_id,))]
    tasks = [dict(r) for r in db.execute(
        "SELECT t.id,t.line_id,t.name,t.content,t.goal,t.owner,t.priority,"
        "t.next_action,t.risk_reason,t.status,t.start_date,t.end_date,"
        "t.status_since,t.updated_at FROM tasks t "
        "JOIN lines l ON l.id=t.line_id "
        "WHERE t.workspace_id=? AND l.workspace_id=? "
        "AND t.deleted=0 AND l.deleted=0 ORDER BY t.start_date,t.id",
        (workspace_id, workspace_id))]
    dependencies = [dict(r) for r in db.execute(
        "SELECT d.dependent_task_id,d.prerequisite_task_id "
        "FROM task_dependencies d "
        "JOIN tasks dependent ON dependent.id=d.dependent_task_id "
        "JOIN tasks prerequisite ON prerequisite.id=d.prerequisite_task_id "
        "WHERE d.workspace_id=? AND dependent.workspace_id=? "
        "AND prerequisite.workspace_id=? AND dependent.deleted=0 "
        "AND prerequisite.deleted=0 "
        "ORDER BY d.dependent_task_id,d.prerequisite_task_id",
        (workspace_id, workspace_id, workspace_id),
    )]
    milestone_rows = [dict(r) for r in db.execute(
        "SELECT milestone.id,milestone.line_id,milestone.name,"
        "milestone.target_description,milestone.milestone_date,"
        "milestone.updated_at FROM milestones milestone "
        "JOIN lines line ON line.id=milestone.line_id "
        "WHERE milestone.workspace_id=? AND line.workspace_id=? "
        "AND milestone.deleted=0 AND line.deleted=0 "
        "ORDER BY milestone.milestone_date,milestone.id",
        (workspace_id, workspace_id),
    )]
    milestone_task_ids = {}
    for relation in db.execute(
        "SELECT relation.milestone_id,relation.task_id "
        "FROM milestone_tasks relation "
        "JOIN milestones milestone ON milestone.id=relation.milestone_id "
        "JOIN tasks task ON task.id=relation.task_id "
        "WHERE relation.workspace_id=? AND milestone.workspace_id=? "
        "AND task.workspace_id=? AND milestone.deleted=0 AND task.deleted=0 "
        "ORDER BY relation.milestone_id,relation.task_id",
        (workspace_id, workspace_id, workspace_id),
    ):
        milestone_task_ids.setdefault(relation["milestone_id"], []).append(
            relation["task_id"]
        )
    for milestone in milestone_rows:
        milestone["acceptance_task_ids"] = milestone_task_ids.get(
            milestone["id"], []
        )
    task_images = [dict(r) for r in db.execute(
        "SELECT image.id,image.task_id,image.mime_type "
        "FROM task_images image "
        "JOIN tasks task ON task.id=image.task_id "
        "JOIN lines line ON line.id=task.line_id "
        "WHERE image.workspace_id=? AND task.workspace_id=? "
        "AND line.workspace_id=? AND task.deleted=0 AND line.deleted=0 "
        "ORDER BY image.id",
        (workspace_id, workspace_id, workspace_id),
    )]
    task_attachments = [dict(r) for r in db.execute(
        "SELECT attachment.id,attachment.task_id,attachment.filename,"
        "attachment.mime_type,length(attachment.data) AS size "
        "FROM task_attachments attachment "
        "JOIN tasks task ON task.id=attachment.task_id "
        "JOIN lines line ON line.id=task.line_id "
        "WHERE attachment.workspace_id=? AND task.workspace_id=? "
        "AND line.workspace_id=? AND task.deleted=0 AND line.deleted=0 "
        "ORDER BY attachment.id",
        (workspace_id, workspace_id, workspace_id),
    )]
    dashboard_snapshots = update_dashboard_snapshot(
        db, workspace_id, tasks, dependencies
    )
    ensure_due_notifications(db, workspace_id, g.user["id"])
    db.commit()
    unread_notifications = db.execute(
        "SELECT COUNT(*) AS count FROM notifications "
        "WHERE workspace_id=? AND user_id=? AND read_at IS NULL",
        (workspace_id, g.user["id"]),
    ).fetchone()["count"]
    return jsonify({
        "lines": lines,
        "tasks": tasks,
        "milestones": milestone_rows,
        "dependencies": dependencies,
        "task_images": task_images,
        "task_attachments": task_attachments,
        "can_undo": has_undo(db),
        "can_redo": has_redo(db),
        "status_enum": statuses,
        "status_colors": get_status_colors(db, statuses),
        "priority_enum": PRIORITY_ENUM,
        "owners": get_workspace_member_names(db, workspace_id),
        "collaboration_members": [dict(row) for row in workspace_member_rows(
            db, workspace_id
        )],
        "unread_notifications": unread_notifications,
        "today": date.today().isoformat(),
        "dashboard_snapshots": dashboard_snapshots,
    })


@app.route("/api/notifications")
def api_notifications():
    db = get_db()
    workspace_id = current_workspace_id()
    ensure_due_notifications(db, workspace_id, g.user["id"])
    db.commit()
    rows = [dict(row) for row in db.execute(
        "SELECT n.id,n.task_id,n.kind,n.message,n.read_at,n.created_at,"
        "actor.display_name AS actor_name,t.name AS task_name,"
        "CASE WHEN t.id IS NOT NULL AND t.deleted=0 AND l.deleted=0 "
        "THEN 1 ELSE 0 END AS task_available "
        "FROM notifications n "
        "LEFT JOIN users actor ON actor.id=n.actor_id "
        "LEFT JOIN tasks t ON t.id=n.task_id AND t.workspace_id=n.workspace_id "
        "LEFT JOIN lines l ON l.id=t.line_id AND l.workspace_id=n.workspace_id "
        "WHERE n.workspace_id=? AND n.user_id=? "
        "ORDER BY CASE WHEN n.read_at IS NULL THEN 0 ELSE 1 END,n.created_at DESC,n.id DESC "
        "LIMIT 100",
        (workspace_id, g.user["id"]),
    )]
    unread_count = db.execute(
        "SELECT COUNT(*) AS count FROM notifications "
        "WHERE workspace_id=? AND user_id=? AND read_at IS NULL",
        (workspace_id, g.user["id"]),
    ).fetchone()["count"]
    return jsonify({
        "notifications": rows,
        "unread_count": unread_count,
    })


@app.route("/api/notifications/read-all", methods=["POST"])
def read_all_notifications():
    db = get_db()
    db.execute(
        "UPDATE notifications SET read_at=? WHERE workspace_id=? AND user_id=? "
        "AND read_at IS NULL",
        (now_iso(), current_workspace_id(), g.user["id"]),
    )
    db.commit()
    return jsonify({"ok": True, "unread_count": 0})


@app.route("/api/notifications/<int:notification_id>/read", methods=["POST"])
def read_notification(notification_id):
    db = get_db()
    result = db.execute(
        "UPDATE notifications SET read_at=COALESCE(read_at,?) "
        "WHERE id=? AND workspace_id=? AND user_id=?",
        (now_iso(), notification_id, current_workspace_id(), g.user["id"]),
    )
    if result.rowcount == 0:
        raise ApiError("通知不存在", 404)
    db.commit()
    unread_count = db.execute(
        "SELECT COUNT(*) AS count FROM notifications "
        "WHERE workspace_id=? AND user_id=? AND read_at IS NULL",
        (current_workspace_id(), g.user["id"]),
    ).fetchone()["count"]
    return jsonify({"ok": True, "unread_count": unread_count})


@app.route("/api/tasks/<int:task_id>/collaboration")
def task_collaboration(task_id):
    db = get_db()
    workspace_id = current_workspace_id()
    active_task(db, workspace_id, task_id)
    followers = [dict(row) for row in db.execute(
        "SELECT u.id,u.username,u.display_name FROM task_followers f "
        "JOIN users u ON u.id=f.user_id "
        "JOIN workspace_members m ON m.workspace_id=f.workspace_id "
        "AND m.user_id=f.user_id "
        "WHERE f.workspace_id=? AND f.task_id=? AND u.active=1 "
        "ORDER BY u.display_name,u.id",
        (workspace_id, task_id),
    )]
    timeline = [dict(row) for row in db.execute(
        "SELECT 'comment' AS kind,c.id,c.content AS detail,'' AS summary,"
        "c.created_at,u.id AS actor_id,u.display_name AS actor_name "
        "FROM task_comments c JOIN users u ON u.id=c.author_id "
        "WHERE c.workspace_id=? AND c.task_id=? "
        "UNION ALL "
        "SELECT 'activity' AS kind,a.id,'' AS detail,a.summary,"
        "a.created_at,u.id AS actor_id,"
        "COALESCE(u.display_name,'系统') AS actor_name "
        "FROM task_activities a LEFT JOIN users u ON u.id=a.actor_id "
        "WHERE a.workspace_id=? AND a.task_id=? "
        "ORDER BY 5 DESC,2 DESC LIMIT 100",
        (workspace_id, task_id, workspace_id, task_id),
    )]
    return jsonify({
        "following": any(row["id"] == g.user["id"] for row in followers),
        "followers": followers,
        "members": [dict(row) for row in workspace_member_rows(db, workspace_id)],
        "timeline": timeline,
    })


@app.route("/api/tasks/<int:task_id>/follow", methods=["POST"])
def follow_task(task_id):
    db = get_db()
    workspace_id = current_workspace_id()
    active_task(db, workspace_id, task_id)
    db.execute(
        "INSERT OR IGNORE INTO task_followers(workspace_id,task_id,user_id,created_at) "
        "VALUES(?,?,?,?)", (workspace_id, task_id, g.user["id"], now_iso()),
    )
    db.commit()
    return jsonify({"ok": True, "following": True})


@app.route("/api/tasks/<int:task_id>/follow", methods=["DELETE"])
def unfollow_task(task_id):
    db = get_db()
    workspace_id = current_workspace_id()
    active_task(db, workspace_id, task_id)
    db.execute(
        "DELETE FROM task_followers WHERE workspace_id=? AND task_id=? AND user_id=?",
        (workspace_id, task_id, g.user["id"]),
    )
    db.commit()
    return jsonify({"ok": True, "following": False})


@app.route("/api/tasks/<int:task_id>/comments", methods=["POST"])
def add_task_comment(task_id):
    data = json_object()
    content = text_field(data, "content", "评论内容").strip()
    if not content:
        raise ApiError("评论内容不能为空")
    if len(content) > 2000:
        raise ApiError("评论内容不能超过 2000 个字符")
    db = get_db()
    workspace_id = current_workspace_id()
    task = active_task(db, workspace_id, task_id)
    created_at = now_iso()
    cur = db.execute(
        "INSERT INTO task_comments(workspace_id,task_id,author_id,content,created_at) "
        "VALUES(?,?,?,?,?)",
        (workspace_id, task_id, g.user["id"], content, created_at),
    )
    db.execute(
        "INSERT OR IGNORE INTO task_followers(workspace_id,task_id,user_id,created_at) "
        "VALUES(?,?,?,?)", (workspace_id, task_id, g.user["id"], created_at),
    )
    mentioned = mentioned_user_ids(db, workspace_id, content) - {g.user["id"]}
    audience = task_audience_user_ids(
        db, workspace_id, task_id, task["owner"]
    ) - mentioned
    actor_name = g.user["display_name"]
    add_notifications(
        db, workspace_id, mentioned, "mention",
        f"{actor_name} 在事务「{task['name']}」中提到了你",
        task_id, g.user["id"],
    )
    add_notifications(
        db, workspace_id, audience, "comment",
        f"{actor_name} 评论了事务「{task['name']}」",
        task_id, g.user["id"],
    )
    db.commit()
    return jsonify({"id": cur.lastrowid, "created_at": created_at}), 201


@app.route("/api/task-images/<int:image_id>")
def task_image(image_id):
    db = get_db()
    workspace_id = current_workspace_id()
    row = db.execute(
        "SELECT image.mime_type,image.data FROM task_images image "
        "JOIN tasks task ON task.id=image.task_id "
        "JOIN lines line ON line.id=task.line_id "
        "WHERE image.id=? AND image.workspace_id=? AND task.workspace_id=? "
        "AND line.workspace_id=? AND task.deleted=0 AND line.deleted=0",
        (image_id, workspace_id, workspace_id, workspace_id),
    ).fetchone()
    if not row:
        raise ApiError("事务图片不存在", 404)
    response = send_file(
        BytesIO(row["data"]), mimetype=row["mime_type"],
        as_attachment=False, max_age=0,
    )
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Cache-Control"] = "private, no-store"
    return response


@app.route("/api/task-attachments/<int:attachment_id>")
def task_attachment(attachment_id):
    db = get_db()
    workspace_id = current_workspace_id()
    row = db.execute(
        "SELECT attachment.filename,attachment.mime_type,attachment.data "
        "FROM task_attachments attachment "
        "JOIN tasks task ON task.id=attachment.task_id "
        "JOIN lines line ON line.id=task.line_id "
        "WHERE attachment.id=? AND attachment.workspace_id=? "
        "AND task.workspace_id=? AND line.workspace_id=? "
        "AND task.deleted=0 AND line.deleted=0",
        (attachment_id, workspace_id, workspace_id, workspace_id),
    ).fetchone()
    if not row:
        raise ApiError("事务附件不存在", 404)
    response = send_file(
        BytesIO(row["data"]), mimetype=row["mime_type"], as_attachment=True,
        download_name=row["filename"], max_age=0,
    )
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Cache-Control"] = "private, no-store"
    return response


# ----- owner options (directly derived from current workspace members)
def get_workspace_member_names(db, workspace_id):
    rows = db.execute(
        "SELECT u.display_name FROM workspace_members m "
        "JOIN users u ON u.id=m.user_id "
        "WHERE m.workspace_id=? AND u.active=1 "
        "ORDER BY CASE m.role WHEN 'admin' THEN 0 ELSE 1 END,u.display_name,u.id",
        (workspace_id,),
    )
    names, seen = [], set()
    for row in rows:
        name = row["display_name"].strip()
        if name and name not in seen:
            seen.add(name)
            names.append(name)
    return names


# ----- statuses (进展状态配置)
@app.route("/api/statuses", methods=["GET"])
def api_get_statuses():
    db = get_db()
    statuses = get_statuses(db)
    return jsonify({
        "statuses": statuses,
        "colors": get_status_colors(db, statuses),
    })


@app.route("/api/statuses", methods=["PUT"])
def api_set_statuses():
    d = json_object()
    statuses = d.get("statuses")
    if not isinstance(statuses, list) or \
            not all(isinstance(s, str) for s in statuses):
        return jsonify({"error": "statuses 必须是字符串数组"}), 400
    cleaned, seen = [], set()
    for s in statuses:
        s = s.strip()
        if s and s not in seen:
            seen.add(s)
            cleaned.append(s)
    if not cleaned:
        return jsonify({"error": "进展状态不能为空"}), 400
    db = get_db()
    colors = d.get("colors", {})
    if not isinstance(colors, dict) or \
            not all(isinstance(k, str) and isinstance(v, str)
                    for k, v in colors.items()):
        return jsonify({"error": "colors 必须是状态名到颜色的对象"}), 400
    unknown_colors = set(colors) - set(cleaned)
    if unknown_colors:
        return jsonify({"error": "颜色配置包含未知状态"}), 400
    normalized_colors = {}
    for status, color in colors.items():
        normalized = line_color(color)
        if normalized is None:
            return jsonify({"error": "状态颜色不能为空"}), 400
        normalized_colors[status] = normalized

    previous_colors = get_status_colors(db)
    set_meta(db, "statuses", json.dumps(cleaned, ensure_ascii=False))
    effective_statuses = get_statuses(db)
    effective_colors = {
        status: normalized_colors.get(
            status,
            previous_colors.get(
                status, DEFAULT_STATUS_COLORS.get(status, FALLBACK_STATUS_COLOR)
            ),
        )
        for status in effective_statuses
    }
    set_meta(db, "status_colors", json.dumps(effective_colors, ensure_ascii=False))
    db.commit()
    return jsonify({
        "ok": True,
        "statuses": effective_statuses,
        "colors": effective_colors,
    })


# ----- unified Excel import/export
@app.route("/api/data/import-template")
def download_data_import_template():
    output = data_import_template_workbook(get_db(), current_workspace_id())
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"AnyLine-数据导入模板-{date.today():%Y%m%d}.xlsx",
    )


@app.route("/api/data/import", methods=["POST"])
def import_data():
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        raise ApiError("请选择要导入的 Excel 文件")
    if not uploaded.filename.lower().endswith(".xlsx"):
        raise ApiError("仅支持 .xlsx 格式的 Excel 文件")
    content = uploaded.read(MAX_TASK_IMPORT_BYTES + 1)
    if not content:
        raise ApiError("导入文件不能为空")
    if len(content) > MAX_TASK_IMPORT_BYTES:
        raise ApiError("导入文件不能超过 5MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError):
        raise ApiError("无法读取 Excel 文件，请使用下载的导入模板")
    db = get_db()
    workspace_id = current_workspace_id()
    try:
        if not ({"线导入", "事务导入"} & set(workbook.sheetnames)):
            raise ApiError("Excel 文件至少需要“线导入”或“事务导入”工作表")
        line_rows, line_errors, line_count = ([], [], 0)
        if "线导入" in workbook.sheetnames:
            line_rows, line_errors, line_count = parse_line_import_sheet(
                workbook["线导入"]
            )
        imported_lines = imported_line_records(line_rows) if not line_errors else []
        task_rows, task_errors, task_count = ([], [], 0)
        if "事务导入" in workbook.sheetnames:
            task_rows, task_errors, task_count = parse_task_import_sheet(
                workbook["事务导入"], db, workspace_id, imported_lines
            )
    finally:
        workbook.close()
    if not line_count and not task_count:
        raise ApiError("导入工作表中没有可导入的数据")
    row_errors = [
        {**item, "sheet": "线导入"} for item in line_errors
    ] + [
        {**item, "sheet": "事务导入"} for item in task_errors
    ]
    if row_errors:
        return jsonify({
            "error": f"导入文件存在 {len(row_errors)} 行错误，未导入任何数据",
            "error_count": len(row_errors),
            "row_errors": row_errors[:50],
        }), 400

    today = date.today().isoformat()
    on_edit(db)
    ids_by_key = {}
    imported_line_ids = []
    for line in line_rows:
        cur = db.execute(
            "INSERT INTO lines(workspace_id,name,description,color,parent_id,fork_date,"
            "merge_date,updated_at) VALUES(?,?,?,?,?,?,?,?)",
            (
                workspace_id, line["name"], line["description"], line["color"],
                ids_by_key.get(line["_parent_key"]), line["fork_date"],
                line["merge_date"], today,
            ),
        )
        ids_by_key[line["_key"]] = cur.lastrowid
        imported_line_ids.append(cur.lastrowid)
    imported_task_ids = []
    for task in task_rows:
        line_id = ids_by_key.get(task["_line_import_key"], task["line_id"])
        cur = db.execute(
            "INSERT INTO tasks(workspace_id,line_id,name,content,goal,owner,priority,"
            "next_action,risk_reason,status,start_date,end_date,status_since,updated_at) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                workspace_id, line_id, task["name"], task["content"], task["goal"],
                task["owner"], task["priority"], task["next_action"],
                task["risk_reason"], task["status"], task["start_date"],
                task["end_date"], today, today,
            ),
        )
        imported_task_ids.append(cur.lastrowid)
        add_task_activity(
            db, workspace_id, cur.lastrowid, "imported", "通过 Excel 导入了事务"
        )
    db.commit()
    return jsonify({
        "ok": True,
        "count": len(imported_line_ids) + len(imported_task_ids),
        "line_count": len(imported_line_ids),
        "task_count": len(imported_task_ids),
        "line_ids": imported_line_ids,
        "task_ids": imported_task_ids,
        "can_undo": True,
    }), 201


@app.route("/api/data/export", methods=["POST"])
def export_data():
    d = json_object()
    scope = d.get("scope")
    workspace_id = current_workspace_id()
    params = [workspace_id, workspace_id]
    selected_count = None
    id_clause = ""
    if scope == "selected":
        ids = d.get("ids")
        if not isinstance(ids, list) or not ids or not all(
            isinstance(item, int) and not isinstance(item, bool) for item in ids
        ):
            raise ApiError("ids 必须是非空整数数组")
        if len(ids) != len(set(ids)):
            raise ApiError("ids 不能包含重复项")
        id_clause = f" AND t.id IN ({','.join('?' * len(ids))})"
        params.extend(ids)
        selected_count = len(ids)
    elif scope != "all":
        raise ApiError("scope 必须是 all 或 selected")

    db = get_db()
    task_rows = [dict(row) for row in db.execute(
        "SELECT t.id,t.line_id,t.name,t.content,t.goal,t.next_action,t.risk_reason,"
        "t.priority,t.owner,t.status,t.start_date,t.end_date,t.status_since,t.updated_at "
        "FROM tasks t JOIN lines l ON l.id=t.line_id "
        "WHERE t.workspace_id=? AND l.workspace_id=? AND t.deleted=0 AND l.deleted=0" +
        id_clause + " ORDER BY t.start_date,t.id",
        params,
    ).fetchall()]
    if scope == "selected" and len(task_rows) != selected_count:
        raise ApiError("部分事务不存在或已删除", 404)

    all_lines = workspace_line_records(db, workspace_id)
    line_by_id = {line["id"]: line for line in all_lines}
    if scope == "all":
        line_rows = all_lines
    else:
        included_ids = {task["line_id"] for task in task_rows}
        pending = list(included_ids)
        while pending:
            parent_id = line_by_id[pending.pop()].get("parent_id")
            if parent_id is not None and parent_id not in included_ids:
                included_ids.add(parent_id)
                pending.append(parent_id)
        line_rows = [line for line in all_lines if line["id"] in included_ids]
    for task in task_rows:
        task["line_path"] = line_by_id[task["line_id"]]["path"]
    output = data_export_workbook(line_rows, task_rows)
    scope_name = "全部数据" if scope == "all" else "选中事务及关联线"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"AnyLine-{scope_name}-{date.today():%Y%m%d}.xlsx",
    )


# ----- lines
@app.route("/api/lines/import-template")
def download_line_import_template():
    output = line_import_template_workbook()
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"AnyLine-主线支线导入模板-{date.today():%Y%m%d}.xlsx",
    )


@app.route("/api/lines/import", methods=["POST"])
def import_lines():
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        raise ApiError("请选择要导入的 Excel 文件")
    if not uploaded.filename.lower().endswith(".xlsx"):
        raise ApiError("仅支持 .xlsx 格式的 Excel 文件")
    content = uploaded.read(MAX_TASK_IMPORT_BYTES + 1)
    if not content:
        raise ApiError("导入文件不能为空")
    if len(content) > MAX_TASK_IMPORT_BYTES:
        raise ApiError("导入文件不能超过 5MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError):
        raise ApiError("无法读取 Excel 文件，请使用下载的导入模板")
    try:
        if "线导入" not in workbook.sheetnames:
            raise ApiError("Excel 文件缺少“线导入”工作表")
        rows, row_errors, data_count = parse_line_import_sheet(workbook["线导入"])
    finally:
        workbook.close()
    if not data_count:
        raise ApiError("“线导入”工作表中没有可导入的数据")
    if row_errors:
        return jsonify({
            "error": f"导入文件存在 {len(row_errors)} 行错误，未导入任何主线或支线",
            "error_count": len(row_errors),
            "row_errors": row_errors[:50],
        }), 400

    db = get_db()
    workspace_id = current_workspace_id()
    today = date.today().isoformat()
    on_edit(db)
    ids_by_key = {}
    imported_ids = []
    for line in rows:
        parent_id = ids_by_key.get(line["_parent_key"])
        cur = db.execute(
            "INSERT INTO lines(workspace_id,name,description,color,parent_id,fork_date,"
            "merge_date,updated_at) VALUES(?,?,?,?,?,?,?,?)",
            (
                workspace_id, line["name"], line["description"], line["color"],
                parent_id, line["fork_date"], line["merge_date"], today,
            ),
        )
        ids_by_key[line["_key"]] = cur.lastrowid
        imported_ids.append(cur.lastrowid)
    db.commit()
    return jsonify({
        "ok": True,
        "count": len(imported_ids),
        "ids": imported_ids,
        "can_undo": True,
    }), 201


@app.route("/api/lines/export")
def export_lines():
    rows = workspace_line_records(get_db(), current_workspace_id())
    output = line_export_workbook(rows)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"AnyLine-主线与支线-{date.today():%Y%m%d}.xlsx",
    )


@app.route("/api/lines", methods=["POST"])
def create_line():
    d = json_object()
    name = text_field(d, "name", "线名").strip()
    description = text_field(d, "description", "描述").strip()
    color = line_color(d.get("color"))
    if not name:
        return jsonify({"error": "线名不能为空"}), 400
    parent_id = d.get("parent_id")
    if parent_id is not None:
        required_id(parent_id, "parent_id")
    fork_date = text_field(
        d, "fork_date", "起始日期", date.today().isoformat()
    ) or date.today().isoformat()
    if not fork_date:
        return jsonify({"error": "起始日期不能为空"}), 400
    try:
        parse_iso_date(fork_date, "起始日期")
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    db = get_db()
    workspace_id = current_workspace_id()
    if parent_id is not None:
        p = db.execute(
            "SELECT id,fork_date FROM lines WHERE id=? AND workspace_id=? "
            "AND deleted=0", (parent_id, workspace_id)
        ).fetchone()
        if not p:
            return jsonify({"error": "父线不存在"}), 404
        if fork_date < p["fork_date"]:
            return jsonify({"error": "支线起始日期不能早于父线起始日期"}), 400
    on_edit(db)
    cur = db.execute(
        "INSERT INTO lines(workspace_id,name,description,color,parent_id,fork_date,updated_at) "
        "VALUES(?,?,?,?,?,?,?)",
        (
            workspace_id, name, description, color, parent_id, fork_date,
            date.today().isoformat(),
        ),
    )
    db.commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.route("/api/lines/<int:lid>", methods=["PATCH"])
def update_line(lid):
    d = json_object()
    if d.get("merge_date") == "":
        d["merge_date"] = None
    db = get_db()
    workspace_id = current_workspace_id()
    row = db.execute(
        "SELECT * FROM lines WHERE id=? AND workspace_id=? AND deleted=0",
        (lid, workspace_id),
    ).fetchone()
    if not row:
        return jsonify({"error": "线不存在"}), 404
    new_name = text_field(d, "name", "线名", row["name"])
    new_description = text_field(
        d, "description", "描述", row["description"] or ""
    )
    new_color = line_color(d.get("color", row["color"]))
    new_fork = text_field(d, "fork_date", "起始日期", row["fork_date"])
    new_merge = text_field(
        d, "merge_date", "反合日期", row["merge_date"], nullable=True
    )
    if "name" in d and not (new_name or "").strip():
        return jsonify({"error": "线名不能为空"}), 400
    if "name" in d:
        d["name"] = new_name.strip()
    if "description" in d:
        d["description"] = new_description.strip()
    if "color" in d:
        d["color"] = new_color
    if not new_fork:
        return jsonify({"error": "起始日期不能为空"}), 400
    try:
        parse_iso_date(new_fork, "起始日期")
        parse_iso_date(new_merge, "反合日期")
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if new_merge is not None and row["parent_id"] is None:
        return jsonify({"error": "主线不能反合"}), 400
    if new_merge is not None and new_merge < new_fork:
        return jsonify({"error": "反合日期不能早于支线起始日期"}), 400
    if row["parent_id"] is not None:
        parent = db.execute(
            "SELECT fork_date FROM lines WHERE id=? AND workspace_id=? AND deleted=0",
            (row["parent_id"], workspace_id),
        ).fetchone()
        if parent and new_fork < parent["fork_date"]:
            return jsonify({"error": "支线起始日期不能早于父线起始日期"}), 400
    child = db.execute(
        "SELECT MIN(fork_date) AS first_date FROM lines "
        "WHERE parent_id=? AND workspace_id=? AND deleted=0", (lid, workspace_id)
    ).fetchone()
    if child["first_date"] and child["first_date"] < new_fork:
        return jsonify({"error": "起始日期不能晚于子支线的起始日期"}), 400
    task = db.execute(
        "SELECT MIN(start_date) AS first_date FROM tasks "
        "WHERE line_id=? AND workspace_id=? AND deleted=0", (lid, workspace_id)
    ).fetchone()
    if task["first_date"] and task["first_date"] < new_fork:
        return jsonify({"error": "起始日期不能晚于线上已有事务的起始日期"}), 400
    milestone = db.execute(
        "SELECT MIN(milestone_date) AS first_date FROM milestones "
        "WHERE line_id=? AND workspace_id=? AND deleted=0", (lid, workspace_id)
    ).fetchone()
    if milestone["first_date"] and milestone["first_date"] < new_fork:
        return jsonify({"error": "起始日期不能晚于线上已有里程碑日期"}), 400

    fields, vals = [], []
    for k in ("name", "description", "color", "fork_date", "merge_date"):
        if k in d:
            fields.append(f"{k}=?")
            vals.append(d[k])
    if fields:
        on_edit(db)
        fields.append("updated_at=?")
        vals.append(date.today().isoformat())
        vals.append(lid)
        vals.append(workspace_id)
        db.execute(
            f"UPDATE lines SET {','.join(fields)} WHERE id=? AND workspace_id=?", vals
        )
        db.commit()
    return jsonify({"ok": True})


def collect_descendants(db, lid, workspace_id):
    ids = [lid]
    i = 0
    while i < len(ids):
        rows = db.execute(
            "SELECT id FROM lines WHERE parent_id=? AND workspace_id=? AND deleted=0",
            (ids[i], workspace_id),
        ).fetchall()
        ids.extend(r["id"] for r in rows)
        i += 1
    return ids


@app.route("/api/lines/<int:lid>", methods=["DELETE"])
def delete_line(lid):
    db = get_db()
    workspace_id = current_workspace_id()
    row = db.execute(
        "SELECT id FROM lines WHERE id=? AND workspace_id=? AND deleted=0",
        (lid, workspace_id),
    ).fetchone()
    if not row:
        return jsonify({"error": "线不存在"}), 404
    ids = collect_descendants(db, lid, workspace_id)
    marks = ",".join("?" * len(ids))
    task_ids = [row["id"] for row in db.execute(
        f"SELECT id FROM tasks WHERE workspace_id=? AND deleted=0 "
        f"AND line_id IN ({marks})",
        [workspace_id] + ids,
    )]
    milestone_ids = [row["id"] for row in db.execute(
        f"SELECT id FROM milestones WHERE workspace_id=? AND deleted=0 "
        f"AND line_id IN ({marks})",
        [workspace_id] + ids,
    )]
    ensure_tasks_not_required(
        db, workspace_id, task_ids, excluded_milestone_ids=milestone_ids
    )
    on_edit(db)
    batch = new_batch(db)
    today = date.today().isoformat()
    db.execute(
        f"UPDATE lines SET deleted=1, del_batch=?, deleted_at=? "
        f"WHERE workspace_id=? AND id IN ({marks})",
        [batch, today, workspace_id] + ids,
    )
    db.execute(
        f"UPDATE tasks SET deleted=1, del_batch=?, deleted_at=? WHERE deleted=0 "
        f"AND workspace_id=? AND line_id IN ({marks})",
        [batch, today, workspace_id] + ids,
    )
    db.execute(
        f"UPDATE milestones SET deleted=1, del_batch=?, deleted_at=? "
        f"WHERE deleted=0 AND workspace_id=? AND line_id IN ({marks})",
        [batch, today, workspace_id] + ids,
    )
    db.commit()
    return jsonify({"ok": True, "can_undo": True})


# ----- milestones
@app.route("/api/milestones", methods=["POST"])
def create_milestone():
    d = json_object()
    name = text_field(d, "name", "里程碑名称").strip()
    target_description = text_field(
        d, "target_description", "里程碑目标描述"
    ).strip()
    milestone_date = text_field(d, "milestone_date", "里程碑日期").strip()
    if not name:
        raise ApiError("里程碑名称不能为空")
    if not target_description:
        raise ApiError("里程碑目标描述不能为空")
    if not milestone_date:
        raise ApiError("里程碑日期不能为空")
    try:
        parse_iso_date(milestone_date, "里程碑日期")
    except ValueError as error:
        raise ApiError(str(error))
    line_id = required_id(d.get("line_id"), "line_id")
    db = get_db()
    workspace_id = current_workspace_id()
    line = db.execute(
        "SELECT id,fork_date FROM lines WHERE id=? AND workspace_id=? AND deleted=0",
        (line_id, workspace_id),
    ).fetchone()
    if not line:
        raise ApiError("所属线不存在", 404)
    if milestone_date < line["fork_date"]:
        raise ApiError("里程碑日期不能早于所属线的起始日期")
    task_ids = validate_milestone_tasks(
        db, workspace_id, d.get("acceptance_task_ids", [])
    )
    today = date.today().isoformat()
    on_edit(db)
    cursor = db.execute(
        "INSERT INTO milestones(workspace_id,line_id,name,target_description,"
        "milestone_date,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
        (workspace_id, line_id, name, target_description, milestone_date, today, today),
    )
    replace_milestone_tasks(db, workspace_id, cursor.lastrowid, task_ids)
    db.commit()
    return jsonify({"id": cursor.lastrowid}), 201


@app.route("/api/milestones/<int:milestone_id>", methods=["PATCH"])
def update_milestone(milestone_id):
    d = json_object()
    db = get_db()
    workspace_id = current_workspace_id()
    milestone = db.execute(
        "SELECT * FROM milestones WHERE id=? AND workspace_id=? AND deleted=0",
        (milestone_id, workspace_id),
    ).fetchone()
    if not milestone:
        raise ApiError("里程碑不存在", 404)
    name = text_field(d, "name", "里程碑名称", milestone["name"]).strip()
    target_description = text_field(
        d, "target_description", "里程碑目标描述",
        milestone["target_description"],
    ).strip()
    milestone_date = text_field(
        d, "milestone_date", "里程碑日期", milestone["milestone_date"]
    ).strip()
    if not name:
        raise ApiError("里程碑名称不能为空")
    if not target_description:
        raise ApiError("里程碑目标描述不能为空")
    if not milestone_date:
        raise ApiError("里程碑日期不能为空")
    try:
        parse_iso_date(milestone_date, "里程碑日期")
    except ValueError as error:
        raise ApiError(str(error))
    line = db.execute(
        "SELECT fork_date FROM lines WHERE id=? AND workspace_id=? AND deleted=0",
        (milestone["line_id"], workspace_id),
    ).fetchone()
    if not line:
        raise ApiError("所属线不存在", 404)
    if milestone_date < line["fork_date"]:
        raise ApiError("里程碑日期不能早于所属线的起始日期")
    if "acceptance_task_ids" in d:
        task_ids = validate_milestone_tasks(
            db, workspace_id, d["acceptance_task_ids"]
        )
    else:
        task_ids = [row["task_id"] for row in db.execute(
            "SELECT task_id FROM milestone_tasks WHERE workspace_id=? "
            "AND milestone_id=? ORDER BY task_id",
            (workspace_id, milestone_id),
        )]
    on_edit(db)
    db.execute(
        "UPDATE milestones SET name=?,target_description=?,milestone_date=?,"
        "updated_at=? WHERE id=? AND workspace_id=?",
        (name, target_description, milestone_date, date.today().isoformat(),
         milestone_id, workspace_id),
    )
    replace_milestone_tasks(db, workspace_id, milestone_id, task_ids)
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/auth/avatar")
def auth_avatar():
    db = get_db()
    row = db.execute(
        "SELECT avatar_mime,avatar_data FROM users WHERE id=?", (g.user["id"],),
    ).fetchone()
    if not row or not row["avatar_data"]:
        raise ApiError("尚未设置自定义头像", 404)
    response = send_file(
        BytesIO(row["avatar_data"]), mimetype=row["avatar_mime"],
        as_attachment=False, max_age=0,
    )
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Cache-Control"] = "private, no-store"
    return response


@app.route("/api/auth/avatar", methods=["PUT"])
def auth_avatar_update():
    mime_type, avatar_data = prepare_avatar_data(json_object().get("data_url"))
    changed_at = now_iso()
    db = get_db()
    db.execute(
        "UPDATE users SET avatar_mime=?,avatar_data=?,avatar_updated_at=?,updated_at=? "
        "WHERE id=?",
        (mime_type, avatar_data, changed_at, date.today().isoformat(), g.user["id"]),
    )
    db.commit()
    return jsonify({"avatar_url": f"/api/auth/avatar?v={changed_at}"})


@app.route("/api/milestones/<int:milestone_id>", methods=["DELETE"])
def delete_milestone(milestone_id):
    db = get_db()
    workspace_id = current_workspace_id()
    milestone = db.execute(
        "SELECT id FROM milestones WHERE id=? AND workspace_id=? AND deleted=0",
        (milestone_id, workspace_id),
    ).fetchone()
    if not milestone:
        raise ApiError("里程碑不存在", 404)
    on_edit(db)
    batch = new_batch(db)
    db.execute(
        "UPDATE milestones SET deleted=1,del_batch=?,deleted_at=? "
        "WHERE id=? AND workspace_id=?",
        (batch, date.today().isoformat(), milestone_id, workspace_id),
    )
    db.commit()
    return jsonify({"ok": True, "can_undo": True})


# ----- tasks
@app.route("/api/tasks", methods=["POST"])
def create_task():
    d = json_object()
    name = text_field(d, "name", "事务名").strip()
    if not name:
        return jsonify({"error": "事务名不能为空"}), 400
    db = get_db()
    workspace_id = current_workspace_id()
    line_id = required_id(d.get("line_id"), "line_id")
    line = db.execute(
        "SELECT id,fork_date FROM lines WHERE id=? AND workspace_id=? AND deleted=0",
        (line_id, workspace_id),
    ).fetchone()
    if not line:
        return jsonify({"error": "所属线不存在"}), 404
    status = text_field(d, "status", "进展状态").strip()
    if not status:
        raise ApiError("进展状态不能为空")
    if status not in get_statuses(db):
        return jsonify({"error": "非法的进展状态"}), 400
    priority = text_field(d, "priority", "优先级", "中") or "中"
    if priority not in PRIORITY_ENUM:
        return jsonify({"error": "非法的优先级"}), 400
    today = date.today().isoformat()
    start_date = text_field(d, "start_date", "起始日期").strip()
    end_date = text_field(d, "end_date", "结束日期").strip()
    content = text_field(d, "content", "事务内容")
    goal = text_field(d, "goal", "闭环目标")
    owner = text_field(d, "owner", "责任人").strip()
    next_action = text_field(d, "next_action", "下一步动作")
    risk_reason = text_field(d, "risk_reason", "风险原因")
    for label, value in (("事务内容", content), ("责任人", owner),
                         ("起始日期", start_date), ("结束日期", end_date)):
        if not value.strip():
            raise ApiError(f"{label}不能为空")
    if owner not in get_workspace_member_names(db, workspace_id):
        raise ApiError("责任人不是当前项目空间成员")
    prerequisite_ids = validate_dependencies(
        db, workspace_id, None, d.get("prerequisite_ids", [])
    )
    image_changes = validate_task_images(
        db, workspace_id, None, d.get("images", [])
    )
    attachment_changes = validate_task_attachments(
        db, workspace_id, None, d.get("attachments", [])
    )
    if status == "已闭环":
        ensure_dependencies_closed(db, workspace_id, prerequisite_ids)
    try:
        validate_date_range(start_date, end_date)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if start_date < line["fork_date"]:
        return jsonify({"error": "事务起始日期不能早于所属线起始日期"}), 400
    on_edit(db)
    cur = db.execute(
        "INSERT INTO tasks(workspace_id,line_id,name,content,goal,owner,priority,next_action,"
        "risk_reason,status,start_date,end_date,status_since,updated_at) "
        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            workspace_id, line_id, name, content, goal, owner, priority,
            next_action, risk_reason,
            status, start_date, end_date, today,
            today,
        ),
    )
    replace_task_dependencies(
        db, workspace_id, cur.lastrowid, prerequisite_ids
    )
    replace_task_images(db, workspace_id, cur.lastrowid, image_changes)
    replace_task_attachments(
        db, workspace_id, cur.lastrowid, attachment_changes
    )
    db.execute(
        "INSERT OR IGNORE INTO task_followers(workspace_id,task_id,user_id,created_at) "
        "VALUES(?,?,?,?)", (workspace_id, cur.lastrowid, g.user["id"], now_iso()),
    )
    add_task_activity(
        db, workspace_id, cur.lastrowid, "created", "创建了事务"
    )
    add_notifications(
        db, workspace_id, owner_user_ids(db, workspace_id, owner), "assigned",
        f"{g.user['display_name']} 将事务「{name}」指派给你",
        cur.lastrowid, g.user["id"],
    )
    db.commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.route("/api/tasks/<int:tid>", methods=["PATCH"])
def update_task(tid):
    d = json_object()
    if "end_date" in d and (
        d["end_date"] is None or
        isinstance(d["end_date"], str) and not d["end_date"].strip()
    ):
        raise ApiError("结束日期不能为空")
    db = get_db()
    workspace_id = current_workspace_id()
    row = db.execute(
        "SELECT * FROM tasks WHERE id=? AND workspace_id=? AND deleted=0",
        (tid, workspace_id),
    ).fetchone()
    if not row:
        return jsonify({"error": "事务不存在"}), 404
    new_start = text_field(d, "start_date", "起始日期", row["start_date"])
    new_end = text_field(
        d, "end_date", "结束日期", row["end_date"], nullable=True
    )
    new_line_id = d.get("line_id", row["line_id"])
    required_id(new_line_id, "line_id")
    if not new_start:
        return jsonify({"error": "起始日期不能为空"}), 400
    if not new_end:
        return jsonify({"error": "结束日期不能为空"}), 400
    try:
        validate_date_range(new_start, new_end)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    target_line = db.execute(
        "SELECT fork_date FROM lines WHERE id=? AND workspace_id=? AND deleted=0",
        (new_line_id, workspace_id),
    ).fetchone()
    if not target_line:
        return jsonify({"error": "所属线不存在"}), 404
    if new_start < target_line["fork_date"]:
        return jsonify({"error": "事务起始日期不能早于所属线起始日期"}), 400

    prerequisite_ids = None
    if "prerequisite_ids" in d:
        prerequisite_ids = validate_dependencies(
            db, workspace_id, tid, d["prerequisite_ids"]
        )
    image_changes = None
    if "images" in d:
        image_changes = validate_task_images(
            db, workspace_id, tid, d["images"]
        )
    attachment_changes = None
    if "attachments" in d:
        attachment_changes = validate_task_attachments(
            db, workspace_id, tid, d["attachments"]
        )

    fields, vals = [], []
    changed_labels = []
    field_labels = {
        "line_id": "所属线", "name": "事务名", "content": "事务内容",
        "goal": "闭环目标", "owner": "责任人", "priority": "优先级",
        "next_action": "下一步动作", "risk_reason": "风险原因",
        "status": "进展状态", "start_date": "起始日期", "end_date": "结束日期",
    }
    for k in ("line_id", "name", "content", "goal", "owner", "priority",
              "next_action", "risk_reason", "status", "start_date", "end_date"):
        if k not in d:
            continue
        if k == "line_id":
            required_id(d[k], "line_id")
        else:
            d[k] = text_field(
                d, k,
                {"name": "事务名", "content": "事务内容", "goal": "闭环目标",
                 "owner": "责任人", "priority": "优先级", "next_action": "下一步动作",
                 "risk_reason": "风险原因", "status": "进展状态",
                 "start_date": "起始日期", "end_date": "结束日期"}[k],
                nullable=(k == "end_date"),
            )
        if k == "name" and not (d[k] or "").strip():
            return jsonify({"error": "事务名不能为空"}), 400
        if k in ("content", "owner", "status", "start_date", "end_date") and \
                not (d[k] or "").strip():
            label = {"content": "事务内容", "owner": "责任人",
                     "status": "进展状态", "start_date": "起始日期",
                     "end_date": "结束日期"}[k]
            raise ApiError(f"{label}不能为空")
        if k == "owner":
            d[k] = d[k].strip()
            if d[k] not in get_workspace_member_names(db, workspace_id):
                raise ApiError("责任人不是当前项目空间成员")
        if k == "name":
            d[k] = d[k].strip()
        if k == "priority" and d[k] not in PRIORITY_ENUM:
            return jsonify({"error": "非法的优先级"}), 400
        if k == "status":
            if d[k] not in get_statuses(db):
                return jsonify({"error": "非法的进展状态"}), 400
            if d[k] != row["status"]:   # 状态变化 -> 重新计时
                fields.append("status_since=?")
                vals.append(date.today().isoformat())
        if d[k] != row[k]:
            changed_labels.append(field_labels[k])
        fields.append(f"{k}=?")
        vals.append(d[k])
    final_status = d.get("status", row["status"])
    final_prerequisite_ids = prerequisite_ids
    if final_prerequisite_ids is None:
        final_prerequisite_ids = current_dependency_ids(db, workspace_id, tid)
    if final_status == "已闭环":
        ensure_dependencies_closed(db, workspace_id, final_prerequisite_ids)

    old_prerequisite_ids = current_dependency_ids(db, workspace_id, tid)
    dependencies_changed = prerequisite_ids is not None and set(
        prerequisite_ids
    ) != set(old_prerequisite_ids)
    was_blocked = False
    if dependencies_changed:
        was_blocked = db.execute(
            "SELECT 1 FROM task_dependencies edge "
            "JOIN tasks prerequisite ON prerequisite.id=edge.prerequisite_task_id "
            "WHERE edge.workspace_id=? AND edge.dependent_task_id=? "
            "AND prerequisite.deleted=0 "
            "AND prerequisite.status NOT IN ('已闭环','已取消') LIMIT 1",
            (workspace_id, tid),
        ).fetchone() is not None
    if (fields or prerequisite_ids is not None or image_changes is not None or
            attachment_changes is not None):
        on_edit(db)
        if fields:
            fields.append("updated_at=?")
            vals.append(date.today().isoformat())
            vals.append(tid)
            vals.append(workspace_id)
            db.execute(
                f"UPDATE tasks SET {','.join(fields)} WHERE id=? AND workspace_id=?", vals
            )
        if prerequisite_ids is not None:
            replace_task_dependencies(db, workspace_id, tid, prerequisite_ids)
        if image_changes is not None:
            replace_task_images(db, workspace_id, tid, image_changes)
        if attachment_changes is not None:
            replace_task_attachments(
                db, workspace_id, tid, attachment_changes
            )
        if dependencies_changed:
            changed_labels.append("前置依赖")
            if was_blocked:
                notify_task_if_unblocked(
                    db, workspace_id, tid, g.user["id"],
                    "前置依赖调整后，事务已解除阻塞，可以继续推进",
                )
        if changed_labels:
            if d.get("status", row["status"]) != row["status"]:
                summary = (
                    f"将进展状态从「{row['status']}」改为"
                    f"「{d['status']}」"
                )
                other_labels = [label for label in changed_labels if label != "进展状态"]
                if other_labels:
                    summary += f"，并更新了{'、'.join(other_labels)}"
            else:
                summary = f"更新了{'、'.join(changed_labels)}"
            add_task_activity(
                db, workspace_id, tid, "updated", summary,
                {"fields": changed_labels},
            )
        new_owner = d.get("owner", row["owner"])
        if new_owner != row["owner"]:
            add_notifications(
                db, workspace_id, owner_user_ids(db, workspace_id, new_owner),
                "assigned",
                f"{g.user['display_name']} 将事务「{d.get('name', row['name'])}」指派给你",
                tid, g.user["id"],
            )
        new_status = d.get("status", row["status"])
        if new_status != row["status"]:
            add_notifications(
                db, workspace_id,
                task_audience_user_ids(db, workspace_id, tid, new_owner),
                "status_changed",
                f"{g.user['display_name']} 将事务「{d.get('name', row['name'])}」"
                f"从「{row['status']}」改为「{new_status}」",
                tid, g.user["id"],
            )
            if row["status"] not in {"已闭环", "已取消"} and \
                    new_status in {"已闭环", "已取消"}:
                notify_dependents_unblocked(
                    db, workspace_id, tid, d.get("name", row["name"]),
                    g.user["id"],
                )
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/tasks/<int:tid>/dependencies", methods=["POST", "DELETE"])
def task_dependency(tid):
    d = json_object()
    prerequisite_task_id = required_id(
        d.get("prerequisite_task_id"), "prerequisite_task_id"
    )
    db = get_db()
    workspace_id = current_workspace_id()
    task = db.execute(
        "SELECT id,status FROM tasks WHERE id=? AND workspace_id=? AND deleted=0",
        (tid, workspace_id),
    ).fetchone()
    if not task:
        raise ApiError("事务不存在", 404)
    existing_ids = current_dependency_ids(db, workspace_id, tid)

    if request.method == "DELETE":
        if prerequisite_task_id not in existing_ids:
            raise ApiError("依赖关系不存在", 404)
        on_edit(db)
        db.execute(
            "DELETE FROM task_dependencies WHERE workspace_id=? "
            "AND dependent_task_id=? AND prerequisite_task_id=?",
            (workspace_id, tid, prerequisite_task_id),
        )
        prerequisite = db.execute(
            "SELECT name FROM tasks WHERE id=? AND workspace_id=?",
            (prerequisite_task_id, workspace_id),
        ).fetchone()
        prerequisite_name = prerequisite["name"] if prerequisite else "前置事务"
        add_task_activity(
            db, workspace_id, tid, "dependency_removed",
            f"移除了前置依赖「{prerequisite_name}」",
        )
        notify_task_if_unblocked(
            db, workspace_id, tid, g.user["id"],
            f"前置依赖「{prerequisite_name}」已解除，事务可以继续推进",
        )
        db.commit()
        return jsonify({"ok": True})

    if prerequisite_task_id in existing_ids:
        return jsonify({"ok": True, "created": False})
    dependency_ids = validate_dependencies(
        db, workspace_id, tid, existing_ids + [prerequisite_task_id]
    )
    if task["status"] == "已闭环":
        ensure_dependencies_closed(db, workspace_id, dependency_ids)
    on_edit(db)
    db.execute(
        "INSERT INTO task_dependencies("
        "workspace_id,dependent_task_id,prerequisite_task_id) VALUES(?,?,?)",
        (workspace_id, tid, prerequisite_task_id),
    )
    prerequisite = db.execute(
        "SELECT name FROM tasks WHERE id=? AND workspace_id=?",
        (prerequisite_task_id, workspace_id),
    ).fetchone()
    add_task_activity(
        db, workspace_id, tid, "dependency_added",
        f"添加了前置依赖「{prerequisite['name']}」",
    )
    db.commit()
    return jsonify({"ok": True, "created": True}), 201


@app.route("/api/tasks/<int:tid>", methods=["DELETE"])
def delete_task(tid):
    db = get_db()
    workspace_id = current_workspace_id()
    row = db.execute(
        "SELECT id FROM tasks WHERE id=? AND workspace_id=? AND deleted=0",
        (tid, workspace_id),
    ).fetchone()
    if not row:
        return jsonify({"error": "事务不存在"}), 404
    ensure_tasks_not_required(db, workspace_id, [tid])
    on_edit(db)
    batch = new_batch(db)
    db.execute(
        "UPDATE tasks SET deleted=1, del_batch=?, deleted_at=? "
        "WHERE id=? AND workspace_id=?",
        (batch, date.today().isoformat(), tid, workspace_id),
    )
    db.commit()
    return jsonify({"ok": True, "can_undo": True})


@app.route("/api/tasks/import-template")
def download_task_import_template():
    output = task_import_template_workbook(get_db(), current_workspace_id())
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"AnyLine-事务导入模板-{date.today():%Y%m%d}.xlsx",
    )


@app.route("/api/tasks/import", methods=["POST"])
def import_tasks():
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        raise ApiError("请选择要导入的 Excel 文件")
    if not uploaded.filename.lower().endswith(".xlsx"):
        raise ApiError("仅支持 .xlsx 格式的 Excel 文件")
    content = uploaded.read(MAX_TASK_IMPORT_BYTES + 1)
    if not content:
        raise ApiError("导入文件不能为空")
    if len(content) > MAX_TASK_IMPORT_BYTES:
        raise ApiError("导入文件不能超过 5MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError):
        raise ApiError("无法读取 Excel 文件，请使用下载的导入模板")
    try:
        if "事务导入" not in workbook.sheetnames:
            raise ApiError("Excel 文件缺少“事务导入”工作表")
        db = get_db()
        workspace_id = current_workspace_id()
        rows, row_errors, data_count = parse_task_import_sheet(
            workbook["事务导入"], db, workspace_id
        )
    finally:
        workbook.close()
    if not data_count:
        raise ApiError("“事务导入”工作表中没有可导入的数据")
    if row_errors:
        return jsonify({
            "error": f"导入文件存在 {len(row_errors)} 行错误，未导入任何事务",
            "error_count": len(row_errors),
            "row_errors": row_errors[:50],
        }), 400

    today = date.today().isoformat()
    on_edit(db)
    imported_ids = []
    for task in rows:
        cur = db.execute(
            "INSERT INTO tasks(workspace_id,line_id,name,content,goal,owner,priority,"
            "next_action,risk_reason,status,start_date,end_date,status_since,updated_at) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                workspace_id, task["line_id"], task["name"], task["content"],
                task["goal"], task["owner"], task["priority"], task["next_action"],
                task["risk_reason"], task["status"], task["start_date"],
                task["end_date"], today, today,
            ),
        )
        imported_ids.append(cur.lastrowid)
        add_task_activity(
            db, workspace_id, cur.lastrowid, "imported", "通过 Excel 导入了事务"
        )
    db.commit()
    return jsonify({
        "ok": True,
        "count": len(imported_ids),
        "ids": imported_ids,
        "can_undo": True,
    }), 201


@app.route("/api/tasks/export", methods=["POST"])
def export_tasks():
    d = json_object()
    scope = d.get("scope")
    workspace_id = current_workspace_id()
    params = [workspace_id, workspace_id]
    selected_count = None
    id_clause = ""
    if scope == "selected":
        ids = d.get("ids")
        if not isinstance(ids, list) or not ids or \
                not all(isinstance(i, int) and not isinstance(i, bool) for i in ids):
            return jsonify({"error": "ids 必须是非空整数数组"}), 400
        if len(ids) != len(set(ids)):
            return jsonify({"error": "ids 不能包含重复项"}), 400
        marks = ",".join("?" * len(ids))
        id_clause = f" AND t.id IN ({marks})"
        params.extend(ids)
        selected_count = len(ids)
    elif scope != "all":
        return jsonify({"error": "scope 必须是 all 或 selected"}), 400

    db = get_db()
    rows = [dict(row) for row in db.execute(
        "SELECT t.id,l.name AS line_name,"
        "CASE WHEN l.parent_id IS NULL THEN '主线' ELSE '支线' END AS line_type,"
        "p.name AS parent_name,t.name,t.content,t.goal,t.next_action,"
        "t.risk_reason,t.priority,t.owner,t.status,t.start_date,t.end_date,"
        "t.status_since,t.updated_at FROM tasks t "
        "JOIN lines l ON l.id=t.line_id "
        "LEFT JOIN lines p ON p.id=l.parent_id "
        "WHERE t.workspace_id=? AND l.workspace_id=? "
        "AND t.deleted=0 AND l.deleted=0" + id_clause +
        " ORDER BY t.start_date,t.id",
        params,
    ).fetchall()]
    if scope == "selected" and len(rows) != selected_count:
        return jsonify({"error": "部分事务不存在或已删除"}), 404

    output = task_export_workbook(rows)
    scope_name = "全部事务" if scope == "all" else "选中事务"
    filename = f"AnyLine-{scope_name}-{date.today():%Y%m%d}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/tasks/bulk", methods=["PATCH", "DELETE"])
def bulk_tasks():
    d = json_object()
    ids = d.get("ids")
    if not isinstance(ids, list) or not ids or \
            not all(isinstance(i, int) and not isinstance(i, bool) for i in ids):
        return jsonify({"error": "ids 必须是非空整数数组"}), 400
    if len(ids) != len(set(ids)):
        return jsonify({"error": "ids 不能包含重复项"}), 400
    db = get_db()
    workspace_id = current_workspace_id()
    marks = ",".join("?" * len(ids))
    existing = db.execute(
        f"SELECT id,name,owner,status,line_id,priority FROM tasks "
        f"WHERE workspace_id=? AND deleted=0 "
        f"AND id IN ({marks})", [workspace_id] + ids
    ).fetchall()
    if len(existing) != len(set(ids)):
        return jsonify({"error": "部分事务不存在或已删除"}), 404

    if request.method == "DELETE":
        ensure_tasks_not_required(db, workspace_id, ids)
        on_edit(db)
        batch = new_batch(db)
        db.execute(
            f"UPDATE tasks SET deleted=1, del_batch=?, deleted_at=? "
            f"WHERE workspace_id=? AND id IN ({marks})",
            [batch, date.today().isoformat(), workspace_id] + ids,
        )
        db.commit()
        return jsonify({"ok": True, "count": len(ids), "can_undo": True})

    patch = d.get("patch")
    if not isinstance(patch, dict) or not patch:
        return jsonify({"error": "patch 不能为空"}), 400
    allowed = {"line_id", "owner", "priority", "status"}
    unknown = set(patch) - allowed
    if unknown:
        return jsonify({"error": "批量更新不支持该字段"}), 400
    if "priority" in patch and patch["priority"] not in PRIORITY_ENUM:
        return jsonify({"error": "非法的优先级"}), 400
    if "status" in patch and patch["status"] not in get_statuses(db):
        return jsonify({"error": "非法的进展状态"}), 400
    if patch.get("status") == "已闭环":
        for task_id in ids:
            ensure_dependencies_closed(
                db, workspace_id,
                current_dependency_ids(db, workspace_id, task_id),
            )
    if "owner" in patch:
        if not isinstance(patch["owner"], str):
            return jsonify({"error": "责任人必须是字符串"}), 400
        patch["owner"] = patch["owner"].strip()
        if not patch["owner"]:
            return jsonify({"error": "责任人不能为空"}), 400
        if patch["owner"] not in get_workspace_member_names(db, workspace_id):
            return jsonify({"error": "责任人不是当前项目空间成员"}), 400
    if "line_id" in patch:
        required_id(patch["line_id"], "line_id")
        ln = db.execute(
            "SELECT id,fork_date FROM lines WHERE id=? AND workspace_id=? "
            "AND deleted=0", (patch["line_id"], workspace_id)
        ).fetchone()
        if not ln:
            return jsonify({"error": "所属线不存在"}), 404
        first_task = db.execute(
            f"SELECT MIN(start_date) AS first_date FROM tasks "
            f"WHERE workspace_id=? AND deleted=0 AND id IN ({marks})",
            [workspace_id] + ids,
        ).fetchone()
        if first_task["first_date"] < ln["fork_date"]:
            return jsonify({"error": "部分事务的起始日期早于目标线起始日期"}), 400

    fields, vals = [], []
    for k, v in patch.items():
        fields.append(f"{k}=?")
        vals.append(v)
        if k == "status":
            fields.append("status_since=?")
            vals.append(date.today().isoformat())
    fields.append("updated_at=?")
    vals.append(date.today().isoformat())
    on_edit(db)
    db.execute(
        f"UPDATE tasks SET {','.join(fields)} WHERE workspace_id=? "
        f"AND id IN ({marks})",
        vals + [workspace_id] + ids,
    )
    bulk_labels = {
        "line_id": "所属线", "owner": "责任人",
        "priority": "优先级", "status": "进展状态",
    }
    for task in existing:
        changed = [
            bulk_labels[key] for key, value in patch.items()
            if task[key] != value
        ]
        if not changed:
            continue
        add_task_activity(
            db, workspace_id, task["id"], "bulk_updated",
            f"批量更新了{'、'.join(changed)}", {"fields": changed},
        )
        new_owner = patch.get("owner", task["owner"])
        if new_owner != task["owner"]:
            add_notifications(
                db, workspace_id, owner_user_ids(db, workspace_id, new_owner),
                "assigned",
                f"{g.user['display_name']} 将事务「{task['name']}」指派给你",
                task["id"], g.user["id"],
            )
        new_status = patch.get("status", task["status"])
        if new_status != task["status"]:
            add_notifications(
                db, workspace_id,
                task_audience_user_ids(db, workspace_id, task["id"], new_owner),
                "status_changed",
                f"{g.user['display_name']} 将事务「{task['name']}」"
                f"从「{task['status']}」改为「{new_status}」",
                task["id"], g.user["id"],
            )
            if task["status"] not in {"已闭环", "已取消"} and \
                    new_status in {"已闭环", "已取消"}:
                notify_dependents_unblocked(
                    db, workspace_id, task["id"], task["name"], g.user["id"]
                )
    db.commit()
    return jsonify({"ok": True, "count": len(ids)})


# ----- undo
@app.route("/api/undo", methods=["POST"])
def undo():
    db = get_db()
    workspace_id = current_workspace_id()
    saved = db.execute(
        "SELECT snapshot FROM undo_snapshots WHERE workspace_id=?", (workspace_id,)
    ).fetchone()
    if saved:
        try:
            snapshot = json.loads(saved["snapshot"])
        except (TypeError, ValueError):
            raise ApiError("撤销数据已损坏", 500)
        if not isinstance(snapshot, dict):
            raise ApiError("撤销数据已损坏", 500)
        save_workspace_snapshot(
            db, "redo_snapshots", workspace_id,
            current_workspace_snapshot(db, workspace_id),
        )
        restore_snapshot(db, snapshot)
        db.execute("DELETE FROM undo_snapshots WHERE workspace_id=?", (workspace_id,))
        set_meta(db, "undo_batch", None)
        db.commit()
        return jsonify({"ok": True})

    # 兼容升级前已产生、尚未使用的最近删除批次。
    batch = get_meta(db, "undo_batch")
    if batch is None:
        return jsonify({"error": "没有可撤销的操作"}), 400
    save_workspace_snapshot(
        db, "redo_snapshots", workspace_id,
        current_workspace_snapshot(db, workspace_id),
    )
    db.execute(
        "UPDATE lines SET deleted=0, del_batch=NULL, deleted_at=NULL "
        "WHERE workspace_id=? AND del_batch=?", (workspace_id, batch)
    )
    db.execute(
        "UPDATE tasks SET deleted=0, del_batch=NULL, deleted_at=NULL "
        "WHERE workspace_id=? AND del_batch=?", (workspace_id, batch)
    )
    db.execute(
        "UPDATE milestones SET deleted=0, del_batch=NULL, deleted_at=NULL "
        "WHERE workspace_id=? AND del_batch=?", (workspace_id, batch)
    )
    set_meta(db, "undo_batch", None)
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/redo", methods=["POST"])
def redo():
    db = get_db()
    workspace_id = current_workspace_id()
    saved = db.execute(
        "SELECT snapshot FROM redo_snapshots WHERE workspace_id=?", (workspace_id,)
    ).fetchone()
    if not saved:
        raise ApiError("没有可恢复的操作")
    try:
        snapshot = json.loads(saved["snapshot"])
    except (TypeError, ValueError):
        raise ApiError("恢复数据已损坏", 500)
    if not isinstance(snapshot, dict):
        raise ApiError("恢复数据已损坏", 500)
    save_workspace_snapshot(
        db, "undo_snapshots", workspace_id,
        current_workspace_snapshot(db, workspace_id),
    )
    restore_snapshot(db, snapshot)
    db.execute("DELETE FROM redo_snapshots WHERE workspace_id=?", (workspace_id,))
    set_meta(db, "undo_batch", None)
    db.commit()
    return jsonify({"ok": True})


# ----- trash
@app.route("/api/trash", methods=["GET"])
def trash():
    db = get_db()
    workspace_id = current_workspace_id()
    line_rows = [dict(r) for r in db.execute(
        "SELECT id,name,parent_id,fork_date,merge_date,del_batch,deleted_at "
        "FROM lines WHERE workspace_id=? AND deleted=1 "
        "ORDER BY deleted_at DESC,id DESC", (workspace_id,)
    )]
    task_rows = [dict(r) for r in db.execute(
        "SELECT id,line_id,name,status,owner,priority,start_date,end_date,"
        "del_batch,deleted_at FROM tasks WHERE workspace_id=? AND deleted=1 "
        "ORDER BY deleted_at DESC,id DESC", (workspace_id,)
    )]
    milestone_rows = [dict(r) for r in db.execute(
        "SELECT id,line_id,name,milestone_date,del_batch,deleted_at "
        "FROM milestones WHERE workspace_id=? AND deleted=1 "
        "ORDER BY deleted_at DESC,id DESC", (workspace_id,)
    )]
    batches = {}
    for row in line_rows:
        b = str(row["del_batch"])
        batches.setdefault(b, {
            "batch": row["del_batch"], "deleted_at": row["deleted_at"],
            "line_count": 0, "task_count": 0, "milestone_count": 0, "names": [],
        })
        batches[b]["line_count"] += 1
        batches[b]["names"].append(row["name"])
    for row in task_rows:
        b = str(row["del_batch"])
        batches.setdefault(b, {
            "batch": row["del_batch"], "deleted_at": row["deleted_at"],
            "line_count": 0, "task_count": 0, "milestone_count": 0, "names": [],
        })
        batches[b]["task_count"] += 1
        batches[b]["names"].append(row["name"])
    for row in milestone_rows:
        b = str(row["del_batch"])
        batches.setdefault(b, {
            "batch": row["del_batch"], "deleted_at": row["deleted_at"],
            "line_count": 0, "task_count": 0, "milestone_count": 0, "names": [],
        })
        batches[b]["milestone_count"] += 1
        batches[b]["names"].append(row["name"])
    return jsonify({
        "batches": sorted(
            batches.values(),
            key=lambda x: (x["deleted_at"] or "", x["batch"] or 0),
            reverse=True,
        ),
        "lines": line_rows,
        "tasks": task_rows,
        "milestones": milestone_rows,
    })


@app.route("/api/trash/restore", methods=["POST"])
def restore_trash():
    d = json_object()
    batch = d.get("batch")
    if batch is None:
        return jsonify({"error": "batch 不能为空"}), 400
    required_id(batch, "batch")
    db = get_db()
    workspace_id = current_workspace_id()
    found = db.execute(
        "SELECT 1 FROM lines WHERE workspace_id=? AND del_batch=? AND deleted=1 "
        "UNION SELECT 1 FROM tasks WHERE workspace_id=? AND del_batch=? AND deleted=1 "
        "UNION SELECT 1 FROM milestones WHERE workspace_id=? AND del_batch=? "
        "AND deleted=1",
        (workspace_id, batch, workspace_id, batch, workspace_id, batch),
    ).fetchone()
    if not found:
        return jsonify({"error": "未找到该删除批次"}), 404
    blocked_line = db.execute(
        "SELECT child.id FROM lines child JOIN lines parent "
        "ON parent.id=child.parent_id WHERE child.del_batch=? AND child.deleted=1 "
        "AND child.workspace_id=? AND parent.workspace_id=? "
        "AND parent.deleted=1 AND parent.del_batch<>? LIMIT 1",
        (batch, workspace_id, workspace_id, batch)
    ).fetchone()
    blocked_task = db.execute(
        "SELECT task.id FROM tasks task JOIN lines line ON line.id=task.line_id "
        "WHERE task.del_batch=? AND task.deleted=1 AND line.deleted=1 "
        "AND task.workspace_id=? AND line.workspace_id=? "
        "AND line.del_batch<>? LIMIT 1",
        (batch, workspace_id, workspace_id, batch)
    ).fetchone()
    blocked_dependency = db.execute(
        "SELECT dependent.id FROM task_dependencies d "
        "JOIN tasks dependent ON dependent.id=d.dependent_task_id "
        "JOIN tasks prerequisite ON prerequisite.id=d.prerequisite_task_id "
        "WHERE d.workspace_id=? AND dependent.workspace_id=? "
        "AND dependent.deleted=1 AND dependent.del_batch=? "
        "AND prerequisite.deleted=1 AND prerequisite.del_batch<>? LIMIT 1",
        (workspace_id, workspace_id, batch, batch),
    ).fetchone()
    blocked_milestone_line = db.execute(
        "SELECT milestone.id FROM milestones milestone JOIN lines line "
        "ON line.id=milestone.line_id WHERE milestone.workspace_id=? "
        "AND line.workspace_id=? AND milestone.deleted=1 "
        "AND milestone.del_batch=? AND line.deleted=1 AND line.del_batch<>? LIMIT 1",
        (workspace_id, workspace_id, batch, batch),
    ).fetchone()
    blocked_milestone_task = db.execute(
        "SELECT milestone.id FROM milestones milestone "
        "JOIN milestone_tasks relation ON relation.milestone_id=milestone.id "
        "AND relation.workspace_id=milestone.workspace_id "
        "JOIN tasks task ON task.id=relation.task_id "
        "WHERE milestone.workspace_id=? AND task.workspace_id=? "
        "AND milestone.deleted=1 AND milestone.del_batch=? "
        "AND task.deleted=1 AND task.del_batch<>? LIMIT 1",
        (workspace_id, workspace_id, batch, batch),
    ).fetchone()
    if (blocked_line or blocked_task or blocked_dependency or
            blocked_milestone_line or blocked_milestone_task):
        return jsonify({"error": "请先恢复该批次依赖的所属线或前置事务"}), 409
    on_edit(db)
    db.execute(
        "UPDATE lines SET deleted=0, del_batch=NULL, deleted_at=NULL "
        "WHERE workspace_id=? AND del_batch=?", (workspace_id, batch)
    )
    db.execute(
        "UPDATE tasks SET deleted=0, del_batch=NULL, deleted_at=NULL "
        "WHERE workspace_id=? AND del_batch=?", (workspace_id, batch)
    )
    db.execute(
        "UPDATE milestones SET deleted=0, del_batch=NULL, deleted_at=NULL "
        "WHERE workspace_id=? AND del_batch=?", (workspace_id, batch)
    )
    if str(get_meta(db, "undo_batch")) == str(batch):
        set_meta(db, "undo_batch", None)
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/trash/purge", methods=["POST"])
def purge_trash():
    db = get_db()
    purge_deleted(db)
    db.commit()
    return jsonify({"ok": True})


init_db()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=80, debug=False)
