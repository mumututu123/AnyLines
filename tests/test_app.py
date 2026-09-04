import http.client
import io
import json
import os
import base64
import sqlite3
import tempfile
import threading
import unittest
from datetime import date, timedelta

from openpyxl import Workbook, load_workbook
from PIL import Image
from werkzeug.serving import make_server

_IMPORT_TEMP_DIR = tempfile.TemporaryDirectory()
os.environ["ANYLINE_DB_PATH"] = os.path.join(_IMPORT_TEMP_DIR.name, "import.db")
os.environ["ANYLINE_ADMIN_USERNAME"] = "admin"
os.environ["ANYLINE_ADMIN_PASSWORD"] = "admin123"
import app as anyline


class AnyLineHttpTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        cls.server = make_server("127.0.0.1", 0, anyline.app, threaded=True)
        cls.port = cls.server.server_port
        cls.thread = threading.Thread(target=cls.server.serve_forever, daemon=True)
        cls.thread.start()

    @classmethod
    def tearDownClass(cls):
        cls.server.shutdown()
        cls.thread.join(timeout=5)
        cls.server.server_close()
        cls.temp_dir.cleanup()
        _IMPORT_TEMP_DIR.cleanup()

    def setUp(self):
        db_path = os.path.join(self.temp_dir.name, f"{self._testMethodName}.db")
        anyline.app.config.update(DATABASE=db_path, TESTING=True)
        anyline.init_db(db_path)
        self.today = date.today()
        self.cookie = None
        status, data = self.request(
            "POST", "/api/auth/login", {"username": "admin", "password": "admin123"}
        )
        self.assertEqual(status, 200, data)

    def request(self, method, path, payload=None, raw_body=None):
        body = raw_body
        headers = {}
        if payload is not None:
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            headers["Content-Type"] = "application/json"
        elif raw_body is not None:
            headers["Content-Type"] = "application/json"
        if self.cookie:
            headers["Cookie"] = self.cookie
        connection = http.client.HTTPConnection("127.0.0.1", self.port, timeout=5)
        connection.request(method, path, body=body, headers=headers)
        response = connection.getresponse()
        set_cookie = response.getheader("Set-Cookie")
        if set_cookie:
            self.cookie = set_cookie.split(";", 1)[0]
        content_type = response.getheader("Content-Type", "")
        raw = response.read()
        connection.close()
        data = json.loads(raw.decode("utf-8")) if "application/json" in content_type else raw
        return response.status, data

    def upload_xlsx(self, content, filename="tasks.xlsx", path="/api/tasks/import"):
        boundary = "----AnyLineTestBoundary"
        body = (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
            "Content-Type: application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet\r\n\r\n"
        ).encode("utf-8") + content + f"\r\n--{boundary}--\r\n".encode("ascii")
        headers = {"Content-Type": f"multipart/form-data; boundary={boundary}"}
        if self.cookie:
            headers["Cookie"] = self.cookie
        connection = http.client.HTTPConnection("127.0.0.1", self.port, timeout=5)
        connection.request("POST", path, body=body, headers=headers)
        response = connection.getresponse()
        raw = response.read()
        content_type = response.getheader("Content-Type", "")
        connection.close()
        data = json.loads(raw.decode("utf-8")) if "application/json" in content_type else raw
        return response.status, data

    def login(self, username, password):
        self.cookie = None
        return self.request(
            "POST", "/api/auth/login", {"username": username, "password": password}
        )

    def create_line(
        self, name="主线", fork_date=None, parent_id=None,
        description="", color=None,
    ):
        payload = {
            "name": name,
            "description": description,
            "color": color,
            "fork_date": fork_date or self.today.isoformat(),
            "parent_id": parent_id,
        }
        status, data = self.request("POST", "/api/lines", payload)
        self.assertEqual(status, 201, data)
        return data["id"]

    def create_task(self, line_id, name="事务", **overrides):
        payload = {
            "line_id": line_id,
            "name": name,
            "content": "内容",
            "goal": "目标",
            "owner": "系统管理员",
            "priority": "高",
            "next_action": "下一步",
            "risk_reason": "",
            "status": "进行中",
            "start_date": self.today.isoformat(),
            "end_date": (self.today + timedelta(days=7)).isoformat(),
        }
        payload.update(overrides)
        status, data = self.request("POST", "/api/tasks", payload)
        self.assertEqual(status, 201, data)
        return data["id"]

    def create_milestone(self, line_id, task_ids=None, **overrides):
        payload = {
            "line_id": line_id,
            "name": "阶段验收",
            "target_description": "确认阶段目标已经达成",
            "milestone_date": (self.today + timedelta(days=5)).isoformat(),
            "acceptance_task_ids": task_ids or [],
        }
        payload.update(overrides)
        status, data = self.request("POST", "/api/milestones", payload)
        self.assertEqual(status, 201, data)
        return data["id"]

    def add_member(self, username, display_name, role="member"):
        workspace_id = self.request(
            "GET", "/api/auth/session"
        )[1]["current_workspace"]["id"]
        status, data = self.request(
            "POST", f"/api/workspaces/{workspace_id}/members", {
                "username": username,
                "display_name": display_name,
                "password": "member123",
                "role": role,
            },
        )
        self.assertEqual(status, 201, data)
        return data["user_id"]

    def test_index_and_empty_state(self):
        status, body = self.request("GET", "/")
        self.assertEqual(status, 200)
        self.assertIn(b"AnyLine", body)
        self.assertIn(b'id="workspace-select"', body)
        self.assertIn(b'id="btn-workspaces"', body)
        self.assertIn(b'id="btn-view-dashboard"', body)
        self.assertIn(b'id="btn-my-status"', body)
        self.assertIn(b'id="my-status-count"', body)
        self.assertNotIn(b'id="btn-my-todos"', body)
        self.assertNotIn(b'id="btn-notifications"', body)
        self.assertNotIn(b'<span>\xe6\x88\x91\xe7\x9a\x84\xe5\xbe\x85\xe5\x8a\x9e</span>', body)
        self.assertIn(b'id="dashboard-view"', body)
        self.assertIn(b'id="dashboard-report-summary"', body)
        self.assertIn(b'id="dashboard-achievements"', body)
        self.assertIn(b'id="dashboard-variance"', body)
        self.assertIn(b'id="btn-dashboard-meeting"', body)
        self.assertIn(b'id="btn-dashboard-print"', body)
        self.assertIn(b'id="image-lightbox"', body)
        self.assertNotIn(b'id="btn-workspace-create"', body)
        self.assertNotIn(b'id="btn-owners"', body)
        self.assertNotIn(b'id="btn-delete-line"', body)
        self.assertNotIn(b'id="btn-undo"', body)

        status, state = self.request("GET", "/api/state")
        self.assertEqual(status, 200)
        self.assertEqual(state["lines"], [])
        self.assertEqual(state["tasks"], [])
        self.assertFalse(state["can_undo"])
        self.assertFalse(state["can_redo"])
        self.assertEqual(state["priority_enum"], ["低", "中", "高", "紧急"])
        self.assertEqual(state["status_colors"]["进行中"], "#0969da")
        self.assertEqual(state["owners"], ["系统管理员"])

    def test_user_can_upload_a_constrained_custom_avatar(self):
        source = io.BytesIO()
        Image.new("RGB", (320, 180), (30, 120, 210)).save(source, format="PNG")
        data_url = "data:image/png;base64," + base64.b64encode(
            source.getvalue()
        ).decode("ascii")

        status, result = self.request(
            "PUT", "/api/auth/avatar", {"data_url": data_url}
        )
        self.assertEqual(status, 200, result)
        self.assertTrue(result["avatar_url"].startswith("/api/auth/avatar?v="))

        status, content = self.request("GET", "/api/auth/avatar")
        self.assertEqual(status, 200)
        with Image.open(io.BytesIO(content)) as avatar:
            self.assertEqual(avatar.size, (256, 256))
            self.assertEqual(avatar.format, "JPEG")

        status, session_data = self.request("GET", "/api/auth/session")
        self.assertEqual(status, 200)
        self.assertEqual(session_data["user"]["avatar_url"], result["avatar_url"])

        tiny = io.BytesIO()
        Image.new("RGB", (32, 32), "red").save(tiny, format="PNG")
        tiny_url = "data:image/png;base64," + base64.b64encode(
            tiny.getvalue()
        ).decode("ascii")
        status, error = self.request(
            "PUT", "/api/auth/avatar", {"data_url": tiny_url}
        )
        self.assertEqual(status, 400, error)
        self.assertIn("64", error["error"])

        status, body = self.request("GET", "/")
        self.assertEqual(status, 200)
        self.assertIn(b'id="btn-avatar-edit"', body)
        self.assertIn(b'id="avatar-file"', body)

    def test_personal_todo_entry_count_and_statistics(self):
        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")
        self.assertIn("function personalTodoTasks()", source)
        self.assertIn("ownerNames.has((task.owner || \"\").trim()) && !isDone(task)", source)
        self.assertIn("function renderMyStatusEntry()", source)
        self.assertIn("const count = todoCount + unreadCount", source)
        self.assertIn('badge.textContent = count > 99 ? "99+" : String(count)', source)
        self.assertIn("function openMyStatusModal(", source)
        self.assertIn("function renderMyTodoPanel(", source)
        for label in ("待办总数", "已超期", "有风险", "7天内到期", "被前置阻塞"):
            with self.subTest(statistic=label):
                self.assertIn(label, source)
        self.assertIn("function renderTaskListTable(", source)
        self.assertIn("item.onclick = () => activateFilter(filter, item)", source)
        self.assertIn('row.setAttribute("aria-label", `查看事务详情：${task.name}`)', source)
        self.assertIn('onClosed: () => openMyStatusModal("todos", selected.key, listScrollTop)', source)
        self.assertIn('$("#btn-my-status").onclick = () => openMyStatusModal()', source)

        status, body = self.request("GET", "/static/style.css")
        self.assertEqual(status, 200)
        styles = body.decode("utf-8")
        self.assertIn(".my-status-entry {", styles)
        self.assertIn(".my-status-count {", styles)
        self.assertIn(".my-todo-summary {", styles)
        self.assertIn("#modal.my-status-modal {", styles)
        self.assertIn(".my-status-tabs {", styles)
        self.assertIn(".dashboard-task-list-wrap.my-todo-list-wrap {", styles)

    def test_canvas_merge_uses_latest_task_end_and_vertical_line(self):
        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")

        self.assertIn("function latestLineTaskEndDate(line)", source)
        self.assertIn("const mergeDate = latestLineTaskEndDate(line);", source)
        self.assertIn("const end = { x: horizontalEnd.x, y: parentY };", source)
        self.assertIn("d += ` L ${merge.end.x} ${merge.end.y}`;", source)
        self.assertIn("const childMergeDate = latestLineTaskEndDate(child);", source)
        self.assertIn(
            "const lineHeadX = parent ? geometry.horizontalStart.x : x1;",
            source,
        )
        self.assertIn('"text-anchor": "end"', source)
        self.assertIn("lbl.textContent = line.name", source)
        self.assertIn("cx: merge.end.x, cy: merge.end.y", source)
        self.assertNotIn("BRANCH_SLOPE", source)
        self.assertIn('e.key === "Delete"', source)
        self.assertIn('key === "z"', source)

    def test_nested_branch_starts_on_parent_at_its_fork_date(self):
        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")

        self.assertIn("x: x(line.fork_date)", source)
        self.assertIn("y: lineY(parent.id)", source)
        self.assertIn("function lineGeometry(line)", source)
        self.assertIn("const horizontalStart = { x: start.x, y };", source)
        self.assertIn(
            "`M ${start.x} ${start.y} L ${horizontalStart.x} ${horizontalStart.y} `",
            source,
        )
        self.assertNotIn("connectorBend", source)
        self.assertIn(
            "state.lines.filter((candidate) => candidate.parent_id === line.id)",
            source,
        )
        self.assertNotIn(
            "(parentGeometry.start.x + parentGeometry.diagonalEnd.x) / 2",
            source,
        )
        self.assertNotIn("pointOnLineAtX", source)

    def test_canvas_shortcuts_include_redo_today_branch_and_task(self):
        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")

        self.assertIn('key === "r"', source)
        self.assertIn('await api("/api/redo", "POST")', source)
        self.assertIn(
            'else if (key === "m") createMilestoneOnSelectedLine();', source
        )
        self.assertIn('if (key === "h") goToToday();', source)
        self.assertIn(
            'else if (key === "b") createBranchOnSelectedLine();', source
        )
        self.assertIn('else createTaskOnSelectedLine();', source)
        self.assertIn('["h", "b", "m", "a", "n"].includes(key)', source)
        self.assertIn('state.view !== "canvas"', source)
        self.assertIn('target.matches("input, textarea, select")', source)
        self.assertIn('!$("#modal-mask").classList.contains("hidden")', source)

        status, body = self.request("GET", "/")
        self.assertEqual(status, 200)
        markup = body.decode("utf-8")
        self.assertIn('id="canvas-shortcuts"', markup)
        for key in ("Ctrl+Z", "Ctrl+R", "M", "H", "B", "A", "N", "Delete", "Ctrl+滚轮"):
            with self.subTest(shortcut_hint=key):
                self.assertIn(f"<kbd>{key}</kbd>", markup)

    def test_same_day_tasks_spread_horizontally_at_high_zoom(self):
        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")

        self.assertIn(
            "const SAME_DAY_AUTO_SPREAD_ZOOM = CANVAS_DETAIL_MIN_ZOOM;", source
        )
        self.assertIn("const SAME_DAY_NODE_DIAMETER = 24;", source)
        self.assertIn("const SAME_DAY_NODE_GAP = 8;", source)
        self.assertIn(
            "const SAME_DAY_SPREAD_STEP = SAME_DAY_NODE_DIAMETER + SAME_DAY_NODE_GAP;",
            source,
        )
        self.assertIn(
            "const autoSpreadSameDay = z >= SAME_DAY_AUTO_SPREAD_ZOOM;", source
        )
        self.assertIn("const totalSpread = (arr.length - 1) * SAME_DAY_SPREAD_STEP;", source)
        self.assertIn("lineGeometry(line).horizontalStart.x", source)
        self.assertIn("drawTask(t, baseY, false, xs[i], i)", source)
        self.assertIn("详细密度下自动水平排开，并留足最大节点直径", source)
        self.assertIn("state.canvasTaskPositions.set(t.id, { x: cx, y });", source)
        self.assertIn("const roundedSquareAttrs =", source)
        self.assertIn('const node = svgEl("rect", {', source)
        self.assertNotIn("trianglePoints(", source)

        self.assertIn("focusedClusterKey: null", source)
        self.assertIn("function renderClusterFocusLens(", source)
        self.assertIn("task-node cluster-focus-node", source)
        self.assertIn("标准密度不再提供手动展开/折叠", source)
        self.assertNotIn("expandedClusters", source)
        self.assertNotIn('class: "cluster-hint"', source)
        self.assertNotIn('class: "cluster-focus-trigger"', source)
        self.assertIn("画布倍率保持不变", source)
        self.assertIn('dismissClusterFocus({ restoreFocus: true })', source)

        status, body = self.request("GET", "/static/style.css")
        self.assertEqual(status, 200)
        styles = body.decode("utf-8")
        self.assertIn(".cluster-focus-lens {", styles)
        self.assertIn("#canvas-wrap.cluster-focus-active", styles)
        self.assertIn("@keyframes cluster-focus-pulse", styles)

    def test_canvas_dependency_focus_visual_encoding_and_semantic_zoom(self):
        status, body = self.request("GET", "/")
        self.assertEqual(status, 200)
        markup = body.decode("utf-8")
        for element_id in (
            "canvas-density-label", "dependency-focus-panel",
            "dependency-focus-blockers", "dependency-focus-affected",
            "canvas-legend", "canvas-status-legend", "opt-date",
        ):
            with self.subTest(element_id=element_id):
                self.assertIn(f'id="{element_id}"', markup)

        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")
        self.assertIn("function taskDependencyFocus(taskId)", source)
        self.assertIn("const upstream = collect(taskId, prerequisitesByTask)", source)
        self.assertIn("const downstream = collect(taskId, dependentsByTask)", source)
        self.assertIn("is-focus-dimmed", source)
        self.assertIn("primaryHealthBadge", source)
        self.assertIn('if (health.overdue) return ["!", "overdue", "超期"]', source)
        self.assertIn('if (health.risk) return ["险", "risk", "风险"]', source)
        self.assertIn('transform: `translate(${cx} ${cy})`', source)
        self.assertIn("CANVAS_OVERVIEW_MAX_ZOOM = 0.7", source)
        self.assertIn("CANVAS_DETAIL_MIN_ZOOM = 1.5", source)
        self.assertIn('class: "line-task-summary"', source)
        self.assertIn("const showDependencies = Boolean(dependencyFocus);", source)
        self.assertIn('["#opt-date", "date"]', source)

        status, body = self.request("GET", "/static/style.css")
        self.assertEqual(status, 200)
        styles = body.decode("utf-8")
        self.assertIn(".task-node { stroke: none;", styles)
        self.assertIn(".task-item.is-focus-selected .task-node", styles)
        self.assertIn("stroke: #79aee8", styles)
        self.assertIn("drop-shadow(0 0 2px rgba(121, 174, 232, .3))", styles)
        self.assertIn(".task-item.is-focus-upstream .task-node", styles)
        self.assertIn(".task-item.is-focus-downstream .task-node", styles)
        self.assertIn(".task-alert-overdue circle", styles)
        self.assertIn(".task-layer.semantic-overview", styles)
        self.assertNotIn(".task-node.health-overdue", styles)

    def test_located_canvas_task_uses_transient_pulse_emphasis(self):
        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")
        self.assertIn("requestAnimationFrame(() => emphasizeCanvasTask(id))", source)
        self.assertIn('node.classList.add("locate-emphasis")', source)
        self.assertIn('node.addEventListener("animationend"', source)

        status, body = self.request("GET", "/static/style.css")
        self.assertEqual(status, 200)
        styles = body.decode("utf-8")
        self.assertIn("@keyframes task-locate-pulse", styles)
        self.assertIn("transform: scale(1.18)", styles)
        self.assertIn("@media (prefers-reduced-motion: reduce)", styles)

    def test_authentication_is_required(self):
        self.cookie = None
        status, data = self.request("GET", "/api/state")
        self.assertEqual(status, 401, data)
        self.assertIn("登录", data["error"])

        status, data = self.login("admin", "wrong-password")
        self.assertEqual(status, 401, data)
        status, data = self.login("admin", "admin123")
        self.assertEqual(status, 200, data)
        self.assertTrue(data["authenticated"])
        self.assertEqual(data["current_workspace"]["role"], "admin")

    def test_workspace_isolation_and_member_permissions(self):
        default_line = self.create_line("默认空间主线")
        status, created = self.request(
            "POST", "/api/workspaces", {"name": "第二项目", "description": "隔离测试"}
        )
        self.assertEqual(status, 201, created)
        second_workspace = created["id"]
        second_line = self.create_line("第二空间主线")
        status, state = self.request("GET", "/api/state")
        self.assertEqual(status, 200)
        self.assertEqual([line["id"] for line in state["lines"]], [second_line])

        status, member = self.request(
            "POST", f"/api/workspaces/{second_workspace}/members", {
                "username": "member1", "display_name": "普通成员",
                "password": "member123", "role": "member",
            }
        )
        self.assertEqual(status, 201, member)
        admin_cookie = self.cookie

        status, login = self.login("member1", "member123")
        self.assertEqual(status, 200, login)
        self.assertEqual(login["current_workspace"]["id"], second_workspace)
        status, state = self.request("GET", "/api/state")
        self.assertEqual(status, 200)
        self.assertEqual([line["id"] for line in state["lines"]], [second_line])
        status, _ = self.request("POST", "/api/lines", {
            "name": "成员可写主线", "fork_date": self.today.isoformat(),
        })
        self.assertEqual(status, 201)
        second_task = self.create_task(second_line, "第二空间事务")
        self.assertEqual(
            self.request("POST", "/api/workspaces", {"name": "越权空间"})[0], 403
        )
        self.assertEqual(
            self.request("GET", f"/api/workspaces/{second_workspace}/members")[0], 403
        )
        session_data = self.request("GET", "/api/auth/session")[1]
        self.assertEqual(len(session_data["workspaces"]), 1)

        self.cookie = admin_cookie
        default_workspace = next(
            workspace["id"] for workspace in
            self.request("GET", "/api/auth/session")[1]["workspaces"]
            if workspace["name"] == "默认项目"
        )
        self.assertEqual(
            self.request(
                "POST", f"/api/workspaces/{default_workspace}/select"
            )[0], 200
        )
        state = self.request("GET", "/api/state")[1]
        self.assertEqual([line["id"] for line in state["lines"]], [default_line])
        self.assertEqual(state["owners"], ["系统管理员"])
        self.assertEqual(
            self.request(
                "PATCH", f"/api/lines/{second_line}", {"name": "跨空间修改"}
            )[0], 404
        )
        self.assertEqual(
            self.request(
                "PATCH", f"/api/tasks/{second_task}", {"name": "跨空间事务"}
            )[0], 404
        )
        self.assertEqual(
            self.request(
                "POST", "/api/tasks/export", {"scope": "selected", "ids": [second_task]}
            )[0], 404
        )

    def test_workspace_archive_is_read_only_and_workspace_can_be_deleted(self):
        session_data = self.request("GET", "/api/auth/session")[1]
        default_workspace = session_data["current_workspace"]["id"]
        line_id = self.create_line("待归档主线")
        task_id = self.create_task(line_id, "待归档事务")

        status, data = self.request(
            "DELETE", f"/api/workspaces/{default_workspace}",
            {"confirmation": "默认项目"},
        )
        self.assertEqual(status, 409, data)
        self.assertIn("至少需要保留", data["error"])

        status, created = self.request(
            "POST", "/api/workspaces", {"name": "保留项目", "description": "删除后切换"}
        )
        self.assertEqual(status, 201, created)
        remaining_workspace = created["id"]
        self.assertEqual(
            self.request(
                "POST", f"/api/workspaces/{default_workspace}/select"
            )[0], 200
        )

        status, archived = self.request(
            "POST", f"/api/workspaces/{default_workspace}/archive"
        )
        self.assertEqual(status, 200, archived)
        session_data = self.request("GET", "/api/auth/session")[1]
        self.assertEqual(
            session_data["current_workspace"]["archived_at"],
            self.today.isoformat(),
        )
        self.assertEqual(self.request("GET", "/api/state")[0], 200)
        self.assertEqual(
            self.request("POST", "/api/tasks/export", {"scope": "all"})[0], 200
        )

        readonly_requests = [
            ("POST", "/api/lines", {
                "name": "不可新增", "fork_date": self.today.isoformat(),
            }),
            ("PATCH", f"/api/tasks/{task_id}", {"name": "不可修改"}),
            ("POST", f"/api/tasks/{task_id}/comments", {"content": "不可评论"}),
            ("POST", f"/api/tasks/{task_id}/follow", None),
            ("PUT", "/api/statuses", {"statuses": ["不可修改"]}),
            ("PATCH", f"/api/workspaces/{default_workspace}", {"name": "不可改名"}),
            ("POST", f"/api/workspaces/{default_workspace}/members", {
                "username": "archived-member", "display_name": "归档成员",
                "password": "member123", "role": "member",
            }),
        ]
        for method, path, payload in readonly_requests:
            with self.subTest(method=method, path=path):
                status, data = self.request(method, path, payload)
                self.assertEqual(status, 409, data)
                self.assertIn("已归档", data["error"])

        self.assertEqual(
            self.request(
                "POST", f"/api/workspaces/{default_workspace}/archive"
            )[0], 409
        )
        self.assertEqual(
            self.request(
                "POST", f"/api/workspaces/{default_workspace}/restore"
            )[0], 200
        )
        session_data = self.request("GET", "/api/auth/session")[1]
        self.assertIsNone(session_data["current_workspace"]["archived_at"])
        self.assertEqual(
            self.request(
                "PATCH", f"/api/tasks/{task_id}", {"name": "恢复后可修改"}
            )[0], 200
        )
        self.assertEqual(
            self.request(
                "POST", f"/api/workspaces/{default_workspace}/restore"
            )[0], 409
        )
        self.assertEqual(
            self.request(
                "POST", f"/api/workspaces/{default_workspace}/archive"
            )[0], 200
        )
        status, data = self.request(
            "DELETE", f"/api/workspaces/{default_workspace}",
            {"confirmation": "名称不匹配"},
        )
        self.assertEqual(status, 400, data)
        self.assertIn("名称不匹配", data["error"])
        status, data = self.request(
            "DELETE", f"/api/workspaces/{default_workspace}",
            {"confirmation": "默认项目"},
        )
        self.assertEqual(status, 200, data)
        self.assertEqual(data["current_workspace_id"], remaining_workspace)
        session_data = self.request("GET", "/api/auth/session")[1]
        self.assertEqual(session_data["current_workspace"]["id"], remaining_workspace)
        self.assertEqual(
            self.request(
                "POST", f"/api/workspaces/{default_workspace}/select"
            )[0], 403
        )
        with anyline.app.app_context():
            db = anyline.get_db()
            self.assertIsNone(db.execute(
                "SELECT id FROM workspaces WHERE id=?", (default_workspace,)
            ).fetchone())
            self.assertIsNone(db.execute(
                "SELECT id FROM tasks WHERE id=?", (task_id,)
            ).fetchone())

    def test_admin_can_manage_member_role_and_password(self):
        auth = self.request("GET", "/api/auth/session")[1]
        workspace_id = auth["current_workspace"]["id"]
        status, created = self.request(
            "POST", f"/api/workspaces/{workspace_id}/members", {
                "username": "project-admin", "display_name": "项目成员",
                "password": "initial123", "role": "member",
            }
        )
        self.assertEqual(status, 201, created)
        user_id = created["user_id"]

        members = self.request(
            "GET", f"/api/workspaces/{workspace_id}/members"
        )[1]["members"]
        managed = next(member for member in members if member["id"] == user_id)
        self.assertEqual(managed["role"], "member")
        self.assertTrue(managed["can_manage_account"])

        status, data = self.request(
            "PATCH", f"/api/workspaces/{workspace_id}/members/{user_id}", {
                "display_name": "项目管理员", "password": "changed123",
                "role": "admin",
            }
        )
        self.assertEqual(status, 200, data)
        self.assertEqual(self.login("project-admin", "initial123")[0], 401)
        status, login = self.login("project-admin", "changed123")
        self.assertEqual(status, 200, login)
        self.assertEqual(login["current_workspace"]["role"], "admin")
        self.assertEqual(
            self.request("POST", "/api/workspaces", {"name": "管理员新空间"})[0],
            201,
        )

    def test_configuration_validation_and_member_owner_options(self):
        self.add_member("zhangsan", "张三")
        self.add_member("lisi", "李四")

        status, data = self.request(
            "PUT", "/api/statuses", {
                "statuses": ["待办", "处理中", "待办"],
                "colors": {"待办": "#AABBCC", "处理中": "#123456"},
            }
        )
        self.assertEqual(status, 200)
        self.assertEqual(data["statuses"], ["待办", "处理中"])
        self.assertEqual(data["colors"], {"待办": "#aabbcc", "处理中": "#123456"})
        status_config = self.request("GET", "/api/statuses")[1]
        self.assertEqual(status_config["statuses"], ["待办", "处理中"])
        self.assertEqual(status_config["colors"]["待办"], "#aabbcc")
        state = self.request("GET", "/api/state")[1]
        self.assertEqual(state["status_colors"]["处理中"], "#123456")
        self.assertEqual(state["owners"], ["系统管理员", "张三", "李四"])
        self.assertEqual(self.request("GET", "/api/owners")[0], 404)
        self.assertEqual(
            self.request("PUT", "/api/owners", {"owners": ["张三"]})[0], 404
        )

        status, data = self.request("PUT", "/api/statuses", {"statuses": []})
        self.assertEqual(status, 400)
        self.assertIn("不能为空", data["error"])

        status, data = self.request("PUT", "/api/statuses", {
            "statuses": ["待办"], "colors": {"待办": "red"},
        })
        self.assertEqual(status, 400)
        self.assertIn("#RRGGBB", data["error"])

    def test_milestone_crud_acceptance_and_task_delete_guard(self):
        line_id = self.create_line()
        first_task = self.create_task(line_id, "完成方案")
        second_task = self.create_task(line_id, "通过评审", status="已闭环")
        milestone_id = self.create_milestone(
            line_id, [first_task, second_task], name="方案冻结"
        )

        status, state = self.request("GET", "/api/state")
        self.assertEqual(status, 200, state)
        self.assertEqual(state["milestones"], [{
            "id": milestone_id,
            "line_id": line_id,
            "name": "方案冻结",
            "target_description": "确认阶段目标已经达成",
            "milestone_date": (self.today + timedelta(days=5)).isoformat(),
            "updated_at": self.today.isoformat(),
            "acceptance_task_ids": [first_task, second_task],
        }])

        status, data = self.request("DELETE", f"/api/tasks/{first_task}")
        self.assertEqual(status, 409, data)
        self.assertIn("验收条件", data["error"])
        new_date = (self.today + timedelta(days=9)).isoformat()
        status, data = self.request(
            "PATCH", f"/api/milestones/{milestone_id}", {
                "name": "方案验收",
                "target_description": "验收全部设计材料",
                "milestone_date": new_date,
                "acceptance_task_ids": [second_task],
            },
        )
        self.assertEqual(status, 200, data)
        status, _ = self.request("DELETE", f"/api/tasks/{first_task}")
        self.assertEqual(status, 200)

        status, data = self.request("DELETE", f"/api/milestones/{milestone_id}")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["milestones"], [])
        status, data = self.request("POST", "/api/undo")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["milestones"][0]["name"], "方案验收")
        self.assertEqual(state["milestones"][0]["acceptance_task_ids"], [second_task])

    def test_milestone_validation_archive_and_canvas_ui(self):
        line_id = self.create_line()
        task_id = self.create_task(line_id)
        yesterday = (self.today - timedelta(days=1)).isoformat()
        base = {
            "line_id": line_id,
            "name": "节点",
            "target_description": "目标",
            "milestone_date": yesterday,
            "acceptance_task_ids": [task_id],
        }
        status, _ = self.request("POST", "/api/milestones", base)
        self.assertEqual(status, 400)
        for invalid_ids in ([task_id, task_id], [999999], [str(task_id)]):
            payload = {**base, "milestone_date": self.today.isoformat(),
                       "acceptance_task_ids": invalid_ids}
            status, _ = self.request("POST", "/api/milestones", payload)
            self.assertIn(status, {400, 404})
        status, _ = self.request("POST", "/api/milestones", {
            **base, "milestone_date": "2026/09/03",
        })
        self.assertEqual(status, 400)

        milestone_id = self.create_milestone(line_id, [task_id])
        status, _ = self.request(
            "PATCH", f"/api/lines/{line_id}", {
                "fork_date": (self.today + timedelta(days=6)).isoformat(),
            },
        )
        self.assertEqual(status, 400)
        workspace_id = self.request(
            "GET", "/api/auth/session"
        )[1]["current_workspace"]["id"]
        self.request("POST", f"/api/workspaces/{workspace_id}/archive")
        status, _ = self.request(
            "PATCH", f"/api/milestones/{milestone_id}", {"name": "不可修改"}
        )
        self.assertEqual(status, 409)

        status, page = self.request("GET", "/")
        self.assertEqual(status, 200)
        self.assertIn(b'id="canvas-context-menu"', page)
        self.assertIn(b'id="context-add-milestone"', page)
        _, script = self.request("GET", "/static/app.js")
        source = script.decode("utf-8")
        self.assertIn('wrap.addEventListener("contextmenu"', source)
        self.assertIn("function openCanvasContextMenu(", source)
        self.assertIn("function openMilestoneModal(", source)
        self.assertIn("function milestoneModalDraft(body)", source)
        self.assertIn("Number(selected.has(b.id)) - Number(selected.has(a.id))", source)
        self.assertIn("if (value && !names.includes(value)) names.push(value)", source)
        self.assertIn("const owner = ownerInput(candidate.owner, true)", source)
        self.assertIn('owner.className = "milestone-acceptance-owner"', source)
        self.assertIn('option.addEventListener("dblclick"', source)
        self.assertIn('event.target.closest?.(".milestone-acceptance-checkbox")', source)
        self.assertIn("onClosed: () => {", source)
        self.assertIn("openMilestoneModal(currentMilestone, line.id, savedDraft)", source)
        self.assertIn("function fivePointStarPoints(", source)
        self.assertIn("function milestoneStatusBands(tasks)", source)
        self.assertIn("28 * band.ratio", source)
        self.assertIn("rect.style.fill = statusColor(band.status)", source)
        self.assertIn('"clip-path": `url(#${clipId})`', source)
        self.assertIn('class: "milestone-node"', source)
        _, styles = self.request("GET", "/static/style.css")
        style_source = styles.decode("utf-8")
        self.assertIn(".milestone-node {", style_source)
        self.assertIn("fill: none; stroke: #d4a72c", style_source)

    def test_line_delete_restores_its_milestones_with_same_batch(self):
        main_id = self.create_line()
        branch_id = self.create_line("支线", parent_id=main_id)
        task_id = self.create_task(branch_id)
        milestone_id = self.create_milestone(branch_id, [task_id])
        status, data = self.request("DELETE", f"/api/lines/{main_id}")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["milestones"], [])
        _, trash = self.request("GET", "/api/trash")
        batch = trash["batches"][0]
        self.assertEqual(batch["milestone_count"], 1)
        self.assertIn(milestone_id, [item["id"] for item in trash["milestones"]])
        status, data = self.request(
            "POST", "/api/trash/restore", {"batch": batch["batch"]}
        )
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["milestones"][0]["id"], milestone_id)
        self.assertEqual(state["milestones"][0]["acceptance_task_ids"], [task_id])

    def test_line_crud_and_date_rules(self):
        yesterday = (self.today - timedelta(days=1)).isoformat()
        tomorrow = (self.today + timedelta(days=1)).isoformat()
        main_id = self.create_line(description="主线描述", color="#123ABC")

        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["lines"][0]["description"], "主线描述")
        self.assertEqual(state["lines"][0]["color"], "#123abc")
        status, data = self.request(
            "PATCH", f"/api/lines/{main_id}", {
                "description": "更新后的描述", "color": "#abcdef",
            }
        )
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["lines"][0]["description"], "更新后的描述")
        self.assertEqual(state["lines"][0]["color"], "#abcdef")
        status, data = self.request(
            "PATCH", f"/api/lines/{main_id}", {"color": "red"}
        )
        self.assertEqual(status, 400, data)

        status, data = self.request("POST", "/api/lines", {
            "name": "非法支线", "parent_id": main_id, "fork_date": yesterday,
        })
        self.assertEqual(status, 400)
        self.assertIn("不能早于", data["error"])

        branch_id = self.create_line("支线", tomorrow, main_id)
        status, _ = self.request(
            "PATCH", f"/api/lines/{branch_id}", {"merge_date": self.today.isoformat()}
        )
        self.assertEqual(status, 400)

        status, data = self.request(
            "PATCH", f"/api/lines/{main_id}", {"fork_date": tomorrow}
        )
        self.assertEqual(status, 200, data)

        day_after = (self.today + timedelta(days=2)).isoformat()
        status, data = self.request(
            "PATCH", f"/api/lines/{main_id}", {"fork_date": day_after}
        )
        self.assertEqual(status, 400)
        self.assertIn("子支线", data["error"])

        status, data = self.request(
            "PATCH", f"/api/lines/{branch_id}", {"merge_date": day_after}
        )
        self.assertEqual(status, 200, data)

    def test_task_full_crud_from_canvas_payload(self):
        line_id = self.create_line()
        task_id = self.create_task(line_id)

        status, state = self.request("GET", "/api/state")
        self.assertEqual(status, 200)
        task = state["tasks"][0]
        self.assertEqual(task["id"], task_id)
        self.assertEqual(task["priority"], "高")
        self.assertEqual(task["next_action"], "下一步")

        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "end_date": None,
        })
        self.assertEqual(status, 400, data)
        self.assertIn("结束日期不能为空", data["error"])

        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "start_date": "",
        })
        self.assertEqual(status, 400, data)
        self.assertIn("起始日期不能为空", data["error"])

        with sqlite3.connect(anyline.app.config["DATABASE"]) as db:
            db.execute("UPDATE tasks SET end_date=NULL WHERE id=?", (task_id,))
        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "name": "历史事务也必须补齐日期",
        })
        self.assertEqual(status, 400, data)
        self.assertIn("结束日期不能为空", data["error"])

        updated_end = (self.today + timedelta(days=10)).isoformat()
        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "name": "更新事务", "status": "已闭环", "end_date": updated_end,
        })
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["tasks"][0]["name"], "更新事务")
        self.assertEqual(state["tasks"][0]["status"], "已闭环")
        self.assertEqual(state["tasks"][0]["end_date"], updated_end)

        status, data = self.request("DELETE", f"/api/tasks/{task_id}")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["tasks"], [])
        self.assertTrue(state["can_undo"])

        status, data = self.request("POST", "/api/undo")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual([task["id"] for task in state["tasks"]], [task_id])

    def test_task_dependencies_and_closure_guard(self):
        line_id = self.create_line()
        prerequisite_one = self.create_task(line_id, "前置事务一")
        prerequisite_two = self.create_task(
            line_id, "前置事务二", status="已闭环"
        )
        dependent_one = self.create_task(
            line_id, "依赖事务一",
            prerequisite_ids=[prerequisite_one, prerequisite_two],
        )
        dependent_two = self.create_task(line_id, "依赖事务二")

        status, data = self.request(
            "POST", f"/api/tasks/{dependent_two}/dependencies",
            {"prerequisite_task_id": prerequisite_one},
        )
        self.assertEqual(status, 201, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(
            {
                (row["dependent_task_id"], row["prerequisite_task_id"])
                for row in state["dependencies"]
            },
            {
                (dependent_one, prerequisite_one),
                (dependent_one, prerequisite_two),
                (dependent_two, prerequisite_one),
            },
        )

        status, data = self.request(
            "PATCH", f"/api/tasks/{dependent_one}", {"status": "已闭环"}
        )
        self.assertEqual(status, 409, data)
        self.assertIn("前置事务一", data["error"])
        status, data = self.request(
            "PATCH", "/api/tasks/bulk",
            {"ids": [dependent_one, dependent_two], "patch": {"status": "已闭环"}},
        )
        self.assertEqual(status, 409, data)

        status, data = self.request(
            "PATCH", f"/api/tasks/{prerequisite_one}", {"status": "已闭环"}
        )
        self.assertEqual(status, 200, data)
        status, data = self.request(
            "PATCH", f"/api/tasks/{dependent_one}", {"status": "已闭环"}
        )
        self.assertEqual(status, 200, data)

        status, data = self.request("DELETE", f"/api/tasks/{prerequisite_one}")
        self.assertEqual(status, 409, data)
        status, data = self.request(
            "DELETE", f"/api/tasks/{dependent_two}/dependencies",
            {"prerequisite_task_id": prerequisite_one},
        )
        self.assertEqual(status, 200, data)

        cycle_one = self.create_task(line_id, "环路一")
        cycle_two = self.create_task(line_id, "环路二")
        status, data = self.request(
            "PATCH", f"/api/tasks/{cycle_one}",
            {"prerequisite_ids": [cycle_two]},
        )
        self.assertEqual(status, 200, data)
        status, data = self.request(
            "PATCH", f"/api/tasks/{cycle_two}",
            {"prerequisite_ids": [cycle_one]},
        )
        self.assertEqual(status, 400, data)
        self.assertIn("循环", data["error"])
        status, data = self.request(
            "PATCH", f"/api/tasks/{cycle_one}",
            {"prerequisite_ids": [cycle_one]},
        )
        self.assertEqual(status, 400, data)

        cycle_three = self.create_task(line_id, "环路三")
        status, data = self.request(
            "PATCH", f"/api/tasks/{cycle_two}",
            {"prerequisite_ids": [cycle_three]},
        )
        self.assertEqual(status, 200, data)
        status, data = self.request(
            "POST", f"/api/tasks/{cycle_three}/dependencies",
            {"prerequisite_task_id": cycle_one},
        )
        self.assertEqual(status, 400, data)
        self.assertIn("循环", data["error"])
        _, state = self.request("GET", "/api/state")
        self.assertNotIn(
            (cycle_three, cycle_one),
            {
                (row["dependent_task_id"], row["prerequisite_task_id"])
                for row in state["dependencies"]
            },
        )

        undo_source = self.create_task(line_id, "撤销来源")
        undo_target = self.create_task(line_id, "撤销目标")
        status, data = self.request(
            "POST", f"/api/tasks/{undo_source}/dependencies",
            {"prerequisite_task_id": undo_target},
        )
        self.assertEqual(status, 201, data)
        status, data = self.request("POST", "/api/undo")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertNotIn(
            (undo_source, undo_target),
            {
                (row["dependent_task_id"], row["prerequisite_task_id"])
                for row in state["dependencies"]
            },
        )

    def test_dashboard_snapshot_metrics_and_same_day_update(self):
        line_id = self.create_line(
            fork_date=(self.today - timedelta(days=2)).isoformat()
        )
        prerequisite = self.create_task(
            line_id, "风险前置",
            status="有风险",
            start_date=(self.today - timedelta(days=2)).isoformat(),
            end_date=(self.today - timedelta(days=1)).isoformat(),
        )
        self.create_task(
            line_id, "被阻塞事务", prerequisite_ids=[prerequisite]
        )
        self.create_task(line_id, "已完成事务", status="已闭环")
        self.create_task(line_id, "已取消事务", status="已取消")

        status, state = self.request("GET", "/api/state")
        self.assertEqual(status, 200, state)
        self.assertEqual(len(state["dashboard_snapshots"]), 1)
        snapshot = state["dashboard_snapshots"][0]
        self.assertEqual(snapshot["snapshot_date"], self.today.isoformat())
        self.assertEqual(
            {key: snapshot[key] for key in ("total", "done", "overdue", "risk", "blocked")},
            {"total": 4, "done": 2, "overdue": 1, "risk": 1, "blocked": 1},
        )
        self.assertEqual(snapshot["status_counts"]["有风险"], 1)
        self.assertEqual(snapshot["status_counts"]["已闭环"], 1)
        self.assertEqual(snapshot["status_counts"]["已取消"], 1)

        status, data = self.request(
            "PATCH", f"/api/tasks/{prerequisite}", {"status": "已闭环"}
        )
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["dashboard_snapshots"]), 1)
        snapshot = state["dashboard_snapshots"][0]
        self.assertEqual(snapshot["done"], 3)
        self.assertEqual(snapshot["overdue"], 0)
        self.assertEqual(snapshot["risk"], 0)
        self.assertEqual(snapshot["blocked"], 0)

    def test_dashboard_risk_bubbles_use_stable_click_targets(self):
        status, script = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = script.decode("utf-8")
        self.assertIn('class: "dashboard-risk-hit-area"', source)
        self.assertIn("group.onclick = open", source)
        self.assertNotIn("group.onpointerenter", source)

    def test_general_undo_for_canvas_edits(self):
        main_id = self.create_line("初始主线")
        status, data = self.request("POST", "/api/undo")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["lines"], [])
        self.assertFalse(state["can_undo"])

        main_id = self.create_line("初始主线")
        status, data = self.request(
            "PATCH", f"/api/lines/{main_id}", {"name": "修改后的主线"}
        )
        self.assertEqual(status, 200, data)
        self.request("POST", "/api/undo")
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["lines"][0]["name"], "初始主线")

        task_id = self.create_task(main_id, "初始事务")
        self.request("POST", "/api/undo")
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["tasks"], [])

        task_id = self.create_task(main_id, "初始事务")
        status, data = self.request(
            "PATCH", f"/api/tasks/{task_id}",
            {"name": "修改后的事务", "status": "已闭环"},
        )
        self.assertEqual(status, 200, data)
        self.request("POST", "/api/undo")
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["tasks"][0]["name"], "初始事务")
        self.assertEqual(state["tasks"][0]["status"], "进行中")

        branch_id = self.create_line("待反合支线", parent_id=main_id)
        merge_date = (self.today + timedelta(days=2)).isoformat()
        status, data = self.request(
            "PATCH", f"/api/lines/{branch_id}", {"merge_date": merge_date}
        )
        self.assertEqual(status, 200, data)
        self.request("POST", "/api/undo")
        _, state = self.request("GET", "/api/state")
        branch = next(line for line in state["lines"] if line["id"] == branch_id)
        self.assertIsNone(branch["merge_date"])

        status, data = self.request("POST", "/api/undo")
        self.assertEqual(status, 400, data)
        self.assertIn("没有可撤销", data["error"])

    def test_redo_restores_undo_and_is_cleared_by_a_new_edit(self):
        main_id = self.create_line("待撤销主线")

        status, data = self.request("POST", "/api/undo")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["lines"], [])
        self.assertFalse(state["can_undo"])
        self.assertTrue(state["can_redo"])

        status, data = self.request("POST", "/api/redo")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual([line["id"] for line in state["lines"]], [main_id])
        self.assertTrue(state["can_undo"])
        self.assertFalse(state["can_redo"])

        status, data = self.request("POST", "/api/redo")
        self.assertEqual(status, 400, data)
        self.assertIn("没有可恢复", data["error"])

        status, data = self.request("POST", "/api/undo")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertTrue(state["can_redo"])

        replacement_id = self.create_line("新的主线")
        _, state = self.request("GET", "/api/state")
        self.assertEqual([line["id"] for line in state["lines"]], [replacement_id])
        self.assertTrue(state["can_undo"])
        self.assertFalse(state["can_redo"])

        status, data = self.request("POST", "/api/redo")
        self.assertEqual(status, 400, data)
        self.assertIn("没有可恢复", data["error"])

    def test_task_validation_never_returns_500(self):
        line_id = self.create_line()
        cases = [
            ({"line_id": line_id, "name": "", "start_date": self.today.isoformat()}, 400),
            ({"line_id": "bad", "name": "事务", "start_date": self.today.isoformat()}, 400),
            ({"line_id": line_id, "name": "事务", "content": [], "start_date": self.today.isoformat()}, 400),
            ({"line_id": line_id, "name": "事务", "priority": "最高", "start_date": self.today.isoformat()}, 400),
            ({"line_id": line_id, "name": "事务", "status": "不存在", "start_date": self.today.isoformat()}, 400),
            ({"line_id": line_id, "name": "事务", "start_date": "2026-02-30"}, 400),
            ({"line_id": line_id, "name": "事务", "start_date": "20260202"}, 400),
            ({"line_id": line_id, "name": "事务", "start_date": "2026-02-02", "end_date": "2026-02-01"}, 400),
        ]
        for payload, expected in cases:
            with self.subTest(payload=payload):
                status, data = self.request("POST", "/api/tasks", payload)
                self.assertEqual(status, expected, data)
                self.assertIn("error", data)

        status, data = self.request("POST", "/api/tasks", raw_body=b"[]")
        self.assertEqual(status, 400)
        self.assertIn("JSON", data["error"])

    def test_task_required_fields_and_modal_controls(self):
        line_id = self.create_line()
        valid = {
            "line_id": line_id,
            "name": "必填校验事务",
            "content": "内容",
            "owner": "系统管理员",
            "status": "进行中",
            "start_date": self.today.isoformat(),
            "end_date": (self.today + timedelta(days=1)).isoformat(),
        }
        labels = {
            "name": "事务名", "content": "事务内容", "owner": "责任人",
            "status": "进展状态", "start_date": "起始日期", "end_date": "结束日期",
        }
        for key, label in labels.items():
            payload = dict(valid)
            payload[key] = ""
            status, data = self.request("POST", "/api/tasks", payload)
            self.assertEqual(status, 400, (key, data))
            self.assertIn(f"{label}不能为空", data["error"])

        invalid_owner = dict(valid)
        invalid_owner["owner"] = "非空间成员"
        status, data = self.request("POST", "/api/tasks", invalid_owner)
        self.assertEqual(status, 400, data)
        self.assertIn("当前项目空间成员", data["error"])

        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")
        self.assertIn('search.placeholder = "搜索事务名、内容、责任人或所属线"', source)
        self.assertIn('field(parent, "依赖事务（可多选）", wrapper)', source)
        self.assertIn('moreSummary.textContent = "更多描述"', source)
        self.assertIn('mark.className = "required-mark"', source)
        self.assertIn(
            'task ? task.start_date : (draft?.startDate || initialStart)', source
        )
        self.assertIn(
            '(draft?.endDate || initialStart)), true)', source
        )
        self.assertIn(
            '["起始日期", body._start], ["结束日期", body._end]', source
        )
        self.assertIn('$("#modal-header-tools").appendChild(del)', source)
        self.assertIn('$("#modal-mask").onclick = (event) => {', source)
        self.assertIn('event.target === event.currentTarget', source)
        self.assertIn('mask._onBackdropClose = options.onBackdropClose || null', source)
        self.assertIn('if (onBackdropClose) onBackdropClose();', source)
        self.assertIn('function saveTaskCreateDraft(', source)
        self.assertIn('function discardTaskCreateDraft(', source)
        self.assertIn(
            'state.taskCreateDrafts.get(openingDraftKey)', source
        )
        self.assertIn('createTaskContentEditor(body, task, draft)', source)
        self.assertIn('draft?.prerequisiteIds || []', source)
        self.assertIn('const imageReadPromises = body._imageReadPromises || []', source)
        self.assertIn('Promise.all(imageReadPromises).then(renderImages)', source)
        self.assertIn('onBackdropClose: () => saveTaskCreateDraft(', source)
        self.assertIn('onCancel: () => discardTaskCreateDraft(', source)
        self.assertIn('state.taskCreateDrafts.clear();', source)

        task_id = self.create_task(line_id)
        status, data = self.request(
            "PATCH", "/api/tasks/bulk",
            {"ids": [task_id], "patch": {"owner": " "}},
        )
        self.assertEqual(status, 400, data)
        self.assertIn("责任人不能为空", data["error"])

    def test_table_line_dropdown_only_uses_main_and_branch_lines(self):
        main_id = self.create_line("产品主线")
        branch_id = self.create_line("交付支线", parent_id=main_id)
        self.create_task(main_id, "不应出现在所属线下拉中的事务")

        status, state = self.request("GET", "/api/state")
        self.assertEqual(status, 200)
        self.assertEqual(
            {line["name"] for line in state["lines"]}, {"产品主线", "交付支线"}
        )
        self.assertNotIn(
            state["tasks"][0]["name"], {line["name"] for line in state["lines"]}
        )

        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8").replace("\r\n", "\n")
        helper_start = source.index("function tableLineOptions()")
        helper_end = source.index("\n}\n", helper_start) + 3
        helper_source = source[helper_start:helper_end]
        self.assertIn("state.lines", helper_source)
        self.assertNotIn("state.tasks", helper_source)
        self.assertIn("line.parent_id === null", helper_source)
        self.assertIn("candidateIds.has(line.parent_id)", helper_source)
        self.assertIn("const lineOptions = tableLineOptions()", source)
        self.assertIn("for (const l of lineOptions)", source)
        self.assertIn("o.textContent = lineOptionLabel(l, lineOptions)", source)

    def test_table_add_task_builds_line_labels_without_map_callback_arguments(self):
        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")
        self.assertIn(
            "lines.map((line) => lineOptionLabel(line, lines))", source
        )
        self.assertNotIn("lines.map(lineOptionLabel)", source)
        self.assertIn('$("#btn-table-add").onclick', source)
        self.assertIn("openTaskModal(null, lineId, true)", source)

    def test_task_content_images_are_persisted_served_and_undoable(self):
        line_id = self.create_line()
        png_base64 = (
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk"
            "+A8AAQUBAScY42YAAAAASUVORK5CYII="
        )
        data_url = f"data:image/png;base64,{png_base64}"
        task_id = self.create_task(line_id, images=[{"data_url": data_url}])

        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["task_images"]), 1)
        image_id = state["task_images"][0]["id"]
        self.assertEqual(state["task_images"][0]["task_id"], task_id)
        status, content = self.request("GET", f"/api/task-images/{image_id}")
        self.assertEqual(status, 200)
        self.assertEqual(content, base64.b64decode(png_base64))

        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "images": [{"id": image_id}, {"data_url": data_url}],
        })
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["task_images"]), 2)

        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "images": [{"id": image_id}],
        })
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["task_images"]), 1)
        self.request("POST", "/api/undo")
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["task_images"]), 2)

        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "images": [{"data_url": "data:image/svg+xml;base64,PHN2Zy8+"}],
        })
        self.assertEqual(status, 400, data)
        self.assertIn("仅支持", data["error"])

        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")
        self.assertIn('textarea.addEventListener("paste"', source)
        self.assertIn('reader.readAsDataURL(file)', source)
        self.assertIn('preview.src = image.src || image.data_url', source)
        self.assertIn(
            'openTaskImageViewer(contentImages, index, previewButton)', source
        )
        self.assertIn('if (e.key === "ArrowRight") moveTaskImageViewer(1)', source)

        status, body = self.request("GET", "/static/style.css")
        self.assertEqual(status, 200)
        styles = body.decode("utf-8")
        lightbox_figure_start = styles.index(".image-lightbox-figure {")
        lightbox_figure_end = styles.index("}", lightbox_figure_start)
        self.assertIn("grid-column: 2", styles[
            lightbox_figure_start:lightbox_figure_end
        ])

    def test_task_attachments_are_downloadable_editable_and_undoable(self):
        line_id = self.create_line()
        first_content = "项目附件内容".encode("utf-8")
        first_data_url = (
            "data:text/plain;base64," +
            base64.b64encode(first_content).decode("ascii")
        )
        task_id = self.create_task(line_id, attachments=[{
            "name": "说明文档.txt", "data_url": first_data_url,
        }])

        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["task_attachments"]), 1)
        attachment = state["task_attachments"][0]
        self.assertEqual(attachment["task_id"], task_id)
        self.assertEqual(attachment["filename"], "说明文档.txt")
        self.assertEqual(attachment["mime_type"], "text/plain")
        self.assertEqual(attachment["size"], len(first_content))
        attachment_id = attachment["id"]

        status, content = self.request(
            "GET", f"/api/task-attachments/{attachment_id}"
        )
        self.assertEqual(status, 200)
        self.assertEqual(content, first_content)

        second_content = b"spreadsheet-bytes"
        second_data_url = (
            "data:application/octet-stream;base64," +
            base64.b64encode(second_content).decode("ascii")
        )
        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "attachments": [
                {"id": attachment_id},
                {"name": "数据表.xlsx", "data_url": second_data_url},
            ],
        })
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["task_attachments"]), 2)

        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "attachments": [{"id": attachment_id}],
        })
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["task_attachments"]), 1)
        self.request("POST", "/api/undo")
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["task_attachments"]), 2)

        status, data = self.request("PATCH", f"/api/tasks/{task_id}", {
            "attachments": [{"name": "空.txt", "data_url": "data:text/plain;base64,"}],
        })
        self.assertEqual(status, 400, data)
        self.assertIn("不能为空", data["error"])

        status, body = self.request("GET", "/static/app.js")
        self.assertEqual(status, 200)
        source = body.decode("utf-8")
        self.assertIn("function autoResizeTaskContent(textarea)", source)
        self.assertIn('textarea.addEventListener("input"', source)
        self.assertIn('editor.addEventListener("dragover"', source)
        self.assertIn('editor.addEventListener("drop"', source)
        self.assertIn("/api/task-attachments/${attachment.id}", source)
        self.assertIn("...(body._attachmentReadPromises || [])", source)
        self.assertIn("await reload();", source)

    def test_bulk_update_delete_and_undo(self):
        self.add_member("lisi", "李四")
        first_line = self.create_line("第一条线")
        second_line = self.create_line("第二条线")
        first = self.create_task(first_line, "事务一")
        second = self.create_task(first_line, "事务二")

        status, data = self.request(
            "PATCH", "/api/tasks/bulk", {"ids": [first, first], "patch": {"owner": "李四"}}
        )
        self.assertEqual(status, 400, data)

        status, data = self.request("PATCH", "/api/tasks/bulk", {
            "ids": [first, second],
            "patch": {"line_id": second_line, "owner": "李四", "priority": "紧急", "status": "有风险"},
        })
        self.assertEqual(status, 200, data)
        self.assertEqual(data["count"], 2)
        _, state = self.request("GET", "/api/state")
        self.assertTrue(all(task["line_id"] == second_line for task in state["tasks"]))
        self.assertTrue(all(task["status"] == "有风险" for task in state["tasks"]))

        status, data = self.request("DELETE", "/api/tasks/bulk", {"ids": [first, second]})
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["tasks"], [])
        self.request("POST", "/api/undo")
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["tasks"]), 2)

    def test_excel_export_all_and_selected(self):
        self.add_member("lisi", "李四")
        main_id = self.create_line("产品主线")
        branch_id = self.create_line("交付支线", parent_id=main_id)
        first = self.create_task(main_id, "=1+1")
        second = self.create_task(branch_id, "交付事务", owner="李四")

        status, content = self.request(
            "POST", "/api/tasks/export", {"scope": "all", "ids": None}
        )
        self.assertEqual(status, 200)
        self.assertTrue(content.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook["事务"]
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, "A1:P3")
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            [column[0] for column in anyline.TASK_EXPORT_COLUMNS],
        )
        self.assertEqual(sheet["E2"].value, "=1+1")
        self.assertEqual(sheet["E2"].data_type, "s")
        self.assertEqual(sheet["C2"].value, "主线")
        self.assertEqual(sheet["C3"].value, "支线")
        self.assertEqual(sheet["D3"].value, "产品主线")

        status, content = self.request(
            "POST", "/api/tasks/export", {"scope": "selected", "ids": [second]}
        )
        self.assertEqual(status, 200)
        sheet = load_workbook(io.BytesIO(content))["事务"]
        self.assertEqual(sheet.auto_filter.ref, "A1:P2")
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["A2"].value, second)
        self.assertEqual(sheet["E2"].value, "交付事务")

        for payload, expected in (
            ({"scope": "selected", "ids": []}, 400),
            ({"scope": "selected", "ids": [first, first]}, 400),
            ({"scope": "selected", "ids": [999999]}, 404),
            ({"scope": "unknown"}, 400),
        ):
            with self.subTest(payload=payload):
                status, data = self.request("POST", "/api/tasks/export", payload)
                self.assertEqual(status, expected, data)

    def test_line_excel_template_export_import_and_undo(self):
        status, content = self.request("GET", "/api/lines/import-template")
        self.assertEqual(status, 200)
        workbook = load_workbook(io.BytesIO(content))
        self.assertIn("线导入", workbook.sheetnames)
        self.assertIn("填报说明", workbook.sheetnames)
        self.assertEqual(
            [cell.value for cell in workbook["线导入"][1]],
            [column[0] for column in anyline.LINE_IMPORT_COLUMNS],
        )
        workbook.close()

        import_book = Workbook()
        sheet = import_book.active
        sheet.title = "线导入"
        sheet.append([column[0] for column in anyline.LINE_IMPORT_COLUMNS])
        # Child rows may appear before their parents; import order is resolved by key.
        sheet.append([
            "delivery", "product", "交付支线", "交付工作", "#123ABC",
            date(2026, 8, 10), date(2026, 8, 28),
        ])
        sheet.append([
            "acceptance", "delivery", "验收支线", "", "",
            date(2026, 8, 12), None,
        ])
        sheet.append([
            "product", "", "=产品主线", "产品规划", "#ABCDEF",
            date(2026, 8, 1), None,
        ])
        output = io.BytesIO()
        import_book.save(output)
        status, data = self.upload_xlsx(
            output.getvalue(), "lines.xlsx", "/api/lines/import"
        )
        self.assertEqual(status, 201, data)
        self.assertEqual(data["count"], 3)

        _, state = self.request("GET", "/api/state")
        by_name = {line["name"]: line for line in state["lines"]}
        self.assertIsNone(by_name["=产品主线"]["parent_id"])
        self.assertEqual(
            by_name["交付支线"]["parent_id"], by_name["=产品主线"]["id"]
        )
        self.assertEqual(
            by_name["验收支线"]["parent_id"], by_name["交付支线"]["id"]
        )
        self.assertEqual(by_name["交付支线"]["color"], "#123abc")

        status, content = self.request("GET", "/api/lines/export")
        self.assertEqual(status, 200)
        exported = load_workbook(io.BytesIO(content))
        exported_sheet = exported["线导入"]
        self.assertEqual(exported_sheet.freeze_panes, "A2")
        self.assertEqual(exported_sheet.max_row, 4)
        self.assertEqual(
            [cell.value for cell in exported_sheet[1]],
            [column[0] for column in anyline.LINE_EXPORT_COLUMNS],
        )
        main_name = next(
            cell for cell in exported_sheet["F"] if cell.value == "=产品主线"
        )
        self.assertEqual(main_name.data_type, "s")
        exported.close()

        status, data = self.request("POST", "/api/undo")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["lines"], [])

        status, data = self.upload_xlsx(
            content, "exported-lines.xlsx", "/api/lines/import"
        )
        self.assertEqual(status, 201, data)
        self.assertEqual(data["count"], 3)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["lines"]), 3)

    def test_line_excel_import_is_atomic_on_invalid_hierarchy(self):
        import_book = Workbook()
        sheet = import_book.active
        sheet.title = "线导入"
        sheet.append([column[0] for column in anyline.LINE_IMPORT_COLUMNS])
        sheet.append(["a", "b", "支线 A", "", "", "2026-08-01", ""])
        sheet.append(["b", "a", "支线 B", "", "", "2026-08-01", ""])
        output = io.BytesIO()
        import_book.save(output)
        status, data = self.upload_xlsx(
            output.getvalue(), "invalid-lines.xlsx", "/api/lines/import"
        )
        self.assertEqual(status, 400, data)
        self.assertIn("未导入任何主线或支线", data["error"])
        self.assertTrue(any("循环" in item["message"] for item in data["row_errors"]))
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["lines"], [])

    def test_unified_excel_import_export_and_selected_lineage(self):
        status, content = self.request("GET", "/api/data/import-template")
        self.assertEqual(status, 200)
        workbook = load_workbook(io.BytesIO(content))
        self.assertIn("线导入", workbook.sheetnames)
        self.assertIn("事务导入", workbook.sheetnames)
        line_sheet = workbook["线导入"]
        line_sheet.append([
            "branch", "main", "交付支线", "", "#123456",
            self.today, None,
        ])
        line_sheet.append([
            "main", "", "产品主线", "", "", self.today, None,
        ])
        task_sheet = workbook["事务导入"]
        task_sheet.append([
            "branch", "", "交付事务", "完成交付", "完成", "下一步", "",
            "高", "系统管理员", "进行中", self.today,
            self.today + timedelta(days=7),
        ])
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        status, data = self.upload_xlsx(
            output.getvalue(), "all-data.xlsx", "/api/data/import"
        )
        self.assertEqual(status, 201, data)
        self.assertEqual(data["line_count"], 2)
        self.assertEqual(data["task_count"], 1)
        _, state = self.request("GET", "/api/state")
        by_name = {line["name"]: line for line in state["lines"]}
        self.assertEqual(
            state["tasks"][0]["line_id"], by_name["交付支线"]["id"]
        )

        unrelated_id = self.create_line("无关主线")
        task_id = state["tasks"][0]["id"]
        status, content = self.request(
            "POST", "/api/data/export", {"scope": "selected", "ids": [task_id]}
        )
        self.assertEqual(status, 200)
        selected_book = load_workbook(io.BytesIO(content))
        exported_line_ids = {
            selected_book["线导入"].cell(row, 1).value
            for row in range(2, selected_book["线导入"].max_row + 1)
        }
        self.assertEqual(
            exported_line_ids,
            {by_name["产品主线"]["id"], by_name["交付支线"]["id"]},
        )
        self.assertNotIn(unrelated_id, exported_line_ids)
        self.assertEqual(selected_book["事务导入"].max_row, 2)
        selected_book.close()

        status, content = self.request(
            "POST", "/api/data/export", {"scope": "all", "ids": None}
        )
        self.assertEqual(status, 200)
        exported = load_workbook(io.BytesIO(content))
        self.assertIn("线导入", exported.sheetnames)
        self.assertIn("事务导入", exported.sheetnames)
        self.assertEqual(
            [cell.value for cell in exported["事务导入"][1]],
            [column[0] for column in anyline.DATA_TASK_EXPORT_COLUMNS],
        )
        exported.close()

        second_workspace = self.request(
            "POST", "/api/workspaces", {"name": "回导空间", "description": ""}
        )
        self.assertEqual(second_workspace[0], 201, second_workspace[1])
        status, data = self.upload_xlsx(
            content, "exported-all-data.xlsx", "/api/data/import"
        )
        self.assertEqual(status, 201, data)
        self.assertEqual(data["line_count"], 3)
        self.assertEqual(data["task_count"], 1)
        _, imported_state = self.request("GET", "/api/state")
        imported_lines = {line["id"] for line in imported_state["lines"]}
        self.assertIn(imported_state["tasks"][0]["line_id"], imported_lines)

    def test_unified_excel_import_is_atomic_across_sheets(self):
        workbook = Workbook()
        line_sheet = workbook.active
        line_sheet.title = "线导入"
        line_sheet.append([column[0] for column in anyline.LINE_IMPORT_COLUMNS])
        line_sheet.append([
            "main", "", "不应写入的主线", "", "", self.today, None,
        ])
        task_sheet = workbook.create_sheet("事务导入")
        task_sheet.append([column[0] for column in anyline.TASK_IMPORT_COLUMNS])
        task_sheet.append([
            "main", "", "错误事务", "内容", "", "", "", "中",
            "系统管理员", "不存在的状态", self.today, self.today,
        ])
        output = io.BytesIO()
        workbook.save(output)
        status, data = self.upload_xlsx(
            output.getvalue(), "invalid-all-data.xlsx", "/api/data/import"
        )
        self.assertEqual(status, 400, data)
        self.assertEqual(data["row_errors"][0]["sheet"], "事务导入")
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["lines"], [])
        self.assertEqual(state["tasks"], [])

    def test_excel_import_template_and_atomic_import(self):
        self.add_member("zhangsan", "张三")
        self.add_member("lisi", "李四")
        main_id = self.create_line("产品主线", "2026-08-01")
        branch_id = self.create_line(
            "交付支线", "2026-08-10", parent_id=main_id
        )

        status, content = self.request("GET", "/api/tasks/import-template")
        self.assertEqual(status, 200)
        self.assertTrue(content.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(content))
        self.assertIn("事务导入", workbook.sheetnames)
        self.assertIn("填报说明", workbook.sheetnames)
        self.assertIn("项目数据", workbook.sheetnames)
        template = workbook["事务导入"]
        self.assertEqual(
            [cell.value for cell in template[1]],
            [column[0] for column in anyline.TASK_IMPORT_COLUMNS],
        )
        project_data = workbook["项目数据"]
        self.assertEqual(project_data["A2"].value, main_id)
        self.assertEqual(project_data["B3"].value, "产品主线 / 交付支线")
        owner_options = [
            workbook["选项"].cell(row, 3).value
            for row in range(2, workbook["选项"].max_row + 1)
            if workbook["选项"].cell(row, 3).value
        ]
        self.assertEqual(owner_options, ["系统管理员", "张三", "李四"])
        self.assertGreaterEqual(len(template.data_validations.dataValidation), 2)
        self.assertIn("ImportStatuses", workbook.defined_names)
        self.assertIn("ImportPriorities", workbook.defined_names)
        workbook.close()

        import_book = Workbook()
        sheet = import_book.active
        sheet.title = "事务导入"
        sheet.append([column[0] for column in anyline.TASK_IMPORT_COLUMNS])
        sheet.append([
            main_id, "产品主线", "接口联调", "完成接口联调", "联调通过", "修复问题", "",
            "高", "张三", "进行中", date(2026, 8, 20), date(2026, 9, 5),
        ])
        sheet.append([
            None, "产品主线 / 交付支线", "准备验收", "整理验收材料", "材料齐备", "",
            "等待客户确认", None, "李四", "有风险", "2026-08-12", "2026-08-28",
        ])
        output = io.BytesIO()
        import_book.save(output)
        status, data = self.upload_xlsx(output.getvalue())
        self.assertEqual(status, 201, data)
        self.assertEqual(data["count"], 2)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["tasks"]), 2)

        status, data = self.upload_xlsx(b"not-an-excel-workbook")
        self.assertEqual(status, 400, data)
        self.assertIn("无法读取", data["error"])
        self.assertEqual(state["tasks"][0]["line_id"], branch_id)
        self.assertEqual(state["tasks"][0]["priority"], "中")
        self.assertEqual(state["tasks"][1]["line_id"], main_id)

        invalid_book = Workbook()
        sheet = invalid_book.active
        sheet.title = "事务导入"
        sheet.append([column[0] for column in anyline.TASK_IMPORT_COLUMNS])
        sheet.append([
            main_id, "产品主线", "有效事务", "有效内容", "", "", "", "中", "张三",
            "进行中", "2026-08-20", "2026-08-21",
        ])
        sheet.append([
            branch_id, "产品主线 / 交付支线", "错误事务", "错误内容", "", "", "", "中",
            "李四", "不存在的状态", "2026-08-20", "2026-08-21",
        ])
        output = io.BytesIO()
        invalid_book.save(output)
        status, data = self.upload_xlsx(output.getvalue())
        self.assertEqual(status, 400, data)
        self.assertEqual(data["row_errors"][0]["row"], 3)
        self.assertIn("未导入任何事务", data["error"])
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["tasks"]), 2)

        status, data = self.request("POST", "/api/undo")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["tasks"], [])

    def test_recursive_delete_restore_and_purge(self):
        main_id = self.create_line("主线")
        branch_id = self.create_line("支线", self.today.isoformat(), main_id)
        self.create_task(branch_id)

        status, data = self.request("DELETE", f"/api/lines/{main_id}")
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(state["lines"], [])
        self.assertEqual(state["tasks"], [])

        _, trash = self.request("GET", "/api/trash")
        self.assertEqual(trash["batches"][0]["line_count"], 2)
        self.assertEqual(trash["batches"][0]["task_count"], 1)
        batch = trash["batches"][0]["batch"]
        status, data = self.request("POST", "/api/trash/restore", {"batch": batch})
        self.assertEqual(status, 200, data)
        _, state = self.request("GET", "/api/state")
        self.assertEqual(len(state["lines"]), 2)
        self.assertEqual(len(state["tasks"]), 1)

        self.request("DELETE", f"/api/lines/{main_id}")
        status, data = self.request("POST", "/api/trash/purge")
        self.assertEqual(status, 200, data)
        _, trash = self.request("GET", "/api/trash")
        self.assertEqual(trash["batches"], [])

    def test_restore_dependency_order(self):
        line_id = self.create_line()
        task_id = self.create_task(line_id)
        self.request("DELETE", f"/api/tasks/{task_id}")
        _, trash = self.request("GET", "/api/trash")
        task_batch = trash["batches"][0]["batch"]
        self.request("DELETE", f"/api/lines/{line_id}")
        _, trash = self.request("GET", "/api/trash")
        line_batch = next(batch["batch"] for batch in trash["batches"] if batch["line_count"])

        status, data = self.request("POST", "/api/trash/restore", {"batch": task_batch})
        self.assertEqual(status, 409, data)
        status, _ = self.request("POST", "/api/trash/restore", {"batch": line_batch})
        self.assertEqual(status, 200)
        status, _ = self.request("POST", "/api/trash/restore", {"batch": task_batch})
        self.assertEqual(status, 200)

    def test_collaboration_comments_mentions_followers_and_notifications(self):
        self.add_member("alice", "张三")
        line_id = self.create_line()
        task_id = self.create_task(line_id, owner="张三")

        status, data = self.request(
            "POST", f"/api/tasks/{task_id}/comments",
            {"content": "@张三 请确认今天的处理方案"},
        )
        self.assertEqual(status, 201, data)
        status, collaboration = self.request(
            "GET", f"/api/tasks/{task_id}/collaboration"
        )
        self.assertEqual(status, 200, collaboration)
        self.assertTrue(collaboration["following"])
        self.assertEqual(collaboration["followers"][0]["display_name"], "系统管理员")
        self.assertEqual(collaboration["timeline"][0]["kind"], "comment")
        self.assertIn("请确认", collaboration["timeline"][0]["detail"])
        self.assertTrue(any(
            item["summary"] == "创建了事务"
            for item in collaboration["timeline"] if item["kind"] == "activity"
        ))

        status, data = self.login("alice", "member123")
        self.assertEqual(status, 200, data)
        status, notices = self.request("GET", "/api/notifications")
        self.assertEqual(status, 200, notices)
        kinds = {item["kind"] for item in notices["notifications"]}
        self.assertIn("assigned", kinds)
        self.assertIn("mention", kinds)
        mention = next(
            item for item in notices["notifications"] if item["kind"] == "mention"
        )
        self.assertEqual(mention["task_id"], task_id)
        self.assertEqual(mention["task_available"], 1)

        status, data = self.request("POST", f"/api/tasks/{task_id}/follow")
        self.assertEqual(status, 200, data)
        self.assertTrue(data["following"])
        status, data = self.request(
            "POST", f"/api/notifications/{mention['id']}/read"
        )
        self.assertEqual(status, 200, data)
        status, data = self.request("POST", "/api/notifications/read-all")
        self.assertEqual(status, 200, data)
        self.assertEqual(data["unread_count"], 0)

        status, data = self.request(
            "POST", f"/api/tasks/{task_id}/comments", {"content": "方案已经确认"}
        )
        self.assertEqual(status, 201, data)
        self.login("admin", "admin123")
        _, notices = self.request("GET", "/api/notifications")
        self.assertTrue(any(
            item["kind"] == "comment" and item["task_id"] == task_id
            for item in notices["notifications"]
        ))

    def test_status_activity_and_dependency_unblocked_notification(self):
        self.add_member("alice", "张三")
        line_id = self.create_line()
        prerequisite_id = self.create_task(line_id, "前置事务")
        dependent_id = self.create_task(
            line_id, "后续事务", owner="张三",
            prerequisite_ids=[prerequisite_id],
        )
        self.login("alice", "member123")
        self.request("POST", "/api/notifications/read-all")
        self.login("admin", "admin123")

        status, data = self.request(
            "PATCH", f"/api/tasks/{prerequisite_id}", {"status": "已闭环"}
        )
        self.assertEqual(status, 200, data)
        _, collaboration = self.request(
            "GET", f"/api/tasks/{prerequisite_id}/collaboration"
        )
        self.assertTrue(any(
            "已闭环" in item["summary"]
            for item in collaboration["timeline"] if item["kind"] == "activity"
        ))

        self.login("alice", "member123")
        _, notices = self.request("GET", "/api/notifications")
        unblocked = [
            item for item in notices["notifications"]
            if item["kind"] == "dependency_unblocked"
        ]
        self.assertEqual(len(unblocked), 1)
        self.assertEqual(unblocked[0]["task_id"], dependent_id)
        self.assertIn("可以继续推进", unblocked[0]["message"])

    def test_collaboration_ui_entries_are_present(self):
        _, body = self.request("GET", "/")
        self.assertIn(b'id="btn-my-status"', body)
        self.assertNotIn(b'id="btn-notifications"', body)
        _, source = self.request("GET", "/static/app.js")
        source = source.decode("utf-8")
        self.assertIn("function renderMyNotificationsPanel(", source)
        self.assertIn('notificationTab.textContent = `协作通知（${state.unreadNotifications || 0} 未读）`', source)
        self.assertIn("function createTaskCollaborationPanel(body, task)", source)
        self.assertIn("function createMentionAutocomplete(textarea, members, taskId)", source)
        self.assertIn('textarea.setAttribute("aria-autocomplete", "list")', source)
        self.assertIn('.startsWith(query)', source)
        self.assertIn('event.key === "ArrowDown" || event.key === "ArrowUp"', source)
        self.assertIn('event.key === "Enter" || event.key === "Tab"', source)
        self.assertNotIn("插入 @成员", source)


class DatabaseMigrationTests(unittest.TestCase):
    def test_legacy_schema_is_migrated(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = os.path.join(temp_dir, "legacy.db")
            db = sqlite3.connect(db_path)
            db.executescript(
                """
                CREATE TABLE lines (
                    id INTEGER PRIMARY KEY, name TEXT NOT NULL, parent_id INTEGER,
                    fork_date TEXT NOT NULL, merge_date TEXT, deleted INTEGER DEFAULT 0,
                    del_batch INTEGER
                );
                CREATE TABLE tasks (
                    id INTEGER PRIMARY KEY, line_id INTEGER NOT NULL, name TEXT NOT NULL,
                    content TEXT, goal TEXT, owner TEXT, status TEXT NOT NULL,
                    start_date TEXT NOT NULL, end_date TEXT, status_since TEXT NOT NULL,
                    deleted INTEGER DEFAULT 0, del_batch INTEGER
                );
                CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT);
                INSERT INTO lines(id,name,parent_id,fork_date,deleted)
                VALUES(1,'历史主线',NULL,'2026-01-01',0);
                INSERT INTO tasks(
                    id,line_id,name,status,start_date,status_since,deleted
                ) VALUES(1,1,'历史事务','未启动','2026-01-01','2026-01-01',0);
                INSERT INTO meta(key,value) VALUES('owners','["历史责任人"]');
                """
            )
            db.commit()
            db.close()

            anyline.init_db(db_path)
            db = sqlite3.connect(db_path)
            line_columns = {row[1] for row in db.execute("PRAGMA table_info(lines)")}
            task_columns = {row[1] for row in db.execute("PRAGMA table_info(tasks)")}
            table_names = {
                row[0] for row in db.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                )
            }
            workspace_id = db.execute("SELECT workspace_id FROM lines WHERE id=1").fetchone()[0]
            task_workspace_id = db.execute(
                "SELECT workspace_id FROM tasks WHERE id=1"
            ).fetchone()[0]
            admin_count = db.execute(
                "SELECT COUNT(*) FROM workspace_members WHERE role='admin'"
            ).fetchone()[0]
            migrated_owners = db.execute(
                "SELECT value FROM workspace_meta WHERE workspace_id=? AND key='owners'",
                (workspace_id,),
            ).fetchone()[0]
            db.close()
            self.assertTrue(
                {"description", "color", "deleted_at", "updated_at", "workspace_id"}
                .issubset(line_columns)
            )
            self.assertTrue(
                {
                    "priority", "next_action", "risk_reason", "deleted_at",
                    "updated_at", "workspace_id",
                }
                .issubset(task_columns)
            )
            self.assertIn("task_images", table_names)
            self.assertIn("task_attachments", table_names)
            self.assertIn("milestones", table_names)
            self.assertIn("milestone_tasks", table_names)
            self.assertIn("dashboard_snapshots", table_names)
            self.assertIn("task_followers", table_names)
            self.assertIn("task_comments", table_names)
            self.assertIn("task_activities", table_names)
            self.assertIn("notifications", table_names)
            self.assertEqual(task_workspace_id, workspace_id)
            self.assertEqual(admin_count, 1)
            self.assertEqual(json.loads(migrated_owners), ["历史责任人"])


if __name__ == "__main__":
    unittest.main()
